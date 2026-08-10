import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from email_validator import EmailNotValidError, validate_email
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


COLUMNS = [
    ("employee_name", "Employee Name"), ("personal_email", "Personal Email"),
    ("date_of_birth", "Date of Birth"), ("employee_code", "Employee Code"),
    ("unit", "Unit"), ("unit_address", "Unit Address"), ("designation", "Designation"),
    ("date_of_joining", "Date of Joining"), ("uan", "UAN"), ("esi_number", "ESI Number"),
    ("basic_gross", "Basic Gross"), ("basic_earnings", "Basic Earnings"),
    ("hra_gross", "HRA Gross"), ("hra_earnings", "HRA Earnings"),
    ("special_allowance_gross", "Special Allowance Gross"), ("special_allowance_earnings", "Special Allowance Earnings"),
    ("overtime_gross", "Overtime Gross"), ("overtime_earnings", "Overtime Earnings"),
    ("variable_pay_gross", "Variable Pay Gross"), ("variable_pay_earnings", "Variable Pay Earnings"),
    ("total_gross", "Total Gross"), ("total_earnings", "Total Earnings"),
    ("pf", "PF"), ("esi_deduction", "ESI Deduction"), ("professional_tax", "Professional Tax"),
    ("loan", "Loan"), ("advance", "Advance"), ("other_deductions", "Other Deductions"),
    ("tds", "TDS"), ("deduction_total", "Deduction Total"), ("days", "Days"),
    ("present_days", "Present Days"), ("lop", "LOP"), ("ot_hours", "OT Hours"),
    ("net_wages", "Net Salary"), ("net_wages_words", "Net Salary in Words"),
]
GROSS_FIELDS = ("basic_gross", "hra_gross", "special_allowance_gross", "overtime_gross", "variable_pay_gross")
EARNING_FIELDS = ("basic_earnings", "hra_earnings", "special_allowance_earnings", "overtime_earnings", "variable_pay_earnings")
DEDUCTION_FIELDS = ("pf", "esi_deduction", "professional_tax", "loan", "advance", "other_deductions", "tds")
MONEY_FIELDS = set(GROSS_FIELDS + EARNING_FIELDS + DEDUCTION_FIELDS + ("total_gross", "total_earnings", "deduction_total", "net_wages"))
TEMPLATE_BASE_WIDTH = 595.0
TEMPLATE_BASE_HEIGHT = 835.3682

UNIT_ADDRESSES = {
    "1": "Plot No. B 105/106, Baikampady Industrial Area, Mangalore, Dakshina Kannada, Karnataka - 575011",
    "2": "Ground Floor, Plot No. 42 & 43, Kallur, Mandli Kallur Industrial Area, Shivamogga, Karnataka - 577202",
    "3": "Plot 166A, Baikampady Industrial Road Area, Mangalore, Dakshina Kannada, Karnataka - 575011",
}

HEADER_ALIASES = {
    "name": "employee_name", "employeename": "employee_name",
    "personalemail": "personal_email", "email": "personal_email", "emailid": "personal_email",
    "dob": "date_of_birth", "dateofbirth": "date_of_birth", "birthdate": "date_of_birth", "birthyear": "date_of_birth",
    "employeecode": "employee_code", "empcode": "employee_code", "code": "employee_code",
    "unit": "unit", "unitnumber": "unit", "unitno": "unit", "unitaddress": "unit_address",
    "designation": "designation", "doj": "date_of_joining", "dateofjoining": "date_of_joining",
    "uan": "uan", "uannumber": "uan", "esinumber": "esi_number",
    "days": "days", "totaldays": "days", "presentdays": "present_days", "paiddays": "present_days",
    "lop": "lop", "lossofpay": "lop", "othours": "ot_hours", "overtimehours": "ot_hours",
    "gross": "total_gross", "totalgross": "total_gross",
    "basic": "basic_gross", "basicgross": "basic_gross", "basicearnings": "basic_earnings",
    "hra": "hra_gross", "hragross": "hra_gross", "hraearnings": "hra_earnings",
    "specialallowance": "special_allowance_gross", "specialallowancegross": "special_allowance_gross", "specialallowanceearnings": "special_allowance_earnings",
    "overtime": "overtime_earnings", "overtimegross": "overtime_gross", "overtimeearnings": "overtime_earnings",
    "variablepay": "variable_pay_earnings", "variablepaygross": "variable_pay_gross", "variablepayearnings": "variable_pay_earnings",
    "pf": "pf", "esideduction": "esi_deduction", "ptax": "professional_tax", "professionaltax": "professional_tax",
    "loan": "loan", "tds": "tds", "advance": "advance", "others": "other_deductions", "otherdeductions": "other_deductions",
    "deductiontotal": "deduction_total", "totaldeductions": "deduction_total",
    "net": "net_wages", "netsalary": "net_wages", "netwages": "net_wages", "netsalaryinwords": "net_wages_words",
}


def create_excel_template() -> bytes:
    approved_template = Path(__file__).resolve().parents[2] / "assets" / "payroll" / "AROMAZEN_Salary_Upload_Template.xlsx"
    if approved_template.is_file():
        return approved_template.read_bytes()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Salary Data"
    sheet.append([label for _, label in COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(COLUMNS)).column_letter}1"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    for index, (_, label) in enumerate(COLUMNS, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(14, min(32, len(label) + 3))
    sample = {"employee_name": "Sample Employee", "personal_email": "employee@example.com", "date_of_birth": date(1992, 5, 18), "employee_code": "EMP001", "unit": "Mangalore Unit", "unit_address": "B-105/106, Industrial Area, Baikampady, Mangalore - 575 011", "designation": "Executive", "date_of_joining": date(2022, 1, 10)}
    sheet.append([sample.get(key, "") for key, _ in COLUMNS])
    sheet["C2"].number_format = "DD-MM-YYYY"
    sheet["H2"].number_format = "DD-MM-YYYY"
    note = workbook.create_sheet("Instructions")
    instructions = [
        "AROMAZEN Salary Slip Upload",
        "Use one row per employee. Employee name, personal email, date of birth, employee code, unit, days and present days are mandatory.",
        "Enter Unit 1, 2 or 3. The approved unit address and PDF template are selected automatically.",
        "PDF password: first 4 letters of employee name in uppercase + four-digit birth year.",
        "Only Days, Present Days, LOP and OT Hours are filled by the Employee Leave Calculator.",
        "All salary amounts, totals, deductions, Net Salary and Net Salary in Words are used exactly as uploaded.",
    ]
    for text in instructions:
        note.append([text])
    note.column_dimensions["A"].width = 125
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalise_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _money(value: object) -> Decimal:
    cleaned = re.sub(r"[^0-9.\-]", "", _text(value))
    if not cleaned:
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation as error:
        raise ValueError(f"Invalid amount: {value}") from error


def _date(value: object, field: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    raise ValueError(f"{field} must be a valid date (DD-MM-YYYY).")


def _birth_year(value: object) -> int:
    text = _text(value)
    if re.fullmatch(r"(?:19|20)\d{2}(?:\.0)?", text):
        year = int(float(text))
    else:
        year = _date(value, "Date of Birth").year
    if year < 1900 or year >= date.today().year:
        raise ValueError("Date of Birth is invalid.")
    return year


def _unit_number(value: object, address: object = None) -> str:
    match = re.search(r"(?:unit\s*)?([123])(?:\.0)?$", _text(value), re.IGNORECASE)
    if match:
        return match.group(1)
    location = _normalise_header(f"{_text(value)} {_text(address)}")
    if any(marker in location for marker in ("shivamogga", "kallur", "plotno42", "plot42")):
        return "2"
    if "166a" in location:
        return "3"
    if any(marker in location for marker in ("105106", "b105106", "mangalore", "baikampady")):
        return "1"
    raise ValueError("Unit must identify Unit 1, Unit 2 or Unit 3; a recognized unit name/address is also accepted.")


def _rounded(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _number_text(value: Decimal) -> str:
    value = _rounded(value)
    return str(int(value)) if value == value.to_integral_value() else format(value.normalize(), "f")


def _indian_words(number: int) -> str:
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    def under_hundred(value: int) -> str:
        return ones[value] if value < 20 else " ".join(part for part in (tens[value // 10], ones[value % 10]) if part)
    def under_thousand(value: int) -> str:
        return " ".join(part for part in ((f"{ones[value // 100]} Hundred" if value >= 100 else ""), under_hundred(value % 100)) if part)
    if number == 0:
        return "Zero Rupees Only"
    parts = []
    for divisor, label in ((10_000_000, "Crore"), (100_000, "Lakh"), (1_000, "Thousand")):
        amount, number = divmod(number, divisor)
        if amount:
            parts.append(f"{under_thousand(amount)} {label}")
    if number:
        parts.append(under_thousand(number))
    return " ".join(parts) + " Rupees Only"


def password_for(name: str, birth_year: int) -> str:
    return f"{re.sub(r'[^A-Za-z]', '', name).upper()[:4]}{birth_year}"


def read_salary_excel(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise ValueError("The uploaded file is not a readable .xlsx workbook.") from error
    sheet = workbook["Salary Data"] if "Salary Data" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("The workbook is empty.")
    expected = {_normalise_header(label): key for key, label in COLUMNS}
    expected.update(HEADER_ALIASES)
    indexes: dict[str, int] = {}
    esi_seen = 0
    for index, value in enumerate(headers):
        normalised = _normalise_header(value)
        if normalised == "esi":
            key = "esi_deduction" if "pf" in indexes or esi_seen > 0 else "esi_number"
            esi_seen += 1
        else:
            key = expected.get(normalised)
        if key and key not in indexes:
            indexes[key] = index
    required = ("employee_name", "personal_email", "date_of_birth", "employee_code", "unit", "days", "present_days")
    labels = dict(COLUMNS)
    missing = [labels[key] for key in required if key not in indexes]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))
    parsed, errors = [], []
    for row_number, row in enumerate(rows, 2):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            values = {key: row[index] if index < len(row) else None for key, index in indexes.items()}
            details = {key: _text(values.get(key)) for key, _ in COLUMNS}
            for key in required:
                if not details[key]:
                    raise ValueError(f"{labels[key]} is required.")
            try:
                email = validate_email(details["personal_email"], check_deliverability=False).normalized.lower()
            except EmailNotValidError as error:
                raise ValueError("Personal Email is invalid.") from error
            birth_year = _birth_year(values.get("date_of_birth"))
            unit = _unit_number(values.get("unit"), values.get("unit_address"))
            details.update(personal_email=email, date_of_birth=str(birth_year), unit=unit, unit_address=_text(values.get("unit_address")) or UNIT_ADDRESSES[unit])
            for key in MONEY_FIELDS:
                details[key] = f"{_money(values.get(key)):.2f}"
            days = _money(values.get("days"))
            present_days = _money(values.get("present_days"))
            if days <= 0:
                raise ValueError("Days must be greater than zero.")
            if present_days < 0 or present_days > days:
                raise ValueError("Present Days must be between zero and Days.")
            details["days"] = _number_text(days)
            details["present_days"] = _number_text(present_days)
            lop = _money(values.get("lop")) if _text(values.get("lop")) else days - present_days
            details["lop"] = _number_text(lop)
            details["ot_hours"] = _number_text(_money(values.get("ot_hours")))
            uploaded_net = _money(values.get("net_wages"))
            if uploaded_net < 0:
                raise ValueError("Net Salary cannot be negative.")
            details.update(gross=details["total_gross"], calculation_factor="", uploaded_net_wages=details["net_wages"])
            details["net_wages_words"] = _text(values.get("net_wages_words"))
            if values.get("date_of_joining") not in (None, ""):
                details["date_of_joining"] = _date(values.get("date_of_joining"), "Date of Joining").strftime("%d-%m-%Y")
            parsed.append({"row_number": row_number, "employee_name": details["employee_name"], "employee_code": details["employee_code"], "personal_email": email, "birth_year": birth_year, "details": details})
        except ValueError as error:
            errors.append(f"Row {row_number}: {error}")
    if errors:
        suffix = f" (+{len(errors) - 12} more)" if len(errors) > 12 else ""
        raise ValueError(" | ".join(errors[:12]) + suffix)
    if not parsed:
        raise ValueError("No employee rows were found.")
    if len(parsed) > 500:
        raise ValueError("A payroll batch can contain at most 500 employees.")
    return parsed


def validate_template_pdf(content: bytes) -> None:
    try:
        reader = PdfReader(io.BytesIO(content))
        if len(reader.pages) != 1:
            raise ValueError("The salary-slip template must contain exactly one page.")
        page = reader.pages[0]
        width, height = float(page.mediabox.width), float(page.mediabox.height)
        if width >= height or abs((width / height) - (A4[0] / A4[1])) > 0.025:
            raise ValueError("The template must be an A4 portrait PDF.")
    except ValueError:
        raise
    except Exception as error:
        raise ValueError("The uploaded file is not a readable PDF.") from error


def _amount(details: dict, key: str) -> str:
    try:
        return f"{Decimal(str(details.get(key) or 0)):,.2f}"
    except InvalidOperation:
        return str(details.get(key) or "-")


def _fit_text(pdf: canvas.Canvas, text: str, x: float, y: float, width: float, align: str = "left", size: float = 7.2, bold: bool = False) -> None:
    font = "Helvetica-Bold" if bold else "Helvetica"
    value = str(text or "-")
    while size > 5.2 and pdf.stringWidth(value, font, size) > width - 8:
        size -= .25
    pdf.setFont(font, size)
    if align == "center":
        pdf.drawCentredString(x + width / 2, y, value)
    elif align == "right":
        pdf.drawRightString(x + width - 4, y, value)
    else:
        pdf.drawString(x + 4, y, value)


def _box(pdf: canvas.Canvas, x: float, y: float, width: float, height: float, fill: colors.Color | None = None) -> None:
    if fill:
        pdf.setFillColor(fill)
        pdf.rect(x, y, width, height, stroke=1, fill=1)
        pdf.setFillColor(colors.black)
    else:
        pdf.rect(x, y, width, height, stroke=1, fill=0)


def _bar(pdf: canvas.Canvas, text: str, x: float, y: float, width: float, height: float = 19) -> None:
    _box(pdf, x, y, width, height, colors.HexColor("#10141b"))
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 7.5)
    pdf.drawCentredString(x + width / 2, y + 6, text)
    pdf.setFillColor(colors.black)


def _draw_flower_logo(pdf: canvas.Canvas) -> None:
    cx, cy = 242, 796
    pdf.setLineWidth(1.3)
    pdf.circle(cx, cy, 20, stroke=1, fill=0)
    pdf.circle(cx, cy, 17, stroke=1, fill=0)
    for dx, dy in ((0, 8), (7, 4), (7, -4), (0, -8), (-7, -4), (-7, 4)):
        pdf.circle(cx + dx, cy + dy, 6.8, stroke=1, fill=0)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(270, 799, "AROMAZEN")
    pdf.setFont("Helvetica", 6.5)
    pdf.drawString(299, 788, "PRIVATE LIMITED")


def _draw_default_template(pdf: canvas.Canvas) -> None:
    x, width = 36, 523
    pdf.setLineWidth(.55)
    _draw_flower_logo(pdf)
    _bar(pdf, "SALARY SLIP", x, 727, width)
    _box(pdf, x, 704, width, 22)
    pdf.line(150, 704, 150, 726)
    _fit_text(pdf, "Salary for the Month", x, 712, 114, bold=True)
    column_x = [36, 116, 236, 306, 411, 466, 559]
    for row in range(4):
        pdf.line(x, 644 + row * 17, x + width, 644 + row * 17)
    for value in column_x:
        pdf.line(value, 644, value, 695)
    labels = (("Emp Name", 36, 678), ("Unit", 236, 678), ("UAN", 411, 678), ("Emp Code", 36, 661), ("Designation", 236, 661), ("ESI", 411, 661), ("Date of Joining", 36, 644), ("OT Hours", 236, 644))
    for label, lx, ly in labels:
        _fit_text(pdf, label, lx, ly + 5, 80 if lx == 36 else 70 if lx == 236 else 55, bold=True)
    attendance = ("DAYS", "PRESENT DAYS", "LOP", "OT HOURS")
    cell_width = width / 4
    for index, label in enumerate(attendance):
        _bar(pdf, label, x + index * cell_width, 605, cell_width)
        _box(pdf, x + index * cell_width, 584, cell_width, 21)
    left_x, left_w, right_x, right_w = 36, 254, 305, 254
    _bar(pdf, "EARNINGS", left_x, 548, left_w)
    _bar(pdf, "DEDUCTIONS", right_x, 548, right_w)
    for px, widths, headings in ((left_x, (100, 77, 77), ("COMPONENTS", "GROSS (INR)", "EARNINGS (INR)")), (right_x, (120, 134), ("COMPONENTS", "AMOUNT (INR)"))):
        cursor = px
        for cell, heading in zip(widths, headings):
            _box(pdf, cursor, 528, cell, 20, colors.HexColor("#eceff1"))
            _fit_text(pdf, heading, cursor, 535, cell, "center", 6.2, True)
            cursor += cell
    row_height = 108 / 7
    for row in range(7):
        y = 420 + row * row_height
        for start, widths in ((left_x, (100, 77, 77)), (right_x, (120, 134))):
            cursor = start
            for cell in widths:
                _box(pdf, cursor, y, cell, row_height)
                cursor += cell
    earning_labels = ("Basic", "HRA", "Special Allowance", "Over Time", "Variable Pay")
    deduction_labels = ("PF", "ESI", "P.Tax", "Loan", "Advance", "Others", "TDS")
    for index, label in enumerate(earning_labels):
        _fit_text(pdf, label, left_x, 520 - index * row_height, 100)
    for index, label in enumerate(deduction_labels):
        _fit_text(pdf, label, right_x, 520 - index * row_height, 120)
    _box(pdf, left_x, 400, left_w, 20, colors.HexColor("#eceff1"))
    pdf.line(left_x + 100, 400, left_x + 100, 420); pdf.line(left_x + 177, 400, left_x + 177, 420)
    _fit_text(pdf, "TOTAL GROSS", left_x, 407, 100, bold=True)
    _box(pdf, right_x, 400, right_w, 20, colors.HexColor("#eceff1"))
    pdf.line(right_x + 120, 400, right_x + 120, 420)
    _fit_text(pdf, "TOTAL DEDUCTIONS", right_x, 407, 120, bold=True)
    _bar(pdf, "NET SALARY (INR)", x, 370, 116, 20)
    _box(pdf, x + 116, 370, width - 116, 20)
    _box(pdf, x, 342, width, 20)
    pdf.line(x + 130, 342, x + 130, 362)
    _fit_text(pdf, "NET SALARY (IN WORDS)", x, 349, 130, bold=True)
    pdf.line(x, 326, x + width, 326)
    pdf.setFont("Helvetica", 6.2)
    pdf.drawCentredString(A4[0] / 2, 315, "This is a computer generated salary slip and does not require a signature.")


def _erase_placeholder(pdf: canvas.Canvas, center_x: float, center_y: float, fill: colors.Color, width: float = 28, height: float = 9) -> None:
    pdf.setFillColor(fill)
    pdf.rect(center_x - width / 2, center_y - height / 2, width, height, stroke=0, fill=1)
    pdf.setFillColor(colors.black)


def _draw_values(pdf: canvas.Canvas, details: dict, payroll_month: str, erase: bool, placeholder_fill=None) -> None:
    x, width = 36, 523
    month_label = datetime.strptime(payroll_month, "%Y-%m").strftime("%B %Y")
    if erase:
        for center_x, center_y in ((249, 703), (489, 703), (178, 661), (362, 661), (513, 661), (178, 644), (362, 644), (513, 644)):
            fill = placeholder_fill(center_x, center_y) if placeholder_fill else colors.white
            if fill is not None:
                _erase_placeholder(pdf, center_x, center_y, fill)
        for index in range(4):
            center_x, center_y = 40 + (index + .5) * 514 / 4, 588
            fill = placeholder_fill(center_x, center_y) if placeholder_fill else colors.white
            if fill is not None:
                _erase_placeholder(pdf, center_x, center_y, fill)
        row_height = 108 / 7
        for index in range(5):
            row_y = 515 - index * row_height
            for center_x in (176, 251):
                fill = placeholder_fill(center_x, row_y) if placeholder_fill else colors.white
                if fill is not None:
                    _erase_placeholder(pdf, center_x, row_y, fill)
        for index in range(7):
            center_x, center_y = 489, 515 - index * row_height
            fill = placeholder_fill(center_x, center_y) if placeholder_fill else colors.white
            if fill is not None:
                _erase_placeholder(pdf, center_x, center_y, fill)
        for center_x, center_y, placeholder_width in ((176, 390, 32), (251, 390, 32), (489, 390, 32), (352, 356, 40), (356, 328, 46)):
            fill = placeholder_fill(center_x, center_y) if placeholder_fill else colors.white
            if fill is not None:
                _erase_placeholder(pdf, center_x, center_y, fill, placeholder_width)
    else:
        _fit_text(pdf, details.get("unit_address", ""), x, 752, width, "center", 6.6)
    if erase:
        _fit_text(pdf, month_label, 155, 703, 189, "center", 7.5)
        _fit_text(pdf, details.get("employee_code", ""), 425, 703, 129, "center", 7.5)
        identity = (("employee_name", 117, 661, 122), ("unit", 307, 661, 111), ("uan", 473, 661, 81), ("date_of_joining", 117, 644, 122), ("designation", 307, 644, 111), ("esi_number", 473, 644, 81))
    else:
        _fit_text(pdf, month_label, 150, 712, width - 114, "center", 7.5)
        identity = (("employee_name", 116, 683, 120), ("unit", 306, 683, 105), ("uan", 466, 683, 93), ("employee_code", 116, 666, 120), ("designation", 306, 666, 105), ("esi_number", 466, 666, 93), ("date_of_joining", 116, 649, 120), ("ot_hours", 306, 649, 105))
    for key, px, py, cell in identity:
        _fit_text(pdf, details.get(key, ""), px, py, cell, "center")
    attendance_x, attendance_width, attendance_y = (40, 514, 588) if erase else (x, width, 591)
    for value, index in zip((details.get("days"), details.get("present_days"), details.get("lop"), details.get("ot_hours")), range(4)):
        _fit_text(pdf, value or "-", attendance_x + index * attendance_width / 4, attendance_y, attendance_width / 4, "center")
    earnings = (("basic_gross", "basic_earnings"), ("hra_gross", "hra_earnings"), ("special_allowance_gross", "special_allowance_earnings"), ("overtime_gross", "overtime_earnings"), ("variable_pay_gross", "variable_pay_earnings"))
    row_height = 108 / 7
    earning_y = 515 if erase else 520
    gross_x, earned_x, gross_width, earned_width = (137, 215, 78, 72) if erase else (136, 213, 77, 77)
    for index, (gross_key, earned_key) in enumerate(earnings):
        y = earning_y - index * row_height
        _fit_text(pdf, _amount(details, gross_key), gross_x, y, gross_width, "right")
        _fit_text(pdf, _amount(details, earned_key), earned_x, y, earned_width, "right")
    deductions = ("pf", "esi_deduction", "professional_tax", "loan", "advance", "other_deductions", "tds")
    deduction_x, deduction_width = (428, 122) if erase else (425, 134)
    for index, key in enumerate(deductions):
        _fit_text(pdf, _amount(details, key), deduction_x, earning_y - index * row_height, deduction_width, "right")
    total_y, net_y, words_y = (390, 356, 328) if erase else (407, 377, 349)
    _fit_text(pdf, _amount(details, "total_gross"), gross_x, total_y, gross_width, "right", bold=True)
    _fit_text(pdf, _amount(details, "total_earnings"), earned_x, total_y, earned_width, "right", bold=True)
    _fit_text(pdf, _amount(details, "deduction_total"), deduction_x, total_y, deduction_width, "right", bold=True)
    _fit_text(pdf, _amount(details, "net_wages"), 155 if erase else 152, net_y, 395 if erase else 407, "center", 8, True)
    _fit_text(pdf, details.get("net_wages_words", ""), 162 if erase else 166, words_y, 388 if erase else 393, "center", 7)


def generate_salary_pdf(details: dict, payroll_month: str, output_path: Path, password: str, template_path: Path | None = None) -> None:
    if template_path:
        template_reader = PdfReader(str(template_path))
        template_page = template_reader.pages[0]
        page_width, page_height = float(template_page.mediabox.width), float(template_page.mediabox.height)
        overlay_buffer = io.BytesIO()
        overlay = canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))
        if template_page.images:
            background_buffer = io.BytesIO()
            template_image = template_page.images[0].image.convert("RGB")
            template_image.save(background_buffer, format="PNG")
            background_buffer.seek(0)
            overlay.drawImage(ImageReader(background_buffer), 0, 0, width=page_width, height=page_height, preserveAspectRatio=False, mask="auto")
            image_width, image_height = template_image.size

            def placeholder_fill(center_x: float, center_y: float) -> colors.Color | None:
                pixel_x = round(center_x / TEMPLATE_BASE_WIDTH * image_width)
                pixel_y = round((TEMPLATE_BASE_HEIGHT - center_y) / TEMPLATE_BASE_HEIGHT * image_height)
                radius_x = max(4, round(10 / TEMPLATE_BASE_WIDTH * image_width))
                radius_y = max(2, round(2.5 / TEMPLATE_BASE_HEIGHT * image_height))
                dark_pixels = 0
                for y_value in range(max(0, pixel_y - radius_y), min(image_height, pixel_y + radius_y + 1)):
                    for x_value in range(max(0, pixel_x - radius_x), min(image_width, pixel_x + radius_x + 1)):
                        red, green, blue = template_image.getpixel((x_value, y_value))
                        if (red + green + blue) / 3 < 175:
                            dark_pixels += 1
                if dark_pixels < 3:
                    return None
                sample_x = min(image_width - 2, pixel_x + max(5, round(18 / TEMPLATE_BASE_WIDTH * image_width)))
                samples = []
                for y_value in range(max(0, pixel_y - 2), min(image_height, pixel_y + 3)):
                    for x_value in range(max(0, sample_x - 2), min(image_width, sample_x + 3)):
                        samples.append(template_image.getpixel((x_value, y_value)))
                channel_values = [sorted(pixel[channel] for pixel in samples) for channel in range(3)]
                middle = len(samples) // 2
                return colors.Color(channel_values[0][middle] / 255, channel_values[1][middle] / 255, channel_values[2][middle] / 255)
        else:
            placeholder_fill = None
        overlay.scale(page_width / TEMPLATE_BASE_WIDTH, page_height / TEMPLATE_BASE_HEIGHT)
        _draw_values(overlay, details, payroll_month, True, placeholder_fill)
        overlay.save()
        if template_page.images:
            plain_buffer = overlay_buffer
        else:
            template_page.merge_page(PdfReader(io.BytesIO(overlay_buffer.getvalue())).pages[0])
            plain_writer = PdfWriter()
            plain_writer.add_page(template_page)
            plain_buffer = io.BytesIO()
            plain_writer.write(plain_buffer)
    else:
        plain_buffer = io.BytesIO()
        pdf = canvas.Canvas(plain_buffer, pagesize=A4)
        _draw_default_template(pdf)
        _draw_values(pdf, details, payroll_month, False)
        pdf.save()
    reader = PdfReader(io.BytesIO(plain_buffer.getvalue()))
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    writer.encrypt(password, algorithm="AES-256")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("wb") as file:
        writer.write(file)
