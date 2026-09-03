import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.assets.models import ITAsset
from app.modules.identity.models import Organization

INVENTORY_REGISTER_DIR = Path(__file__).resolve().parents[2] / "assets" / "inventory"
ASSET_TEMPLATE_PATH = INVENTORY_REGISTER_DIR / "IT ASSETS ORG.xlsx"
INVENTORY_REGISTER_PATHS = tuple(sorted(INVENTORY_REGISTER_DIR.glob("*.xlsx"), key=lambda path: path.name.lower()))
EMPTY_IDENTIFIERS = {"", "nil", "nll", "n il", "0", "none", "not found", "not fund", "no label"}


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def meaningful_identifier(value: str | None) -> bool:
    return bool(value and value.strip().lower() not in EMPTY_IDENTIFIERS and not value.startswith("#"))


def parse_date(value: object, workbook_epoch) -> tuple[date | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, workbook_epoch).date(), None
        except (ValueError, OverflowError):
            return None, clean_text(value)
    text = clean_text(value)
    if not text or text.startswith("#"):
        return None, None
    normalized = re.sub(r"-{2,}", "-", text)
    for pattern in (
        "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y",
        "%m/%d/%y", "%d %b %Y", "%d-%b-%Y", "%d/%b/%Y",
    ):
        try:
            return datetime.strptime(normalized, pattern).date(), None
        except ValueError:
            continue
    return None, text


def parse_price(value: object) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01")), None
        except InvalidOperation:
            pass
    text = clean_text(value)
    if not text or text.startswith("#"):
        return None, None
    try:
        return Decimal(text.replace(",", "").replace("₹", "").strip()).quantize(Decimal("0.01")), None
    except InvalidOperation:
        return None, text


def imported_status(*values: object) -> tuple[str, str]:
    combined = " ".join(filter(None, (clean_text(value) for value in values))).lower()
    if "destroyed" in combined:
        return "Scrap proposed", "Damaged"
    if "disposed" in combined:
        return "Disposed", "Obsolete"
    if "scrap" in combined:
        return "Scrapped", "Damaged"
    if "not working" in combined or "not work" in combined or "repair" in combined:
        return "Repair needed", "Poor"
    if "under maintenance" in combined or "waiting for maintenance" in combined:
        return "Under maintenance", "Fair"
    if "former employee" in combined or "not received" in combined:
        return "Recovery required", "Fair"
    if "spare" in combined or "not in use" in combined or "not in us" in combined:
        return "Spare", "Good"
    return "Active", "Good"


def asset_key(
    row_number: int,
    source_sn: str | None,
    label_no: str | None,
    serial_imei: str | None,
    sim_no: str | None,
    source_register: str | None = None,
) -> str:
    register = re.sub(r"[^a-z0-9]+", "-", (source_register or "").lower()).strip("-")
    prefix = "" if register in {"", "it-assets-org-xlsx", "it-assets-org"} else f"{register}:"
    for kind, value in (("label", label_no), ("serial", serial_imei), ("sim", sim_no)):
        if meaningful_identifier(value):
            return f"{prefix}{kind}:{value.lower()}"
    fallback = f"sn:{source_sn}" if meaningful_identifier(source_sn) else f"row:{row_number}"
    return f"{prefix}{fallback}"


def _details(**values: object) -> dict[str, str]:
    return {label: text for label, value in values.items() if (text := clean_text(value))}


def _base_row(register: str, row_number: int, source_sn: object, **values: object) -> dict:
    source_sn_text = clean_text(source_sn)
    label_no = clean_text(values.pop("label_no", None))
    serial_imei = clean_text(values.pop("serial_imei", None))
    sim_no = clean_text(values.pop("sim_no", None))
    row = {
        "source_register": register,
        "source_sn": source_sn_text,
        "label_no": label_no,
        "serial_imei": serial_imei,
        "sim_no": sim_no,
        **values,
    }
    row["asset_key"] = asset_key(row_number, source_sn_text, label_no, serial_imei, sim_no, register)
    return row


def _detect_profile(first_rows: list[tuple], filename: str) -> str:
    sample = " ".join(clean_text(value) or "" for row in first_rows for value in row).upper()
    name = filename.upper()
    if "CHASSIS" in sample or "WHEELER" in name:
        return "vehicle"
    if "STABILIZER" in sample or "AIR CONDITIONER" in name:
        return "air_conditioner"
    if "EMPLOYE" in sample and "SERIAL NO/IMEI NO" in sample:
        return "it"
    if "CCTV" in sample or "CCTV" in name:
        return "cctv"
    if "WEIGHING SCALE" in sample or "WEIGHING" in name:
        return "weighing_scale"
    if "REFRIGERATOR" in sample or "REFRIGERATOR" in name:
        return "refrigerator"
    if "FACTORY MACHINE" in sample:
        return "factory_unit1" if "UNIT1" in name else "factory_unit3"
    raise ValueError("Unsupported asset register layout")


def read_asset_rows(content_path: Path | BinaryIO, source_filename: str | None = None) -> list[dict]:
    workbook = load_workbook(content_path, data_only=True, read_only=True)
    try:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        filename = source_filename or (content_path.name if isinstance(content_path, Path) else "Uploaded asset register.xlsx")
        register = Path(filename).name
        profile = _detect_profile(rows[:8], register)
        result: list[dict] = []

        if profile == "it":
            for row_number, values in enumerate(rows[1:], start=2):
                values = tuple(values) + (None,) * (17 - len(values))
                if not any(value not in (None, "") for value in values[:17]):
                    continue
                invoice_date, invoice_date_raw = parse_date(values[12], workbook.epoch)
                price, price_raw = parse_price(values[15])
                status, condition = imported_status(values[1], values[4], values[5])
                result.append(_base_row(
                    register, row_number, values[0], employee=clean_text(values[1]),
                    physical_location=clean_text(values[2]), department_name=clean_text(values[3]),
                    home_office=clean_text(values[4]), category=clean_text(values[5]), brand=clean_text(values[6]),
                    model=clean_text(values[7]), serial_imei=values[8], sim_no=values[9], ups=clean_text(values[10]),
                    label_no=values[11], invoice_date=invoice_date, invoice_date_raw=invoice_date_raw,
                    invoice_no=clean_text(values[13]), supplier_name=clean_text(values[14]), price=price,
                    price_raw=price_raw, warranty=clean_text(values[16]), custom_fields={}, status=status, condition=condition,
                ))

        elif profile == "air_conditioner":
            for row_number, values in enumerate(rows[4:], start=5):
                values = tuple(values) + (None,) * (12 - len(values))
                if not any(values[index] not in (None, "") for index in (0, 1, 2, 3, 11)):
                    continue
                invoice_date, invoice_date_raw = parse_date(values[6], workbook.epoch)
                service_date, service_date_raw = parse_date(values[5], workbook.epoch)
                price, price_raw = parse_price(values[10])
                status, condition = imported_status(values[4], values[5])
                result.append(_base_row(
                    register, row_number, values[0], physical_location=clean_text(values[1]),
                    department_name=clean_text(values[2]), category="Air Conditioner", brand=clean_text(values[3]),
                    label_no=values[11], invoice_date=invoice_date, invoice_date_raw=invoice_date_raw,
                    invoice_no=clean_text(values[7]), supplier_name=clean_text(values[9]), price=price,
                    price_raw=price_raw, last_maintenance_date=service_date,
                    custom_fields=_details(Stabilizer=values[4], Tonnage=values[8], **{"Service date note": service_date_raw}),
                    status=status, condition=condition,
                ))

        elif profile == "vehicle":
            for row_number, values in enumerate(rows[3:], start=4):
                values = tuple(values) + (None,) * (9 - len(values))
                if not any(values[index] not in (None, "") for index in (0, 1, 2, 3)):
                    continue
                status, condition = imported_status(values[2], values[7], values[8])
                result.append(_base_row(
                    register, row_number, values[0], employee=clean_text(values[1]), category="Vehicle",
                    model=clean_text(values[2]), label_no=values[3], serial_imei=values[4],
                    custom_fields=_details(**{
                        "Engine number": values[5], "Registration / FC valid until": values[6],
                        "Insurance period": values[7], "Pollution certificate": values[8],
                    }), status=status, condition=condition,
                ))

        elif profile == "cctv":
            for row_number, values in enumerate(rows[2:], start=3):
                values = tuple(values) + (None,) * (4 - len(values))
                if not any(values[index] not in (None, "") for index in (0, 1, 2, 3)):
                    continue
                result.append(_base_row(
                    register, row_number, values[0], physical_location=clean_text(values[1]),
                    department_name=clean_text(values[2]), category="CCTV Camera", label_no=values[3],
                    custom_fields={}, status="Active", condition="Good",
                ))

        elif profile in {"factory_unit1", "factory_unit3"}:
            unit1 = profile == "factory_unit1"
            start_index = 4
            for row_number, values in enumerate(rows[start_index:], start=start_index + 1):
                values = tuple(values) + (None,) * (21 - len(values))
                category = values[3]
                label_index = 13 if unit1 else 15
                if not any(value not in (None, "") for value in (values[0], category, values[label_index])):
                    continue
                if unit1:
                    brand, serial, model = values[4], values[5], None
                    operational, usage = values[6], values[7]
                    disposed, maintenance, notes, annual = values[9], values[10], values[11], values[12]
                    invoice_date_raw_value, invoice_no, supplier, raw_price, warranty, issued = values[14:20]
                else:
                    brand, serial, model = values[4], values[5], values[6]
                    operational, usage, disposed, maintenance, notes = values[7:12]
                    annual = None
                    invoice_no, invoice_date_raw_value, supplier = values[12], values[13], values[14]
                    raw_price, warranty, issued = values[16], values[17], values[18]
                invoice_date, invoice_date_raw = parse_date(invoice_date_raw_value, workbook.epoch)
                price, price_raw = parse_price(raw_price)
                status, condition = imported_status(operational, usage, disposed, maintenance, notes)
                result.append(_base_row(
                    register, row_number, values[0], physical_location=clean_text(values[1]),
                    department_name=clean_text(values[2]), category=clean_text(category) or "Factory Machine",
                    brand=clean_text(brand), model=clean_text(model), serial_imei=serial, label_no=values[label_index],
                    invoice_date=invoice_date, invoice_date_raw=invoice_date_raw, invoice_no=clean_text(invoice_no),
                    supplier_name=clean_text(supplier), price=price, price_raw=price_raw, warranty=clean_text(warranty),
                    maintenance_notes=clean_text(maintenance), notes=clean_text(notes),
                    custom_fields=_details(**{
                        "Operating condition": operational, "Usage status": usage, "Disposed": disposed,
                        "Annual maintenance": annual, "Issued by": issued,
                    }), status=status, condition=condition,
                ))

        elif profile == "weighing_scale":
            for row_number, values in enumerate(rows[4:], start=5):
                values = tuple(values) + (None,) * (22 - len(values))
                if not any(value not in (None, "") for value in (values[0], values[3], values[15])):
                    continue
                invoice_date, invoice_date_raw = parse_date(values[18], workbook.epoch)
                price, price_raw = parse_price(values[20])
                status, condition = imported_status(values[8], values[9], values[11], values[12], values[13], values[16])
                result.append(_base_row(
                    register, row_number, values[0], physical_location=clean_text(values[1]),
                    department_name=clean_text(values[2]), category="Weighing Scale", brand=clean_text(values[3]),
                    model=clean_text(values[4]) or clean_text(values[6]), serial_imei=values[7], label_no=values[15],
                    invoice_date=invoice_date, invoice_date_raw=invoice_date_raw, invoice_no=clean_text(values[19]),
                    supplier_name=clean_text(values[17]), price=price, price_raw=price_raw, warranty=clean_text(values[21]),
                    maintenance_notes=clean_text(values[12]), notes=clean_text(values[16]) or clean_text(values[13]),
                    custom_fields=_details(**{
                        "Capacity": values[5], "Operating condition": values[8], "Usage status": values[9],
                        "Disposed": values[11], "Annual maintenance": values[14],
                    }), status=status, condition=condition,
                ))

        elif profile == "refrigerator":
            for row_number, values in enumerate(rows[2:], start=3):
                values = tuple(values) + (None,) * (13 - len(values))
                if not any(value not in (None, "") for value in (values[0], values[3], values[11])):
                    continue
                invoice_date, invoice_date_raw = parse_date(values[8], workbook.epoch)
                maintenance_date, maintenance_date_raw = parse_date(values[9], workbook.epoch)
                price, price_raw = parse_price(values[12])
                result.append(_base_row(
                    register, row_number, values[0], physical_location=clean_text(values[1]),
                    department_name=clean_text(values[2]), category="Refrigerator", brand=clean_text(values[3]),
                    model=clean_text(values[5]), label_no=values[11], invoice_date=invoice_date,
                    invoice_date_raw=invoice_date_raw, invoice_no=clean_text(values[7]), supplier_name=clean_text(values[6]),
                    price=price, price_raw=price_raw, last_maintenance_date=maintenance_date, notes=clean_text(values[10]),
                    custom_fields=_details(Capacity=values[4], **{"Maintenance date note": maintenance_date_raw}),
                    status="Active", condition="Good",
                ))
        for row in result:
            row["asset_group"] = "IT" if profile == "it" else "General"
        return result
    finally:
        workbook.close()


SOURCE_FIELDS = [
    "source_sn", "source_register", "asset_group", "employee", "physical_location", "department_name", "home_office",
    "category", "brand", "model", "serial_imei", "sim_no", "ups", "label_no", "invoice_date",
    "invoice_date_raw", "invoice_no", "supplier_name", "price", "price_raw", "warranty", "custom_fields",
]


async def merge_asset_rows(
    session: AsyncSession,
    organization_id: uuid.UUID,
    rows: list[dict],
    new_defaults: dict | None = None,
) -> tuple[int, int]:
    existing = {item.asset_key: item for item in await session.scalars(select(ITAsset).where(ITAsset.organization_id == organization_id))}
    created = updated = 0
    for row in rows:
        item = existing.get(row["asset_key"])
        if item:
            for field in SOURCE_FIELDS:
                setattr(item, field, row.get(field))
            updated += 1
        else:
            values = {**(new_defaults or {}), **row}
            session.add(ITAsset(organization_id=organization_id, **values))
            created += 1
    return created, updated


async def seed_asset_register(session: AsyncSession) -> None:
    if not INVENTORY_REGISTER_PATHS:
        return
    organizations = list(await session.scalars(select(Organization)))
    for organization in organizations:
        for path in INVENTORY_REGISTER_PATHS:
            await merge_asset_rows(session, organization.id, read_asset_rows(path))
    if organizations:
        await session.commit()
