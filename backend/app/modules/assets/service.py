import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.assets.models import ITAsset
from app.modules.identity.models import Organization

ASSET_TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "assets" / "assets" / "IT ASSETS ORG.xlsx"
HEADERS = ["SN", "EMPLOYE", "PHYSICAL LOCATION", "DEPARTMENT", "HOME & OFFICE", "CATEGORY", "BRAND", "MODEL", "SERIAL NO/IMEI NO", "SIM NO:", "UPS", " LABEL NO:", "INVOICE DATE", "INVOICE NO.", "SUPPLIER NAME", " PRICE ", "WARRANTY"]
EMPTY_IDENTIFIERS = {"", "nil", "nll", "n il", "0", "none", "not found", "not fund"}


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def meaningful_identifier(value: str | None) -> bool:
    return bool(value and value.strip().lower() not in EMPTY_IDENTIFIERS and not value.startswith("#"))


def parse_date(value: object, workbook_epoch) -> tuple[date | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, workbook_epoch).date(), None
        except (ValueError, OverflowError):
            return None, clean_text(value)
    text = clean_text(value)
    if not text or text.startswith("#"):
        return None, None
    for pattern in ("%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y", "%m/%d/%y", "%d %b %Y"):
        try:
            return datetime.strptime(text, pattern).date(), None
        except ValueError:
            continue
    return None, text


def parse_price(value: object) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01")), None
        except InvalidOperation:
            pass
    text = clean_text(value)
    if not text or text.startswith("#"):
        return None, None
    try:
        return Decimal(text.replace(",", "").replace("₹", "").strip()).quantize(Decimal("0.01")), None
    except InvalidOperation:
        return None, text


def imported_status(employee: str | None, home_office: str | None, category: str | None) -> tuple[str, str]:
    combined = " ".join(filter(None, (employee, home_office, category))).lower()
    if "destroyed" in combined:
        return "Scrap proposed", "Damaged"
    if "former employee" in combined or "not received" in combined:
        return "Recovery required", "Fair"
    if "spare" in combined:
        return "Spare", "Good"
    return "Active", "Good"


def asset_key(row_number: int, source_sn: str | None, label_no: str | None, serial_imei: str | None, sim_no: str | None) -> str:
    for prefix, value in (("label", label_no), ("serial", serial_imei), ("sim", sim_no)):
        if meaningful_identifier(value):
            return f"{prefix}:{value.lower()}"
    return f"sn:{source_sn}" if meaningful_identifier(source_sn) else f"row:{row_number}"


def read_asset_rows(content_path: Path | BinaryIO) -> list[dict]:
    workbook = load_workbook(content_path, data_only=True, read_only=True)
    try:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        rows: list[dict] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, max_col=17, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            source_sn, employee, location, department, home_office, category, brand, model, serial_imei, sim_no, ups, label_no, raw_invoice_date, invoice_no, supplier, raw_price, warranty = values
            source_sn_text = clean_text(source_sn)
            employee_text = clean_text(employee)
            home_text = clean_text(home_office)
            category_text = clean_text(category)
            label_text = clean_text(label_no)
            serial_text = clean_text(serial_imei)
            sim_text = clean_text(sim_no)
            invoice_date, invoice_date_raw = parse_date(raw_invoice_date, workbook.epoch)
            price, price_raw = parse_price(raw_price)
            status, condition = imported_status(employee_text, home_text, category_text)
            rows.append({
                "asset_key": asset_key(row_number, source_sn_text, label_text, serial_text, sim_text),
                "source_sn": source_sn_text,
                "employee": employee_text,
                "physical_location": clean_text(location),
                "department_name": clean_text(department),
                "home_office": home_text,
                "category": category_text,
                "brand": clean_text(brand),
                "model": clean_text(model),
                "serial_imei": serial_text,
                "sim_no": sim_text,
                "ups": clean_text(ups),
                "label_no": label_text,
                "invoice_date": invoice_date,
                "invoice_date_raw": invoice_date_raw,
                "invoice_no": clean_text(invoice_no),
                "supplier_name": clean_text(supplier),
                "price": price,
                "price_raw": price_raw,
                "warranty": clean_text(warranty),
                "status": status,
                "condition": condition,
            })
        return rows
    finally:
        workbook.close()


SOURCE_FIELDS = ["source_sn", "employee", "physical_location", "department_name", "home_office", "category", "brand", "model", "serial_imei", "sim_no", "ups", "label_no", "invoice_date", "invoice_date_raw", "invoice_no", "supplier_name", "price", "price_raw", "warranty"]


async def merge_asset_rows(session: AsyncSession, organization_id: uuid.UUID, rows: list[dict]) -> tuple[int, int]:
    existing = {item.asset_key: item for item in await session.scalars(select(ITAsset).where(ITAsset.organization_id == organization_id))}
    created = updated = 0
    for row in rows:
        item = existing.get(row["asset_key"])
        if item:
            for field in SOURCE_FIELDS:
                setattr(item, field, row.get(field))
            updated += 1
        else:
            session.add(ITAsset(organization_id=organization_id, **row))
            created += 1
    return created, updated


async def seed_asset_register(session: AsyncSession) -> None:
    if not ASSET_TEMPLATE_PATH.exists():
        return
    organizations = list(await session.scalars(select(Organization)))
    rows = read_asset_rows(ASSET_TEMPLATE_PATH)
    changed = False
    for organization in organizations:
        count = await session.scalar(select(func.count(ITAsset.id)).where(ITAsset.organization_id == organization.id))
        if not count:
            await merge_asset_rows(session, organization.id, rows)
            changed = True
    if changed:
        await session.commit()
