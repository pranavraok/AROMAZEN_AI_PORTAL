import io
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

from app.modules.payroll.engine import generate_salary_pdf, password_for, read_salary_excel


def main() -> None:
    source, template, protected_output, preview_output = map(Path, sys.argv[1:5])
    workbook = load_workbook(source)
    sheet = workbook.active
    dob_column = sheet.max_column + 1
    sheet.cell(1, dob_column, "DATE OF BIRTH")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 4).value not in (None, ""):
            sheet.cell(row, dob_column, date(1992, 5, 18))
    content = io.BytesIO()
    workbook.save(content)
    rows = read_salary_excel(content.getvalue())
    assert rows
    first = rows[0]
    password = password_for(first["employee_name"], first["birth_year"])
    generate_salary_pdf(first["details"], "2026-07", protected_output, password, template)
    reader = PdfReader(protected_output)
    assert reader.is_encrypted and reader.decrypt(password)
    writer = PdfWriter()
    writer.add_page(reader.pages[0])
    with preview_output.open("wb") as stream:
        writer.write(stream)
    print(f"Real workbook QA passed: {len(rows)} rows, Unit {first['details']['unit']}, net {first['details']['net_wages']}")


if __name__ == "__main__":
    main()
