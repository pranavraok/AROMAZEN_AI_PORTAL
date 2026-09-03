import io
import uuid
from copy import copy
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.db.session import get_db_session
from app.modules.assets.models import AssetMaintenanceEvent, AssetNotificationSetting, ITAsset
from app.modules.assets.schemas import (
    AssetListResponse, AssetNotificationSettingsPayload, AssetNotificationSettingsRead,
    AssetPayload, AssetRead, AssetSummary, AssetUpdatePayload, MaintenancePayload, MaintenanceRead,
)
from app.modules.assets.service import ASSET_TEMPLATE_PATH, asset_key, clean_text, merge_asset_rows, read_asset_rows
from app.modules.identity.authorization import require_permissions
from app.modules.identity.models import AuditEvent, Department, User
from app.modules.identity.service import role_keys_for_user
from app.modules.knowledge.department_uploads import DepartmentUpload, replace_department_uploads

router = APIRouter()
INACTIVE_MAINTENANCE_STATUSES = {"Scrapped", "Disposed", "Lost"}
SCRAP_QUEUE_STATUSES = {"Scrap proposed", "Approved for scrap"}
ASSET_ADMIN_DEPARTMENTS = {"inventory", "hr", "human-resources", "accounts"}


async def ensure_asset_access(user: User, session: AsyncSession) -> None:
    roles = await role_keys_for_user(session, user.id)
    if roles.intersection({"owner", "super_admin"}):
        return
    department = await session.get(Department, user.department_id) if user.department_id else None
    if "department_admin" not in roles or not department or department.slug not in ASSET_ADMIN_DEPARTMENTS:
        raise HTTPException(
            status_code=403,
            detail="Asset Management is available to Inventory, HR and Accounts administrators.",
        )


async def notification_settings(session: AsyncSession, organization_id: uuid.UUID) -> AssetNotificationSetting:
    value = await session.get(AssetNotificationSetting, organization_id)
    if value:
        return value
    value = AssetNotificationSetting(organization_id=organization_id)
    session.add(value)
    await session.flush()
    return value


def maintenance_status(asset: ITAsset, today: date | None = None) -> tuple[str, int | None]:
    if not asset.next_maintenance_date or asset.status in INACTIVE_MAINTENANCE_STATUSES:
        return "not_scheduled", None
    current = today or datetime.now(timezone.utc).date()
    remaining = (asset.next_maintenance_date - current).days
    if remaining < 0:
        return "overdue", remaining
    if remaining <= asset.maintenance_reminder_days:
        return "due", remaining
    return "scheduled", remaining


def asset_read(asset: ITAsset) -> AssetRead:
    state, days = maintenance_status(asset)
    values = {column.name: getattr(asset, column.name) for column in ITAsset.__table__.columns}
    values["maintenance_state"] = state
    values["maintenance_days_remaining"] = days
    values["price"] = float(asset.price) if asset.price is not None else None
    values["scrap_value"] = float(asset.scrap_value) if asset.scrap_value is not None else None
    return AssetRead.model_validate(values)


def summary_for(items: list[ITAsset]) -> AssetSummary:
    states = [maintenance_status(item)[0] for item in items]
    return AssetSummary(
        total=len(items),
        active=sum(item.status == "Active" for item in items),
        spare=sum(item.status == "Spare" for item in items),
        maintenance_due=sum(state in {"due", "overdue"} for state in states),
        maintenance_overdue=sum(state == "overdue" for state in states),
        repair_needed=sum(item.status in {"Repair needed", "Under maintenance"} for item in items),
        recovery_required=sum(item.status == "Recovery required" for item in items),
        scrap_queue=sum(item.status in SCRAP_QUEUE_STATUSES for item in items),
        scrapped_or_disposed=sum(item.status in {"Scrapped", "Disposed"} for item in items),
        total_value=round(sum(float(item.price or 0) for item in items), 2),
    )


def source_order(item: ITAsset) -> tuple[int, int | str]:
    if item.source_sn and item.source_sn.isdigit():
        return 0, int(item.source_sn)
    return 1, item.source_sn or ""


@router.get("", response_model=AssetListResponse)
async def list_assets(
    search: str = Query("", max_length=200),
    status: str = Query("All", max_length=60),
    category: str = Query("All", max_length=160),
    location: str = Query("All", max_length=240),
    department: str = Query("All", max_length=240),
    register: str = Query("All", max_length=240),
    asset_group: str = Query("IT", max_length=40),
    attention_only: bool = False,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetListResponse:
    await ensure_asset_access(user, session)
    all_items = sorted(list(await session.scalars(select(ITAsset).where(ITAsset.organization_id == user.organization_id))), key=source_order)
    group_counts = {
        "IT": sum(item.asset_group == "IT" for item in all_items),
        "General": sum(item.asset_group == "General" for item in all_items),
    }
    grouped_items = all_items if asset_group == "All" else [item for item in all_items if item.asset_group == asset_group]
    filtered = grouped_items
    if search.strip():
        needle = search.strip().lower()
        filtered = [item for item in filtered if needle in " ".join(str(value or "") for value in (item.source_sn, item.employee, item.physical_location, item.department_name, item.category, item.brand, item.model, item.serial_imei, item.sim_no, item.label_no)).lower()]
    if status != "All":
        filtered = [item for item in filtered if item.status == status]
    if category != "All":
        filtered = [item for item in filtered if item.category == category]
    if location != "All":
        filtered = [item for item in filtered if item.physical_location == location]
    if department != "All":
        filtered = [item for item in filtered if item.department_name == department]
    if register != "All":
        filtered = [item for item in filtered if item.source_register == register]
    if attention_only:
        filtered = [item for item in filtered if maintenance_status(item)[0] in {"due", "overdue"} or item.status in {"Repair needed", "Recovery required", *SCRAP_QUEUE_STATUSES}]
    return AssetListResponse(
        items=[asset_read(item) for item in filtered],
        summary=summary_for(grouped_items),
        categories=sorted({item.category for item in grouped_items if item.category}, key=str.lower),
        locations=sorted({item.physical_location for item in grouped_items if item.physical_location}, key=str.lower),
        departments=sorted({item.department_name for item in grouped_items if item.department_name}, key=str.lower),
        registers=sorted({item.source_register for item in grouped_items if item.source_register}, key=str.lower),
        group_counts=group_counts,
    )


@router.post("", response_model=AssetRead, status_code=201)
async def create_asset(
    payload: AssetPayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetRead:
    await ensure_asset_access(user, session)
    existing = list(await session.scalars(select(ITAsset.source_sn).where(ITAsset.organization_id == user.organization_id)))
    numeric = [int(value) for value in existing if value and value.isdigit()]
    source_sn = str(max(numeric, default=0) + 1)
    values = payload.model_dump()
    defaults = await notification_settings(session, user.organization_id)
    if "maintenance_reminder_days" not in payload.model_fields_set:
        values["maintenance_reminder_days"] = defaults.default_reminder_days
    if "notification_enabled" not in payload.model_fields_set:
        values["notification_enabled"] = defaults.default_notification_enabled
    if "maintenance_interval_months" not in payload.model_fields_set:
        values["maintenance_interval_months"] = defaults.default_maintenance_interval_months
    values["source_register"] = values.get("source_register") or "Manual entry"
    item = ITAsset(
        organization_id=user.organization_id,
        asset_key=asset_key(0, source_sn, values.get("label_no"), values.get("serial_imei"), values.get("sim_no")) + f":{uuid.uuid4().hex[:8]}",
        source_sn=source_sn,
        invoice_date_raw=None,
        price_raw=None,
        **values,
    )
    session.add(item)
    await session.flush()
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.created", target_type="it_asset", target_id=str(item.id), metadata_json={"category": item.category, "label_no": item.label_no}))
    await session.commit()
    await session.refresh(item)
    return asset_read(item)


@router.get("/notification-settings", response_model=AssetNotificationSettingsRead)
async def get_notification_settings(
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetNotificationSettingsRead:
    await ensure_asset_access(user, session)
    value = await notification_settings(session, user.organization_id)
    await session.commit()
    return AssetNotificationSettingsRead.model_validate({
        **{column.name: getattr(value, column.name) for column in AssetNotificationSetting.__table__.columns},
        "apply_to_current_assets": False,
    })


@router.put("/notification-settings", response_model=AssetNotificationSettingsRead)
async def update_notification_settings(
    payload: AssetNotificationSettingsPayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetNotificationSettingsRead:
    await ensure_asset_access(user, session)
    value = await notification_settings(session, user.organization_id)
    updates = payload.model_dump(exclude={"apply_to_current_assets"})
    for field, setting in updates.items():
        setattr(value, field, setting)
    if payload.apply_to_current_assets:
        items = list(await session.scalars(select(ITAsset).where(ITAsset.organization_id == user.organization_id)))
        for item in items:
            item.notification_enabled = payload.default_notification_enabled
            item.maintenance_reminder_days = payload.default_reminder_days
            if payload.default_maintenance_interval_months is not None:
                item.maintenance_interval_months = payload.default_maintenance_interval_months
    session.add(AuditEvent(
        organization_id=user.organization_id, actor_user_id=user.id, action="asset.notification_settings_updated",
        target_type="asset_notification_settings", target_id=str(user.organization_id),
        metadata_json={**updates, "apply_to_current_assets": payload.apply_to_current_assets},
    ))
    await session.commit()
    await session.refresh(value)
    return AssetNotificationSettingsRead.model_validate({
        **{column.name: getattr(value, column.name) for column in AssetNotificationSetting.__table__.columns},
        "apply_to_current_assets": False,
    })


async def get_asset(session: AsyncSession, asset_id: uuid.UUID, user: User) -> ITAsset:
    item = await session.scalar(select(ITAsset).where(ITAsset.id == asset_id, ITAsset.organization_id == user.organization_id))
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found.")
    return item


@router.patch("/{asset_id}", response_model=AssetRead)
async def update_asset(
    asset_id: uuid.UUID,
    payload: AssetUpdatePayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetRead:
    await ensure_asset_access(user, session)
    item = await get_asset(session, asset_id, user)
    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        if isinstance(value, str):
            value = clean_text(value)
        setattr(item, field, value)
    item.invoice_date_raw = None if "invoice_date" in updates else item.invoice_date_raw
    item.price_raw = None if "price" in updates else item.price_raw
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.updated", target_type="it_asset", target_id=str(item.id), metadata_json={"fields": sorted(updates)}))
    await session.commit()
    await session.refresh(item)
    return asset_read(item)


@router.delete("/{asset_id}", status_code=204)
async def delete_asset(
    asset_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> None:
    await ensure_asset_access(user, session)
    item = await get_asset(session, asset_id, user)
    session.add(AuditEvent(
        organization_id=user.organization_id, actor_user_id=user.id, action="asset.deleted",
        target_type="it_asset", target_id=str(item.id),
        metadata_json={"category": item.category, "label_no": item.label_no, "source_register": item.source_register},
    ))
    await session.delete(item)
    await session.commit()


@router.post("/{asset_id}/maintenance", response_model=AssetRead)
async def record_maintenance(
    asset_id: uuid.UUID,
    payload: MaintenancePayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetRead:
    await ensure_asset_access(user, session)
    item = await get_asset(session, asset_id, user)
    event = AssetMaintenanceEvent(organization_id=user.organization_id, asset_id=item.id, service_date=payload.service_date, vendor=clean_text(payload.vendor), cost=payload.cost, notes=clean_text(payload.notes), next_due_date=payload.next_due_date, created_by_user_id=user.id)
    session.add(event)
    item.last_maintenance_date = payload.service_date
    item.next_maintenance_date = payload.next_due_date
    item.maintenance_notes = clean_text(payload.notes) or item.maintenance_notes
    if item.status in {"Under maintenance", "Repair needed"}:
        item.status = "Active"
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.maintenance_recorded", target_type="it_asset", target_id=str(item.id), metadata_json={"service_date": payload.service_date.isoformat(), "next_due_date": payload.next_due_date.isoformat() if payload.next_due_date else None}))
    await session.commit()
    await session.refresh(item)
    return asset_read(item)


@router.get("/{asset_id}/maintenance", response_model=list[MaintenanceRead])
async def maintenance_history(
    asset_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> list[MaintenanceRead]:
    await ensure_asset_access(user, session)
    await get_asset(session, asset_id, user)
    events = list(await session.scalars(select(AssetMaintenanceEvent).where(AssetMaintenanceEvent.asset_id == asset_id, AssetMaintenanceEvent.organization_id == user.organization_id).order_by(AssetMaintenanceEvent.service_date.desc())))
    return [MaintenanceRead.model_validate(event) for event in events]


@router.post("/import")
async def import_assets(
    file: UploadFile = File(...),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await ensure_asset_access(user, session)
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=422, detail="Upload the IT asset register as an Excel .xlsx file.")
    content = await file.read(get_settings().max_excel_upload_size_mb * 1024 * 1024 + 1)
    if len(content) > get_settings().max_excel_upload_size_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"The Excel file exceeds {get_settings().max_excel_upload_size_mb} MB.")
    try:
        rows = read_asset_rows(io.BytesIO(content), file.filename)
    except Exception as error:
        raise HTTPException(status_code=422, detail="The IT asset workbook could not be read.") from error
    if not rows:
        raise HTTPException(status_code=422, detail="No asset rows were found in the workbook.")
    defaults = await notification_settings(session, user.organization_id)
    created, updated = await merge_asset_rows(session, user.organization_id, rows, {
        "notification_enabled": defaults.default_notification_enabled,
        "maintenance_reminder_days": defaults.default_reminder_days,
        "maintenance_interval_months": defaults.default_maintenance_interval_months,
    })
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.register_imported", target_type="it_asset", target_id="register", metadata_json={"filename": file.filename, "created": created, "updated": updated}))
    documents = await replace_department_uploads(session, user, "inventory", [DepartmentUpload(
        source_key="assets:register",
        content=content,
        original_filename=file.filename or "IT_Asset_Register.xlsx",
        mime_type=file.content_type,
    )])
    return {"created": created, "updated": updated, "total_rows": len(rows), "knowledge_document_id": str(documents[0].id), "knowledge_version": documents[0].version}


@router.get("/export/register")
async def export_assets(
    asset_group: str = Query("All", max_length=40),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    await ensure_asset_access(user, session)
    items = sorted(list(await session.scalars(select(ITAsset).where(ITAsset.organization_id == user.organization_id))), key=source_order)
    if asset_group != "All":
        items = [item for item in items if item.asset_group == asset_group]
    workbook = load_workbook(ASSET_TEMPLATE_PATH) if ASSET_TEMPLATE_PATH.exists() else None
    if workbook:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        sheet.delete_rows(1, sheet.max_row)
        sheet.title = "Unified Asset Register"
    else:
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Unified Asset Register"
    headers = [
        "SN", "ASSET GROUP", "SOURCE REGISTER", "CATEGORY", "BRAND", "MODEL", "ASSIGNED TO", "LOCATION", "DEPARTMENT",
        "LABEL / REGISTRATION", "SERIAL / IMEI / CHASSIS", "SIM", "STATUS", "CONDITION", "INVOICE DATE",
        "INVOICE NO.", "SUPPLIER", "PRICE", "WARRANTY", "LAST SERVICE", "NEXT SERVICE", "REMINDER DAYS",
        "NOTIFICATIONS", "RESPONSIBLE PERSON", "SCRAP DATE", "SCRAP VALUE", "SCRAP REASON", "SOURCE DETAILS", "NOTES",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = copy(cell.font)
        cell.font = cell.font.copy(bold=True, color="FFFFFF")
    red_fill = PatternFill("solid", fgColor="FF0000")
    green_fill = PatternFill("solid", fgColor="92D050")
    for row_index, item in enumerate(items, start=2):
        source_sn: object = int(item.source_sn) if item.source_sn and item.source_sn.isdigit() else item.source_sn
        values = [
            source_sn, item.asset_group, item.source_register, item.category, item.brand, item.model, item.employee,
            item.physical_location, item.department_name, item.label_no, item.serial_imei, item.sim_no,
            item.status, item.condition, item.invoice_date or item.invoice_date_raw, item.invoice_no,
            item.supplier_name, float(item.price) if item.price is not None else item.price_raw, item.warranty,
            item.last_maintenance_date, item.next_maintenance_date, item.maintenance_reminder_days,
            "On" if item.notification_enabled else "Off", item.maintenance_owner, item.scrap_date,
            float(item.scrap_value) if item.scrap_value is not None else None, item.scrap_reason,
            "; ".join(f"{key}: {value}" for key, value in (item.custom_fields or {}).items()), item.notes,
        ]
        sheet.append(values)
        for cell in sheet[row_index]:
            if item.status in {"Recovery required", "Scrap proposed", "Approved for scrap", "Scrapped", "Disposed"}:
                cell.fill = copy(red_fill)
            elif item.status == "Spare":
                cell.fill = copy(green_fill)
        for column in (15, 20, 21, 25):
            sheet.cell(row_index, column).number_format = "dd-mm-yyyy"
        for column in (18, 26):
            sheet.cell(row_index, column).number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:AC{max(2, len(items) + 1)}"
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(36, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    filename = f"AROMAZEN {asset_group} Asset Register.xlsx" if asset_group != "All" else "AROMAZEN Unified Asset Register.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
