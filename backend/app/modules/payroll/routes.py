import csv
import io
import json
import re
import smtplib
import uuid
from calendar import monthrange
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.concurrency import run_in_threadpool
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import get_settings
from app.db.session import SessionLocal, get_db_session
from app.modules.identity.authorization import department_matches, require_department, require_permissions
from app.modules.identity.models import AuditEvent, Department, KnowledgeCollection, KnowledgeDocument, PayrollBatch, PayrollRecipient, PayrollTemplate, User
from app.modules.identity.service import role_keys_for_user
from app.modules.payroll.attendance_rules import DEFAULT_LATE_GRACE_MINUTES, apply_monthly_late_policy
from app.modules.payroll.engine import COLUMNS, create_excel_template, generate_salary_pdf, password_for, read_salary_excel, salary_template_form_fields, validate_template_pdf

router = APIRouter(dependencies=[Depends(require_department("hr"))])

DEFAULT_EMAIL_SUBJECT = "AROMAZEN Salary Slip - {month}"
DEFAULT_EMAIL_BODY = """Dear {employee_name},

Please find attached your salary slip for {month}.

The PDF password is the first four letters of your name in uppercase followed by your four-digit year of birth.

Regards,
HR Department
AROMAZEN PVT LTD"""


class EmailDraftUpdate(BaseModel):
    subject: str
    body: str


def _unit_from_template_name(filename: str) -> int | None:
    normalised = re.sub(r"[^a-z0-9]", "", filename.lower())
    match = re.search(r"unit([123])salaryslip", normalised)
    return int(match.group(1)) if match else None


async def _knowledge_unit_templates(session: AsyncSession, organization_id: uuid.UUID) -> dict[int, KnowledgeDocument]:
    documents = list(await session.scalars(
        select(KnowledgeDocument)
        .join(KnowledgeCollection, KnowledgeCollection.id == KnowledgeDocument.collection_id)
        .where(
            KnowledgeDocument.organization_id == organization_id,
            KnowledgeDocument.status == "ready",
            KnowledgeCollection.status == "active",
            KnowledgeCollection.slug == "hr",
        )
        .order_by(KnowledgeDocument.version.desc(), KnowledgeDocument.created_at.desc())
    ))
    result: dict[int, KnowledgeDocument] = {}
    for document in documents:
        unit = _unit_from_template_name(document.original_filename)
        if unit and unit not in result and Path(document.original_filename).suffix.lower() == ".pdf":
            result[unit] = document
    return result


async def _ensure_hr_access(user: User, session: AsyncSession) -> None:
    roles = await role_keys_for_user(session, user.id)
    if roles.intersection({"owner", "super_admin"}):
        return
    department = await session.get(Department, user.department_id) if user.department_id else None
    if not department_matches(department, "hr"):
        raise HTTPException(status_code=403, detail="Salary slips are restricted to HR administrators.")


def _recipient_response(item: PayrollRecipient) -> dict:
    return {
        "id": str(item.id), "row_number": item.row_number, "employee_name": item.employee_name,
        "employee_code": item.employee_code, "personal_email": item.personal_email,
        "unit": item.details_json.get("unit", ""), "unit_address": item.details_json.get("unit_address", ""),
        "password_hint": password_for(item.employee_name, item.birth_year), "status": item.status,
        "attempt_count": item.attempt_count, "error_message": item.error_message,
        "sent_at": item.sent_at.isoformat() if item.sent_at else None,
        "gross": item.details_json.get("gross", "0.00"), "deductions": item.details_json.get("deduction_total", "0.00"),
        "net_wages": item.details_json.get("net_wages", "0.00"),
        "template_name": item.details_json.get("template_name", ""),
    }


async def _batch_response(session: AsyncSession, batch: PayrollBatch, include_recipients: bool = True) -> dict:
    result = {
        "id": str(batch.id), "payroll_month": batch.payroll_month, "original_filename": batch.original_filename,
        "status": batch.status, "total_count": batch.total_count, "sent_count": batch.sent_count,
        "failed_count": batch.failed_count, "pending_count": max(batch.total_count - batch.sent_count - batch.failed_count, 0),
        "created_at": batch.created_at.isoformat(), "completed_at": batch.completed_at.isoformat() if batch.completed_at else None,
        "template_name": "Selected automatically by unit",
        "email_subject": batch.email_subject or DEFAULT_EMAIL_SUBJECT,
        "email_body": batch.email_body or DEFAULT_EMAIL_BODY,
        "duplicate_email_count": batch.duplicate_email_count,
    }
    if include_recipients:
        recipients = list(await session.scalars(select(PayrollRecipient).where(PayrollRecipient.batch_id == batch.id).order_by(PayrollRecipient.row_number)))
        result["sent_count"] = sum(item.status == "sent" for item in recipients)
        result["failed_count"] = sum(item.status == "failed" for item in recipients)
        result["pending_count"] = sum(item.status in {"pending", "sending"} for item in recipients)
        result["recipients"] = [_recipient_response(item) for item in recipients]
    return result


def _template_response(template: PayrollTemplate) -> dict:
    return {"id": str(template.id), "name": template.name, "original_filename": template.original_filename, "is_active": template.is_active, "created_at": template.created_at.isoformat(), "unit_number": _unit_from_template_name(template.original_filename), "source": "legacy"}


def _knowledge_template_response(unit: int, document: KnowledgeDocument) -> dict:
    path = Path(get_settings().upload_storage_path) / document.stored_filename
    fields = salary_template_form_fields(path) if path.is_file() else []
    return {"id": str(document.id), "name": f"Unit {unit}", "original_filename": document.original_filename, "is_active": True, "created_at": document.created_at.isoformat(), "unit_number": unit, "source": "Human Resources knowledge", "detected_fields": fields, "supports_dynamic_fields": bool(fields)}


@router.get("/templates")
async def list_templates(
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> list[dict]:
    await _ensure_hr_access(user, session)
    templates = await _knowledge_unit_templates(session, user.organization_id)
    return [_knowledge_template_response(unit, document) for unit, document in sorted(templates.items())]


@router.post("/templates")
async def upload_template(
    template_name: str = Form(""),
    unit_number: int | None = Form(None),
    template_file: UploadFile = File(...),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    name = template_name.strip()
    inferred_unit = unit_number or _unit_from_template_name(template_file.filename or "") or _unit_from_template_name(name)
    if inferred_unit not in {1, 2, 3}:
        raise HTTPException(status_code=422, detail="Select Unit 1, Unit 2 or Unit 3 for this salary-slip template.")
    if Path(template_file.filename or "").suffix.lower() != ".pdf":
        raise HTTPException(status_code=422, detail="Export the Canva template as an A4 portrait PDF before uploading.")
    content = await template_file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="The salary-slip template must be smaller than 10 MB.")
    try:
        validate_template_pdf(content)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error
    collection = await session.scalar(select(KnowledgeCollection).where(
        KnowledgeCollection.organization_id == user.organization_id,
        KnowledgeCollection.slug == "hr",
        KnowledgeCollection.status == "active",
    ))
    if not collection:
        raise HTTPException(status_code=422, detail="The Human Resources knowledge collection is unavailable.")
    template_id = uuid.uuid4()
    stored_name = f"payroll-templates/{user.organization_id}/unit-{inferred_unit}/{template_id}.pdf"
    path = Path(get_settings().upload_storage_path) / stored_name
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    current_documents = list((await session.scalars(select(KnowledgeDocument).where(
        KnowledgeDocument.organization_id == user.organization_id,
        KnowledgeDocument.collection_id == collection.id,
        KnowledgeDocument.status == "ready",
    ))))
    previous = [item for item in current_documents if _unit_from_template_name(item.original_filename) == inferred_unit]
    for item in previous:
        item.status = "superseded"
    canonical_name = f"UNIT-{inferred_unit}_SalarySlip.pdf"
    template = KnowledgeDocument(
        id=template_id,
        organization_id=user.organization_id,
        collection_id=collection.id,
        uploaded_by_user_id=user.id,
        original_filename=canonical_name,
        stored_filename=stored_name,
        mime_type="application/pdf",
        size_bytes=len(content),
        version=max((item.version for item in previous), default=0) + 1,
        status="ready",
        extracted_text="",
        extracted_characters=0,
        processed_at=datetime.now(timezone.utc),
        document_category="salary_slip_template",
    )
    session.add(template)
    detected_fields = salary_template_form_fields(content)
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="payroll.template_uploaded", target_type="knowledge_document", target_id=str(template_id), metadata_json={"name": name or f"Unit {inferred_unit}", "filename": canonical_name, "unit": inferred_unit, "detected_fields": detected_fields}))
    await session.commit()
    await session.refresh(template)
    return _knowledge_template_response(inferred_unit, template)


@router.post("/templates/{template_id}/activate")
async def activate_template(
    template_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    template = await session.get(PayrollTemplate, template_id)
    if not template or template.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Salary-slip template not found.")
    await session.execute(update(PayrollTemplate).where(PayrollTemplate.organization_id == user.organization_id).values(is_active=False))
    template.is_active = True
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="payroll.template_activated", target_type="payroll_template", target_id=str(template_id), metadata_json={"name": template.name}))
    await session.commit()
    await session.refresh(template)
    return _template_response(template)


@router.get("/templates/{template_id}/content")
async def template_content(
    template_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> FileResponse:
    await _ensure_hr_access(user, session)
    document = await session.get(KnowledgeDocument, template_id)
    if document and document.organization_id == user.organization_id and _unit_from_template_name(document.original_filename):
        stored_filename, original_filename = document.stored_filename, document.original_filename
    else:
        template = await session.get(PayrollTemplate, template_id)
        if not template or template.organization_id != user.organization_id:
            raise HTTPException(status_code=404, detail="Salary-slip template not found.")
        stored_filename, original_filename = template.stored_filename, template.original_filename
    path = (Path(get_settings().upload_storage_path) / stored_filename).resolve()
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Salary-slip template file is unavailable.")
    return FileResponse(path, media_type="application/pdf", filename=original_filename, content_disposition_type="inline")


@router.get("/template")
async def salary_excel_template(
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    await _ensure_hr_access(user, session)
    return StreamingResponse(io.BytesIO(create_excel_template()), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=AROMAZEN_Salary_Upload_Template.xlsx"})


@router.post("/batches")
async def create_batch(
    payroll_month: str = Form(...),
    excel_file: UploadFile = File(...),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    if not re.fullmatch(r"\d{4}-(?:0[1-9]|1[0-2])", payroll_month):
        raise HTTPException(status_code=422, detail="Payroll month must use YYYY-MM format.")
    if Path(excel_file.filename or "").suffix.lower() != ".xlsx":
        raise HTTPException(status_code=422, detail="Upload the completed .xlsx salary template.")
    content = await excel_file.read()
    if len(content) > get_settings().max_upload_size_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="The salary workbook is too large.")
    try:
        employee_rows = read_salary_excel(content)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error
    unit_templates = await _knowledge_unit_templates(session, user.organization_id)
    used_units = sorted({int(item["details"]["unit"]) for item in employee_rows})
    missing_units = [unit for unit in used_units if unit not in unit_templates]
    if missing_units:
        names = ", ".join(f"UNIT-{unit}_SalarySlip.pdf" for unit in missing_units)
        raise HTTPException(status_code=422, detail=f"Upload the missing template(s) to HR Policies: {names}.")
    duplicate_email_count = sum(count - 1 for count in Counter(item["personal_email"] for item in employee_rows).values() if count > 1)
    batch_id = uuid.uuid4()
    folder = Path(get_settings().upload_storage_path) / "payroll" / str(user.organization_id) / str(batch_id)
    folder.mkdir(parents=True, exist_ok=True)
    workbook_name = f"payroll/{user.organization_id}/{batch_id}/source.xlsx"
    (Path(get_settings().upload_storage_path) / workbook_name).write_bytes(content)
    batch = PayrollBatch(id=batch_id, organization_id=user.organization_id, created_by_user_id=user.id, template_id=None, payroll_month=payroll_month, original_filename=excel_file.filename or "salary.xlsx", stored_filename=workbook_name, email_subject=DEFAULT_EMAIL_SUBJECT, email_body=DEFAULT_EMAIL_BODY, duplicate_email_count=duplicate_email_count, total_count=len(employee_rows), status="draft")
    session.add(batch)
    for item in employee_rows:
        recipient_id = uuid.uuid4()
        pdf_name = f"payroll/{user.organization_id}/{batch_id}/{recipient_id}.pdf"
        original_name = f"Salary_Slip_{payroll_month}_{re.sub(r'[^A-Za-z0-9_-]', '_', item['employee_code'])}.pdf"
        unit = int(item["details"]["unit"])
        template = unit_templates[unit]
        item["details"]["template_name"] = template.original_filename
        template_path = Path(get_settings().upload_storage_path) / template.stored_filename
        generate_salary_pdf(item["details"], payroll_month, Path(get_settings().upload_storage_path) / pdf_name, password_for(item["employee_name"], item["birth_year"]), template_path)
        session.add(PayrollRecipient(id=recipient_id, batch_id=batch_id, organization_id=user.organization_id, row_number=item["row_number"], employee_name=item["employee_name"], employee_code=item["employee_code"], personal_email=item["personal_email"], birth_year=item["birth_year"], details_json=item["details"], pdf_stored_filename=pdf_name, pdf_original_filename=original_name, status="pending"))
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="payroll.batch_created", target_type="payroll_batch", target_id=str(batch_id), metadata_json={"payroll_month": payroll_month, "employee_count": len(employee_rows), "units": used_units, "duplicate_emails": duplicate_email_count}))
    await session.commit()
    await session.refresh(batch)
    return await _batch_response(session, batch)


@router.get("/batches")
async def list_batches(
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> list[dict]:
    await _ensure_hr_access(user, session)
    batches = await session.scalars(select(PayrollBatch).where(PayrollBatch.organization_id == user.organization_id).order_by(PayrollBatch.created_at.desc()).limit(24))
    return [await _batch_response(session, batch, False) for batch in batches]


@router.get("/batches/{batch_id}")
async def get_batch(
    batch_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    batch = await session.get(PayrollBatch, batch_id)
    if not batch or batch.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Payroll batch not found.")
    return await _batch_response(session, batch)


@router.patch("/batches/{batch_id}/email")
async def update_batch_email(
    batch_id: uuid.UUID,
    payload: EmailDraftUpdate,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    batch = await session.get(PayrollBatch, batch_id)
    if not batch or batch.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Payroll batch not found.")
    if batch.status == "sending":
        raise HTTPException(status_code=409, detail="The email cannot be changed while delivery is running.")
    subject, body = payload.subject.strip(), payload.body.strip()
    if not subject or len(subject) > 240:
        raise HTTPException(status_code=422, detail="Email subject is required and must be under 240 characters.")
    if not body or len(body) > 8000:
        raise HTTPException(status_code=422, detail="Email body is required and must be under 8,000 characters.")
    batch.email_subject = subject
    batch.email_body = body
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="payroll.email_updated", target_type="payroll_batch", target_id=str(batch_id), metadata_json={}))
    await session.commit()
    await session.refresh(batch)
    return await _batch_response(session, batch)


@router.get("/batches/{batch_id}/recipients/{recipient_id}/pdf")
async def download_salary_pdf(
    batch_id: uuid.UUID,
    recipient_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> FileResponse:
    await _ensure_hr_access(user, session)
    item = await session.get(PayrollRecipient, recipient_id)
    if not item or item.batch_id != batch_id or item.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Salary slip not found.")
    path = (Path(get_settings().upload_storage_path) / item.pdf_stored_filename).resolve()
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Generated salary slip is unavailable.")
    return FileResponse(path, media_type="application/pdf", filename=item.pdf_original_filename)


def _render_email(value: str, item: PayrollRecipient, month_label: str) -> str:
    return value.replace("{employee_name}", item.employee_name).replace("{month}", month_label)


def _send_message(item: PayrollRecipient, batch: PayrollBatch) -> None:
    settings = get_settings()
    username = settings.zoho_smtp_username
    password = settings.zoho_smtp_password
    from_email = settings.zoho_from_email or username
    if not username or not password or not from_email:
        raise RuntimeError("Zoho Mail is not configured.")
    month_label = datetime.strptime(batch.payroll_month, "%Y-%m").strftime("%B %Y")
    message = EmailMessage()
    message["From"] = formataddr((settings.zoho_from_name, from_email))
    message["To"] = item.personal_email
    message["Subject"] = _render_email(batch.email_subject or DEFAULT_EMAIL_SUBJECT, item, month_label)
    message.set_content(_render_email(batch.email_body or DEFAULT_EMAIL_BODY, item, month_label))
    path = Path(settings.upload_storage_path) / item.pdf_stored_filename
    message.add_attachment(path.read_bytes(), maintype="application", subtype="pdf", filename=item.pdf_original_filename)
    security = settings.zoho_smtp_security.strip().lower()
    if security not in {"ssl", "starttls"}:
        raise RuntimeError("Unsupported Zoho SMTP security mode.")
    smtp_client = smtplib.SMTP_SSL if security == "ssl" else smtplib.SMTP
    with smtp_client(settings.zoho_smtp_host, settings.zoho_smtp_port, timeout=45) as smtp:
        smtp.ehlo()
        if security == "starttls":
            smtp.starttls()
            smtp.ehlo()
        smtp.login(username, password)
        smtp.send_message(message, from_addr=from_email, to_addrs=[item.personal_email])


async def _deliver_batch(batch_id: uuid.UUID, recipient_ids: list[uuid.UUID]) -> None:
    for recipient_id in recipient_ids:
        async with SessionLocal() as session:
            batch = await session.get(PayrollBatch, batch_id)
            item = await session.get(PayrollRecipient, recipient_id)
            if not batch or not item or item.batch_id != batch_id:
                continue
            item.status = "sending"
            item.attempt_count += 1
            item.error_message = None
            await session.commit()
            try:
                await run_in_threadpool(_send_message, item, batch)
                item.status = "sent"
                item.sent_at = datetime.now(timezone.utc)
            except Exception as error:
                item.status = "failed"
                item.error_message = f"{type(error).__name__}: {str(error)[:850]}"
            await session.commit()
    async with SessionLocal() as session:
        batch = await session.get(PayrollBatch, batch_id)
        if not batch:
            return
        sent = await session.scalar(select(func.count()).select_from(PayrollRecipient).where(PayrollRecipient.batch_id == batch_id, PayrollRecipient.status == "sent")) or 0
        failed = await session.scalar(select(func.count()).select_from(PayrollRecipient).where(PayrollRecipient.batch_id == batch_id, PayrollRecipient.status == "failed")) or 0
        batch.sent_count = sent
        batch.failed_count = failed
        batch.status = "completed" if sent == batch.total_count else "failed" if failed == batch.total_count else "partial"
        batch.completed_at = datetime.now(timezone.utc)
        await session.commit()


async def _queue_delivery(batch_id: uuid.UUID, retry_failed: bool, background_tasks: BackgroundTasks, user: User, session: AsyncSession) -> dict:
    batch = await session.get(PayrollBatch, batch_id)
    if not batch or batch.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Payroll batch not found.")
    settings = get_settings()
    if not settings.zoho_smtp_username or not settings.zoho_smtp_password or not (settings.zoho_from_email or settings.zoho_smtp_username):
        raise HTTPException(status_code=503, detail="The HR Zoho Mail account is not configured on the server.")
    if batch.status == "sending":
        raise HTTPException(status_code=409, detail="This payroll batch is already being sent.")
    target_status = "failed" if retry_failed else "pending"
    recipients = list(await session.scalars(select(PayrollRecipient).where(PayrollRecipient.batch_id == batch_id, PayrollRecipient.status == target_status).order_by(PayrollRecipient.row_number)))
    if not recipients:
        raise HTTPException(status_code=409, detail="There are no failed deliveries to retry." if retry_failed else "There are no pending salary slips to send.")
    batch.status = "sending"
    batch.sending_started_at = datetime.now(timezone.utc)
    batch.completed_at = None
    if retry_failed:
        for recipient in recipients:
            recipient.status = "pending"
            recipient.error_message = None
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="payroll.failed_retried" if retry_failed else "payroll.batch_send_started", target_type="payroll_batch", target_id=str(batch_id), metadata_json={"recipient_count": len(recipients)}))
    await session.commit()
    background_tasks.add_task(_deliver_batch, batch_id, [item.id for item in recipients])
    return await _batch_response(session, batch)


@router.post("/batches/{batch_id}/send")
async def send_batch(batch_id: uuid.UUID, background_tasks: BackgroundTasks, user: User = Depends(require_permissions("users.manage")), session: AsyncSession = Depends(get_db_session)) -> dict:
    await _ensure_hr_access(user, session)
    return await _queue_delivery(batch_id, False, background_tasks, user, session)


@router.post("/batches/{batch_id}/retry-failed")
async def retry_failed(batch_id: uuid.UUID, background_tasks: BackgroundTasks, user: User = Depends(require_permissions("users.manage")), session: AsyncSession = Depends(get_db_session)) -> dict:
    await _ensure_hr_access(user, session)
    return await _queue_delivery(batch_id, True, background_tasks, user, session)


def _attendance_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _attendance_clock(value: object) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    text = str(value or "").strip()
    for pattern in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
        try:
            return datetime.strptime(text, pattern).time()
        except ValueError:
            pass
    return None


def _attendance_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


ATTENDANCE_STATUS_LABELS = {
    "P": "Present",
    "LT": "Late",
    "EL": "Early leave",
    "A": "Absent",
    "WO": "Weekly off",
    "HD": "Half day",
}
DEFAULT_ATTENDANCE_SHIFTS = [
    {"name": "Shift 1", "start": "06:00", "end": "14:00", "grace_minutes": DEFAULT_LATE_GRACE_MINUTES},
    {"name": "General", "start": "09:00", "end": "17:30", "grace_minutes": DEFAULT_LATE_GRACE_MINUTES},
    {"name": "Shift 2", "start": "14:00", "end": "22:00", "grace_minutes": DEFAULT_LATE_GRACE_MINUTES},
    {"name": "Shift 3", "start": "22:00", "end": "06:00", "grace_minutes": DEFAULT_LATE_GRACE_MINUTES},
]


def _attendance_hours(value: object) -> float:
    if isinstance(value, timedelta):
        return round(value.total_seconds() / 3600, 2)
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return round(value.hour + value.minute / 60 + value.second / 3600, 2)
    if isinstance(value, (int, float)) and 0 <= float(value) < 2:
        return round(float(value) * 24, 2)
    parsed = _attendance_clock(value)
    return round(parsed.hour + parsed.minute / 60, 2) if parsed else 0.0


def _parse_matrix_attendance(workbook) -> list[dict]:
    records: list[dict] = []
    employee_pattern = re.compile(
        r"Employee Name\s*:\s*(.*?)\s*,\s*Employee ID\s*:\s*(.*?)\s*,\s*Gender\s*:\s*(.*?)\s*,\s*Department\s*:\s*(.*?)\s*,\s*Position\s*:\s*(.*)",
        re.IGNORECASE,
    )
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        heading = " ".join(str(cell or "") for row in rows[:6] for cell in row)
        period_match = re.search(r"From\s+([A-Za-z]+)\s+\d{1,2}\s+(\d{4})", heading, re.IGNORECASE)
        if period_match:
            try:
                month_number = datetime.strptime(period_match.group(1), "%B").month
            except ValueError:
                month_number = datetime.strptime(period_match.group(1)[:3], "%b").month
            year_number = int(period_match.group(2))
        else:
            sheet_date = re.search(r"(20\d{2})(\d{2})(\d{2})", sheet.title)
            if not sheet_date:
                continue
            year_number, month_number = int(sheet_date.group(1)), int(sheet_date.group(2))
        month_days = monthrange(year_number, month_number)[1]
        for row_index, row in enumerate(rows):
            metadata = employee_pattern.search(str(row[0] or "")) if row else None
            if not metadata or row_index + 4 >= len(rows):
                continue
            employee_name, employee_code, _gender, department, position = (part.strip() for part in metadata.groups())
            header, status_values, in_values, out_values, total_values = rows[row_index + 1:row_index + 6]
            if _attendance_key(status_values[1] if len(status_values) > 1 else "") != "status":
                continue
            for column in range(2, len(header)):
                day_match = re.match(r"\s*(\d{1,2})", str(header[column] or ""))
                if not day_match:
                    continue
                day_number = int(day_match.group(1))
                if day_number < 1 or day_number > month_days:
                    continue
                status_code = str(status_values[column] or "").strip().upper() if column < len(status_values) else ""
                first_in = _attendance_clock(in_values[column] if column < len(in_values) else None)
                last_out = _attendance_clock(out_values[column] if column < len(out_values) else None)
                worked_hours = _attendance_hours(total_values[column] if column < len(total_values) else None)
                if not any((status_code, first_in, last_out, worked_hours)):
                    continue
                records.append({
                    "employee_code": employee_code,
                    "employee_name": employee_name,
                    "department": department or "Unassigned",
                    "position": position,
                    "date": date(year_number, month_number, day_number),
                    "first_in": first_in,
                    "last_out": last_out,
                    "worked_hours": worked_hours,
                    "status_code": status_code,
                })
    return records


def _parse_tabular_attendance(workbook) -> list[dict]:
    aliases = {
        "code": {"employeeid", "employeecode", "empid", "empcode", "userid", "user id", "id"},
        "name": {"employeename", "empname", "name", "employee"},
        "department": {"department", "dept"},
        "date": {"date", "attendancedate", "punchdate", "transactiondate"},
        "in": {"intime", "checkin", "firstin", "timein", "in"},
        "out": {"outtime", "checkout", "lastout", "timeout", "out"},
        "punch": {"punchtime", "time", "transactiontime", "datetime", "punchdatetime"},
    }
    groups: dict[tuple[str, date], dict] = {}
    all_aliases = set().union(*aliases.values())
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next((index for index, row in enumerate(rows[:20]) if sum(_attendance_key(cell) in all_aliases for cell in row) >= 2), None)
        if header_index is None:
            continue
        header = [_attendance_key(value) for value in rows[header_index]]
        columns = {kind: next((index for index, value in enumerate(header) if value in {_attendance_key(item) for item in names}), None) for kind, names in aliases.items()}
        if columns["date"] is None and columns["punch"] is None:
            continue
        for values in rows[header_index + 1:]:
            def cell(kind: str):
                index = columns[kind]
                return values[index] if index is not None and index < len(values) else None
            punch_value = cell("punch")
            work_date = _attendance_date(cell("date")) or _attendance_date(punch_value)
            if not work_date:
                continue
            code = str(cell("code") or "").strip()
            name = str(cell("name") or code or "Unknown employee").strip()
            identity = code or _attendance_key(name)
            group = groups.setdefault((identity, work_date), {"employee_code": code, "employee_name": name, "department": str(cell("department") or "Unassigned").strip(), "date": work_date, "punches": [], "in": None, "out": None})
            explicit_in, explicit_out = _attendance_clock(cell("in")), _attendance_clock(cell("out"))
            punch = _attendance_clock(punch_value)
            if explicit_in:
                group["in"] = explicit_in
            if explicit_out:
                group["out"] = explicit_out
            if punch:
                group["punches"].append(punch)
    records = []
    for group in groups.values():
        punches = sorted(group["punches"])
        first_in = group["in"] or (punches[0] if punches else None)
        last_out = group["out"] or (punches[-1] if len(punches) > 1 else None)
        records.append({**group, "first_in": first_in, "last_out": last_out, "worked_hours": 0.0, "status_code": "P"})
    return records


def _time_minutes(value: time) -> int:
    return value.hour * 60 + value.minute


def _shift_duration(start: time, end: time) -> float:
    minutes = (_time_minutes(end) - _time_minutes(start)) % (24 * 60)
    return round(minutes / 60, 2)


def _nearest_shift(clock: time, shifts: list[dict]) -> dict:
    clock_minutes = _time_minutes(clock)
    return min(shifts, key=lambda shift: min(abs(clock_minutes - shift["start_minutes"]), 1440 - abs(clock_minutes - shift["start_minutes"])))


def _parse_shift_roster(content: bytes, filename: str) -> list[dict]:
    extension = Path(filename).suffix.lower()
    rows: list[tuple] = []
    if extension == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        rows = [tuple(row) for row in csv.reader(io.StringIO(text))]
    elif extension in {".xlsx", ".xlsm"}:
        roster_workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            for sheet in roster_workbook.worksheets:
                rows.extend(tuple(row) for row in sheet.iter_rows(values_only=True))
        finally:
            roster_workbook.close()
    else:
        raise HTTPException(status_code=422, detail="Upload the shift roster as .xlsx or .csv.")
    aliases = {
        "code": {"employeeid", "employeecode", "empid", "empcode"},
        "name": {"employeename", "employee", "name"},
        "from": {"fromdate", "fromdateyyyymmdd", "startdate", "effectivefrom"},
        "to": {"todate", "todateyyyymmdd", "enddate", "effectiveto"},
        "shift": {"shiftname", "shift"},
    }
    header_index = next((index for index, row in enumerate(rows[:20]) if sum(_attendance_key(cell) in set().union(*aliases.values()) for cell in row) >= 3), None)
    if header_index is None:
        raise HTTPException(status_code=422, detail="The shift roster needs Employee ID/Name, From Date, To Date and Shift Name columns.")
    header = [_attendance_key(value) for value in rows[header_index]]
    columns = {kind: next((index for index, value in enumerate(header) if value in {_attendance_key(item) for item in names}), None) for kind, names in aliases.items()}
    if columns["shift"] is None or columns["from"] is None or (columns["code"] is None and columns["name"] is None):
        raise HTTPException(status_code=422, detail="The shift roster needs Employee ID or Employee Name, From Date and Shift Name.")
    assignments = []
    for row in rows[header_index + 1:]:
        def cell(kind: str):
            index = columns[kind]
            return row[index] if index is not None and index < len(row) else None
        code = str(cell("code") or "").strip()
        name = str(cell("name") or "").strip()
        start_date = _attendance_date(cell("from"))
        end_date = _attendance_date(cell("to")) or start_date
        shift_name = str(cell("shift") or "").strip()
        if not shift_name or not start_date or (not code and not name):
            continue
        assignments.append({"employee_code": code, "employee_name_key": _attendance_key(name), "from_date": start_date, "to_date": end_date or start_date, "shift_name": shift_name})
    if not assignments:
        raise HTTPException(status_code=422, detail="No valid shift assignments were found in the roster.")
    return assignments


def _roster_shift(record: dict, assignments: list[dict], valid_shift_names: set[str]) -> str | None:
    record_code = str(record.get("employee_code") or "").strip()
    record_name = _attendance_key(record.get("employee_name"))
    work_date = record["date"]
    for assignment in assignments:
        identity_matches = (record_code and assignment["employee_code"] == record_code) or (record_name and assignment["employee_name_key"] == record_name)
        if identity_matches and assignment["from_date"] <= work_date <= assignment["to_date"] and assignment["shift_name"] in valid_shift_names:
            return assignment["shift_name"]
    return None

def _attendance_analysis(records: list[dict], shifts: list[dict], filename: str, assignments: list[dict] | None = None) -> dict:
    if not records:
        raise HTTPException(status_code=422, detail="No attendance rows were recognized in this workbook.")
    assignments = assignments or []
    shifts_by_name = {shift["name"]: shift for shift in shifts}
    valid_shift_names = set(shifts_by_name)
    employee_shift_counts: dict[str, Counter] = {}
    for record in records:
        key = record["employee_code"] or record["employee_name"]
        roster_shift = _roster_shift(record, assignments, valid_shift_names)
        if roster_shift:
            record["shift_name"] = roster_shift
            record["assignment_source"] = "Roster"
        elif record.get("first_in"):
            shift = _nearest_shift(record["first_in"], shifts)
            record["shift_name"] = shift["name"]
            record["assignment_source"] = "Automatic"
        if record.get("shift_name"):
            employee_shift_counts.setdefault(key, Counter())[record["shift_name"]] += 1
    primary_shifts = {key: counts.most_common(1)[0][0] for key, counts in employee_shift_counts.items() if counts}
    employees: dict[str, dict] = {}
    shift_summary: dict[str, dict] = {}
    department_summary: dict[str, dict] = {}
    normalized_records = []
    status_totals = Counter()
    monthly_late_counts = Counter()
    for record in sorted(records, key=lambda item: (item["date"], item["employee_name"])):
        key = record["employee_code"] or record["employee_name"]
        shift_name = record.get("shift_name") or primary_shifts.get(key) or "Unassigned"
        shift = shifts_by_name.get(shift_name)
        first_in, last_out = record.get("first_in"), record.get("last_out")
        worked_hours = float(record.get("worked_hours") or 0)
        if not worked_hours and first_in and last_out:
            start_dt = datetime.combine(record["date"], first_in)
            end_dt = datetime.combine(record["date"], last_out)
            if end_dt < start_dt:
                end_dt += timedelta(days=1)
            worked_hours = round((end_dt - start_dt).total_seconds() / 3600, 2)
        raw_status_code = str(record.get("status_code") or "").upper()
        computed_late = False
        if shift and first_in:
            computed_late = _time_minutes(first_in) > shift["start_minutes"] + shift["grace_minutes"]
        detected_late = computed_late if shift and first_in else raw_status_code == "LT"
        late_policy = apply_monthly_late_policy(monthly_late_counts, key, record["date"], raw_status_code, detected_late)
        status_code = late_policy.status_code
        is_late = late_policy.is_late
        is_early = raw_status_code == "EL"
        late_penalty_half_day = late_policy.half_day_penalty
        status_label = "Half day (4th late)" if late_penalty_half_day else ATTENDANCE_STATUS_LABELS.get(status_code, status_code.title() if status_code else "Present")
        scheduled_hours = shift["duration_hours"] if shift else 0
        overtime_hours = round(max(0.0, worked_hours - scheduled_hours), 2) if scheduled_hours else 0.0
        scheduled = status_code != "WO"
        present_value = 0.5 if status_code == "HD" else (1.0 if status_code in {"P", "LT", "EL"} else 0.0)
        status_totals[status_label] += 1
        item = employees.setdefault(key, {"employee_code": record["employee_code"], "employee_name": record["employee_name"], "department": record.get("department") or "Unassigned", "primary_shift": primary_shifts.get(key, "Unassigned"), "scheduled_days": 0, "present_days": 0.0, "absent_days": 0, "weekly_off_days": 0, "late_days": 0, "early_leave_days": 0, "half_days": 0, "late_penalty_half_days": 0, "total_hours": 0.0, "overtime_hours": 0.0, "days_with_hours": 0})
        item["scheduled_days"] += int(scheduled)
        item["present_days"] += present_value
        item["absent_days"] += int(status_code == "A")
        item["weekly_off_days"] += int(status_code == "WO")
        item["late_days"] += int(is_late)
        item["early_leave_days"] += int(is_early)
        item["half_days"] += int(status_code == "HD")
        item["late_penalty_half_days"] += int(late_penalty_half_day)
        item["total_hours"] = round(item["total_hours"] + worked_hours, 2)
        item["overtime_hours"] = round(item["overtime_hours"] + overtime_hours, 2)
        item["days_with_hours"] += int(worked_hours > 0)
        for summary, summary_key in ((shift_summary, shift_name), (department_summary, item["department"])):
            group = summary.setdefault(summary_key, {"name": summary_key, "employee_ids": set(), "scheduled_days": 0, "present_days": 0.0, "absent_days": 0, "late_days": 0, "early_leave_days": 0, "late_penalty_half_days": 0, "total_hours": 0.0, "overtime_hours": 0.0})
            group["employee_ids"].add(key)
            group["scheduled_days"] += int(scheduled)
            group["present_days"] += present_value
            group["absent_days"] += int(status_code == "A")
            group["late_days"] += int(is_late)
            group["early_leave_days"] += int(is_early)
            group["late_penalty_half_days"] += int(late_penalty_half_day)
            group["total_hours"] = round(group["total_hours"] + worked_hours, 2)
            group["overtime_hours"] = round(group["overtime_hours"] + overtime_hours, 2)
        normalized_records.append({"employee_code": record["employee_code"], "employee_name": record["employee_name"], "department": record.get("department") or "Unassigned", "date": record["date"].isoformat(), "shift_name": shift_name, "first_in": first_in.strftime("%H:%M") if first_in else "", "last_out": last_out.strftime("%H:%M") if last_out else "", "worked_hours": worked_hours, "overtime_hours": overtime_hours, "status_code": status_code, "status": status_label, "late_penalty_half_day": late_penalty_half_day, "assignment_source": record.get("assignment_source") or ("Primary shift" if shift_name != "Unassigned" else "Unassigned")})
    employee_items = []
    for item in employees.values():
        days_with_hours = item.pop("days_with_hours")
        item["average_hours"] = round(item["total_hours"] / days_with_hours, 2) if days_with_hours else 0
        item["attendance_rate"] = round(item["present_days"] / item["scheduled_days"] * 100, 1) if item["scheduled_days"] else 0
        employee_items.append(item)
    employee_items.sort(key=lambda item: (item["department"], item["employee_name"]))
    def finish_groups(groups: dict[str, dict]) -> list[dict]:
        finished = []
        for group in groups.values():
            group["employee_count"] = len(group.pop("employee_ids"))
            group["attendance_rate"] = round(group["present_days"] / group["scheduled_days"] * 100, 1) if group["scheduled_days"] else 0
            group["average_hours"] = round(group["total_hours"] / group["present_days"], 2) if group["present_days"] else 0
            finished.append(group)
        return sorted(finished, key=lambda item: item["name"])
    dates = [record["date"] for record in records]
    return {
        "filename": filename,
        "period": {"from": min(dates).isoformat(), "to": max(dates).isoformat()},
        "employee_count": len(employee_items),
        "record_count": len(normalized_records),
        "roster_assigned_records": sum(1 for record in normalized_records if record["assignment_source"] == "Roster"),
        "automatic_assigned_records": sum(1 for record in normalized_records if record["assignment_source"] == "Automatic"),
        "scheduled_days": sum(item["scheduled_days"] for item in employee_items),
        "present_days": round(sum(item["present_days"] for item in employee_items), 1),
        "absent_days": sum(item["absent_days"] for item in employee_items),
        "weekly_off_days": sum(item["weekly_off_days"] for item in employee_items),
        "late_days": sum(item["late_days"] for item in employee_items),
        "early_leave_days": sum(item["early_leave_days"] for item in employee_items),
        "half_days": sum(item["half_days"] for item in employee_items),
        "late_penalty_half_days": sum(item["late_penalty_half_days"] for item in employee_items),
        "total_hours": round(sum(item["total_hours"] for item in employee_items), 2),
        "overtime_hours": round(sum(item["overtime_hours"] for item in employee_items), 2),
        "status_counts": dict(status_totals),
        "shift_rules": [{"name": shift["name"], "start": shift["start"], "end": shift["end"], "grace_minutes": shift["grace_minutes"]} for shift in shifts],
        "shifts": finish_groups(shift_summary),
        "departments": finish_groups(department_summary),
        "employees": employee_items,
        "records": normalized_records,
    }


def _parse_attendance_shifts(shift_rules: str) -> list[dict]:
    try:
        parsed_shifts = json.loads(shift_rules) if shift_rules.strip() else DEFAULT_ATTENDANCE_SHIFTS
        if not isinstance(parsed_shifts, list) or not parsed_shifts:
            raise ValueError("invalid_shift_rules")
        shifts = []
        for index, item in enumerate(parsed_shifts):
            start = _attendance_clock(item.get("start"))
            end = _attendance_clock(item.get("end"))
            if not start or not end:
                raise ValueError("invalid_shift_time")
            shifts.append({"name": str(item.get("name") or f"Shift {index + 1}").strip(), "start": start.strftime("%H:%M"), "end": end.strftime("%H:%M"), "start_minutes": _time_minutes(start), "end_minutes": _time_minutes(end), "grace_minutes": max(0, int(item.get("grace_minutes") or 0)), "duration_hours": _shift_duration(start, end)})
        return shifts
    except (ValueError, TypeError, json.JSONDecodeError) as error:
        raise HTTPException(status_code=422, detail="Review the shift names, start/end times and grace minutes.") from error


@router.post("/attendance/analyze")
async def analyze_attendance(
    excel_file: UploadFile = File(...),
    shift_start: str = Form("09:00"),
    shift_rules: str = Form(""),
    shift_roster_file: UploadFile | None = File(None),
    user: User = Depends(require_permissions("ai.workspace.use")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    if Path(excel_file.filename or "").suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=422, detail="Upload a fingerprint attendance .xlsx file.")
    content = await excel_file.read()
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise HTTPException(status_code=422, detail="The attendance workbook could not be opened.") from error
    shifts = _parse_attendance_shifts(shift_rules)
    assignments: list[dict] = []
    if shift_roster_file and shift_roster_file.filename:
        assignments = _parse_shift_roster(await shift_roster_file.read(), shift_roster_file.filename)
        unknown_shifts = sorted({item["shift_name"] for item in assignments} - {shift["name"] for shift in shifts})
        if unknown_shifts:
            raise HTTPException(status_code=422, detail=f"These roster shift names do not match the configured shifts: {', '.join(unknown_shifts)}")
    records = _parse_matrix_attendance(workbook)
    if not records:
        records = _parse_tabular_attendance(workbook)
    workbook.close()
    return _attendance_analysis(records, shifts, excel_file.filename or "attendance.xlsx", assignments)


def _salary_merge_workbook(content: bytes):
    try:
        workbook = load_workbook(io.BytesIO(content))
    except Exception as error:
        raise HTTPException(status_code=422, detail="The salary workbook could not be opened.") from error
    sheet = workbook["Salary Data"] if "Salary Data" in workbook.sheetnames else workbook.active
    aliases = {
        "employee_name": {"employeename", "employee", "name"},
        "employee_code": {"employeecode", "empcode", "employeeid", "empid"},
        "days": {"days", "totaldays"},
        "present_days": {"presentdays", "paiddays"},
        "lop": {"lop", "lossofpay"},
        "ot_hours": {"othours", "overtimehours"},
    }
    header = [_attendance_key(cell.value) for cell in sheet[1]]
    indexes = {key: next((index + 1 for index, value in enumerate(header) if value in names), None) for key, names in aliases.items()}
    missing = [dict(COLUMNS).get(key, key.replace("_", " ").title()) for key, index in indexes.items() if index is None]
    if missing:
        workbook.close()
        raise HTTPException(status_code=422, detail="The final salary template is missing: " + ", ".join(missing))
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row_number, indexes["employee_name"]).value or "").strip()
        code = str(sheet.cell(row_number, indexes["employee_code"]).value or "").strip()
        if not name and not code:
            continue
        rows.append({"row_number": row_number, "employee_name": name, "employee_code": code})
    if not rows:
        workbook.close()
        raise HTTPException(status_code=422, detail="No employee rows were found in Salary Data.")
    return workbook, sheet, indexes, rows


def _leave_calculator_analysis(
    salary_content: bytes,
    attendance_content: bytes,
    payroll_month: str,
    shift_rules: str,
    attendance_filename: str,
    salary_filename: str,
    roster_content: bytes | None = None,
    roster_filename: str = "",
):
    try:
        month_date = datetime.strptime(payroll_month, "%Y-%m")
    except ValueError as error:
        raise HTTPException(status_code=422, detail="Payroll month must use YYYY-MM.") from error
    shifts = _parse_attendance_shifts(shift_rules)
    assignments = _parse_shift_roster(roster_content, roster_filename) if roster_content and roster_filename else []
    unknown_shifts = sorted({item["shift_name"] for item in assignments} - {shift["name"] for shift in shifts})
    if unknown_shifts:
        raise HTTPException(status_code=422, detail=f"These roster shift names do not match the configured shifts: {', '.join(unknown_shifts)}")
    try:
        attendance_workbook = load_workbook(io.BytesIO(attendance_content), data_only=True, read_only=True)
    except Exception as error:
        raise HTTPException(status_code=422, detail="The attendance workbook could not be opened.") from error
    records = _parse_matrix_attendance(attendance_workbook)
    if not records:
        records = _parse_tabular_attendance(attendance_workbook)
    attendance_workbook.close()
    records = [item for item in records if item["date"].year == month_date.year and item["date"].month == month_date.month]
    if not records:
        raise HTTPException(status_code=422, detail=f"No attendance records were found for {month_date.strftime('%B %Y')}.")
    attendance = _attendance_analysis(records, shifts, attendance_filename, assignments)
    workbook, salary_sheet, indexes, salary_rows = _salary_merge_workbook(salary_content)
    by_code: dict[str, list[dict]] = {}
    by_name: dict[str, list[dict]] = {}
    for employee in attendance["employees"]:
        code_key = _attendance_key(employee["employee_code"])
        name_key = _attendance_key(employee["employee_name"])
        if code_key:
            by_code.setdefault(code_key, []).append(employee)
        if name_key:
            by_name.setdefault(name_key, []).append(employee)
    matched_attendance: set[str] = set()
    calendar_days = monthrange(month_date.year, month_date.month)[1]
    result_rows = []
    for salary_row in salary_rows:
        code_matches = by_code.get(_attendance_key(salary_row["employee_code"]), []) if salary_row["employee_code"] else []
        name_matches = by_name.get(_attendance_key(salary_row["employee_name"]), []) if salary_row["employee_name"] else []
        employee = code_matches[0] if len(code_matches) == 1 else name_matches[0] if len(name_matches) == 1 else None
        match_status = "Matched by employee code" if employee and len(code_matches) == 1 else "Matched by employee name" if employee else "Not found in attendance"
        if employee:
            matched_attendance.add(_attendance_key(employee["employee_code"]) or _attendance_key(employee["employee_name"]))
        calculated_lop = round(float(employee["absent_days"]) + float(employee["half_days"]) * 0.5, 1) if employee else None
        result_rows.append({
            **salary_row,
            "department": employee["department"] if employee else "",
            "primary_shift": employee["primary_shift"] if employee else "",
            "calendar_days": calendar_days,
            "scheduled_days": employee["scheduled_days"] if employee else None,
            "attendance_present_days": employee["present_days"] if employee else None,
            "absent_days": employee["absent_days"] if employee else None,
            "half_days": employee["half_days"] if employee else None,
            "late_penalty_half_days": employee["late_penalty_half_days"] if employee else None,
            "weekly_off_days": employee["weekly_off_days"] if employee else None,
            "paid_leave_days": 0,
            "calculated_lop": calculated_lop,
            "lop_override": None,
            "final_lop": calculated_lop,
            "paid_days": round(calendar_days - calculated_lop, 1) if calculated_lop is not None else None,
            "ot_hours": employee["overtime_hours"] if employee else None,
            "match_status": match_status,
        })
    attendance_only = [employee for employee in attendance["employees"] if (_attendance_key(employee["employee_code"]) or _attendance_key(employee["employee_name"])) not in matched_attendance]
    matched_rows = [item for item in result_rows if item["calculated_lop"] is not None]
    payload = {
        "payroll_month": payroll_month,
        "month_label": month_date.strftime("%B %Y"),
        "calendar_days": calendar_days,
        "salary_filename": salary_filename,
        "attendance_filename": attendance_filename,
        "employee_count": len(result_rows),
        "matched_count": len(matched_rows),
        "unmatched_count": len(result_rows) - len(matched_rows),
        "attendance_only_count": len(attendance_only),
        "total_calculated_lop": round(sum(item["calculated_lop"] or 0 for item in matched_rows), 1),
        "total_late_penalty_half_days": sum(item["late_penalty_half_days"] or 0 for item in matched_rows),
        "total_paid_days": round(sum(item["paid_days"] or 0 for item in matched_rows), 1),
        "total_ot_hours": round(sum(item["ot_hours"] or 0 for item in matched_rows), 2),
        "rows": result_rows,
        "attendance_only": [{"employee_code": item["employee_code"], "employee_name": item["employee_name"], "department": item["department"]} for item in attendance_only],
        "shift_rules": attendance["shift_rules"],
    }
    return payload, workbook, salary_sheet, indexes


def _number(value: object, label: str) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise HTTPException(status_code=422, detail=f"{label} must be a number.") from error
    if number < 0:
        raise HTTPException(status_code=422, detail=f"{label} cannot be negative.")
    return round(number, 2)


def _merged_leave_workbook(analysis: dict, workbook, salary_sheet, indexes: dict, adjustments_json: str) -> bytes:
    try:
        adjustments = json.loads(adjustments_json) if adjustments_json.strip() else []
        adjustment_map = {int(item["row_number"]): item for item in adjustments}
    except (ValueError, TypeError, KeyError, json.JSONDecodeError) as error:
        workbook.close()
        raise HTTPException(status_code=422, detail="The leave adjustments could not be read.") from error
    if analysis["unmatched_count"]:
        workbook.close()
        raise HTTPException(status_code=409, detail="Resolve every unmatched salary employee before downloading the merged payroll file.")
    for item in analysis["rows"]:
        adjustment = adjustment_map.get(item["row_number"], {})
        paid_leave = min(_number(adjustment.get("paid_leave_days", 0), "Paid leave"), item["calculated_lop"])
        override_value = adjustment.get("lop_override")
        final_lop = _number(override_value, "LOP override") if override_value not in (None, "") else round(item["calculated_lop"] - paid_leave, 2)
        if final_lop > analysis["calendar_days"]:
            workbook.close()
            raise HTTPException(status_code=422, detail=f"LOP cannot exceed {analysis['calendar_days']} days.")
        item["paid_leave_days"] = paid_leave
        item["lop_override"] = override_value if override_value not in (None, "") else None
        item["final_lop"] = final_lop
        item["paid_days"] = round(analysis["calendar_days"] - final_lop, 2)
        salary_sheet.cell(item["row_number"], indexes["days"], analysis["calendar_days"])
        salary_sheet.cell(item["row_number"], indexes["present_days"], item["paid_days"])
        salary_sheet.cell(item["row_number"], indexes["lop"], final_lop)
        salary_sheet.cell(item["row_number"], indexes["ot_hours"], item["ot_hours"])
    review_name = "Leave Calculation Review"
    if review_name in workbook.sheetnames:
        workbook.remove(workbook[review_name])
    review = workbook.create_sheet(review_name, 1)
    review.sheet_view.showGridLines = False
    review.merge_cells("A1:P1")
    review["A1"] = f"Employee Leave Calculator · {analysis['month_label']}"
    review["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    review["A1"].fill = PatternFill("solid", fgColor="111827")
    review["A1"].alignment = Alignment(horizontal="left", vertical="center")
    review.row_dimensions[1].height = 30
    review["A3"] = "Employees"
    review["B3"] = analysis["employee_count"]
    review["D3"] = "Final paid days"
    review["E3"] = round(sum(item["paid_days"] for item in analysis["rows"]), 1)
    review["G3"] = "Final LOP"
    review["H3"] = round(sum(item["final_lop"] for item in analysis["rows"]), 1)
    review["J3"] = "OT hours"
    review["K3"] = round(sum(item["ot_hours"] or 0 for item in analysis["rows"]), 2)
    for cell in (review["A3"], review["D3"], review["G3"], review["J3"]):
        cell.font = Font(bold=True, color="6B7280")
    headers = ["Employee Code", "Employee Name", "Department", "Primary Shift", "Calendar Days", "Attendance Present", "Absent", "Half Days", "Late Penalty HD", "Weekly Offs", "Paid Leave", "Calculated LOP", "LOP Override", "Final LOP", "Final Paid Days", "OT Hours"]
    for column, value in enumerate(headers, 1):
        cell = review.cell(6, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_number, item in enumerate(analysis["rows"], 7):
        values = [item["employee_code"], item["employee_name"], item["department"], item["primary_shift"], item["calendar_days"], item["attendance_present_days"], item["absent_days"], item["half_days"], item["late_penalty_half_days"], item["weekly_off_days"], item["paid_leave_days"], item["calculated_lop"], item["lop_override"], item["final_lop"], item["paid_days"], item["ot_hours"]]
        for column, value in enumerate(values, 1):
            review.cell(row_number, column, value)
        if row_number % 2 == 0:
            for column in range(1, len(headers) + 1):
                review.cell(row_number, column).fill = PatternFill("solid", fgColor="F3F4F6")
    widths = [16, 24, 18, 16, 14, 18, 11, 11, 16, 12, 12, 15, 14, 12, 16, 12]
    for column, width in enumerate(widths, 1):
        review.column_dimensions[get_column_letter(column)].width = width
    review.freeze_panes = "A7"
    review.auto_filter.ref = f"A6:P{max(6, 6 + len(analysis['rows']))}"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@router.post("/leave-calculator/analyze")
async def analyze_employee_leaves(
    payroll_month: str = Form(...),
    salary_file: UploadFile = File(...),
    attendance_file: UploadFile = File(...),
    shift_rules: str = Form(""),
    shift_roster_file: UploadFile | None = File(None),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await _ensure_hr_access(user, session)
    if Path(salary_file.filename or "").suffix.lower() not in {".xlsx", ".xlsm"} or Path(attendance_file.filename or "").suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=422, detail="Upload the final salary template and fingerprint attendance as Excel files.")
    salary_content, attendance_content = await salary_file.read(), await attendance_file.read()
    roster_content = await shift_roster_file.read() if shift_roster_file and shift_roster_file.filename else None
    result, workbook, _sheet, _indexes = await run_in_threadpool(_leave_calculator_analysis, salary_content, attendance_content, payroll_month, shift_rules, attendance_file.filename or "attendance.xlsx", salary_file.filename or "salary.xlsx", roster_content, shift_roster_file.filename if shift_roster_file else "")
    workbook.close()
    return result


@router.post("/leave-calculator/merge")
async def merge_employee_leaves(
    payroll_month: str = Form(...),
    salary_file: UploadFile = File(...),
    attendance_file: UploadFile = File(...),
    shift_rules: str = Form(""),
    adjustments_json: str = Form("[]"),
    shift_roster_file: UploadFile | None = File(None),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    await _ensure_hr_access(user, session)
    salary_content, attendance_content = await salary_file.read(), await attendance_file.read()
    roster_content = await shift_roster_file.read() if shift_roster_file and shift_roster_file.filename else None
    analysis, workbook, salary_sheet, indexes = await run_in_threadpool(_leave_calculator_analysis, salary_content, attendance_content, payroll_month, shift_rules, attendance_file.filename or "attendance.xlsx", salary_file.filename or "salary.xlsx", roster_content, shift_roster_file.filename if shift_roster_file else "")
    content = await run_in_threadpool(_merged_leave_workbook, analysis, workbook, salary_sheet, indexes, adjustments_json)
    filename = f"AROMAZEN_Salary_With_Attendance_{payroll_month}.xlsx"
    return StreamingResponse(io.BytesIO(content), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
