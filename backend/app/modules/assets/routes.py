import io
import uuid
from copy import copy
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.db.session import get_db_session
from app.modules.assets.models import AssetMaintenanceEvent, ITAsset
from app.modules.assets.schemas import AssetListResponse, AssetPayload, AssetRead, AssetSummary, AssetUpdatePayload, MaintenancePayload, MaintenanceRead
from app.modules.assets.service import ASSET_TEMPLATE_PATH, asset_key, clean_text, merge_asset_rows, read_asset_rows
from app.modules.identity.authorization import require_permissions
from app.modules.identity.models import AuditEvent, Department, User
from app.modules.identity.service import role_keys_for_user

router = APIRouter()
INACTIVE_MAINTENANCE_STATUSES = {"Scrapped", "Disposed", "Lost"}
SCRAP_QUEUE_STATUSES = {"Scrap proposed", "Approved for scrap"}


async def ensure_asset_access(user: User, session: AsyncSession) -> None:
    roles = await role_keys_for_user(session, user.id)
    if roles.intersection({"owner", "super_admin"}):
        return
    department = await session.get(Department, user.department_id) if user.department_id else None
    if not department or department.slug != "hr":
        raise HTTPException(status_code=403, detail="Asset Management is available to HR administrators only.")


def maintenance_status(asset: ITAsset, today: date | None = None) -> tuple[str, int | None]:
    if not asset.next_maintenance_date or asset.status in INACTIVE_MAINTENANCE_STATUSES:
        return "not_scheduled", None
    current = today or datetime.now(timezone.utc).date()
    remaining = (asset.next_maintenance_date - current).days
    if remaining < 0:
        return "overdue", remaining
    if remaining <= asset.maintenance_reminder_days:
        return "due", remaining
    return "scheduled", remaining


def asset_read(asset: ITAsset) -> AssetRead:
    state, days = maintenance_status(asset)
    values = {column.name: getattr(asset, column.name) for column in ITAsset.__table__.columns}
    values["maintenance_state"] = state
    values["maintenance_days_remaining"] = days
    values["price"] = float(asset.price) if asset.price is not None else None
    values["scrap_value"] = float(asset.scrap_value) if asset.scrap_value is not None else None
    return AssetRead.model_validate(values)


def summary_for(items: list[ITAsset]) -> AssetSummary:
    states = [maintenance_status(item)[0] for item in items]
    return AssetSummary(
        total=len(items),
        active=sum(item.status == "Active" for item in items),
        spare=sum(item.status == "Spare" for item in items),
        maintenance_due=sum(state in {"due", "overdue"} for state in states),
        maintenance_overdue=sum(state == "overdue" for state in states),
        repair_needed=sum(item.status in {"Repair needed", "Under maintenance"} for item in items),
        recovery_required=sum(item.status == "Recovery required" for item in items),
        scrap_queue=sum(item.status in SCRAP_QUEUE_STATUSES for item in items),
        scrapped_or_disposed=sum(item.status in {"Scrapped", "Disposed"} for item in items),
        total_value=round(sum(float(item.price or 0) for item in items), 2),
    )


def source_order(item: ITAsset) -> tuple[int, int | str]:
    if item.source_sn and item.source_sn.isdigit():
        return 0, int(item.source_sn)
    return 1, item.source_sn or ""


@router.get("", response_model=AssetListResponse)
async def list_assets(
    search: str = Query("", max_length=200),
    status: str = Query("All", max_length=60),
    category: str = Query("All", max_length=160),
    location: str = Query("All", max_length=240),
    department: str = Query("All", max_length=240),
    attention_only: bool = False,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetListResponse:
    await ensure_asset_access(user, session)
    all_items = sorted(list(await session.scalars(select(ITAsset).where(ITAsset.organization_id == user.organization_id))), key=source_order)
    filtered = all_items
    if search.strip():
        needle = search.strip().lower()
        filtered = [item for item in filtered if needle in " ".join(str(value or "") for value in (item.source_sn, item.employee, item.physical_location, item.department_name, item.category, item.brand, item.model, item.serial_imei, item.sim_no, item.label_no)).lower()]
    if status != "All":
        filtered = [item for item in filtered if item.status == status]
    if category != "All":
        filtered = [item for item in filtered if item.category == category]
    if location != "All":
        filtered = [item for item in filtered if item.physical_location == location]
    if department != "All":
        filtered = [item for item in filtered if item.department_name == department]
    if attention_only:
        filtered = [item for item in filtered if maintenance_status(item)[0] in {"due", "overdue"} or item.status in {"Repair needed", "Recovery required", *SCRAP_QUEUE_STATUSES}]
    return AssetListResponse(
        items=[asset_read(item) for item in filtered],
        summary=summary_for(all_items),
        categories=sorted({item.category for item in all_items if item.category}, key=str.lower),
        locations=sorted({item.physical_location for item in all_items if item.physical_location}, key=str.lower),
        departments=sorted({item.department_name for item in all_items if item.department_name}, key=str.lower),
    )


@router.post("", response_model=AssetRead, status_code=201)
async def create_asset(
    payload: AssetPayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetRead:
    await ensure_asset_access(user, session)
    existing = list(await session.scalars(select(ITAsset.source_sn).where(ITAsset.organization_id == user.organization_id)))
    numeric = [int(value) for value in existing if value and value.isdigit()]
    source_sn = str(max(numeric, default=0) + 1)
    values = payload.model_dump()
    item = ITAsset(
        organization_id=user.organization_id,
        asset_key=asset_key(0, source_sn, values.get("label_no"), values.get("serial_imei"), values.get("sim_no")) + f":{uuid.uuid4().hex[:8]}",
        source_sn=source_sn,
        invoice_date_raw=None,
        price_raw=None,
        **values,
    )
    session.add(item)
    await session.flush()
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.created", target_type="it_asset", target_id=str(item.id), metadata_json={"category": item.category, "label_no": item.label_no}))
    await session.commit()
    await session.refresh(item)
    return asset_read(item)


async def get_asset(session: AsyncSession, asset_id: uuid.UUID, user: User) -> ITAsset:
    item = await session.scalar(select(ITAsset).where(ITAsset.id == asset_id, ITAsset.organization_id == user.organization_id))
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found.")
    return item


@router.patch("/{asset_id}", response_model=AssetRead)
async def update_asset(
    asset_id: uuid.UUID,
    payload: AssetUpdatePayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetRead:
    await ensure_asset_access(user, session)
    item = await get_asset(session, asset_id, user)
    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        if isinstance(value, str):
            value = clean_text(value)
        setattr(item, field, value)
    item.invoice_date_raw = None if "invoice_date" in updates else item.invoice_date_raw
    item.price_raw = None if "price" in updates else item.price_raw
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.updated", target_type="it_asset", target_id=str(item.id), metadata_json={"fields": sorted(updates)}))
    await session.commit()
    await session.refresh(item)
    return asset_read(item)


@router.post("/{asset_id}/maintenance", response_model=AssetRead)
async def record_maintenance(
    asset_id: uuid.UUID,
    payload: MaintenancePayload,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> AssetRead:
    await ensure_asset_access(user, session)
    item = await get_asset(session, asset_id, user)
    event = AssetMaintenanceEvent(organization_id=user.organization_id, asset_id=item.id, service_date=payload.service_date, vendor=clean_text(payload.vendor), cost=payload.cost, notes=clean_text(payload.notes), next_due_date=payload.next_due_date, created_by_user_id=user.id)
    session.add(event)
    item.last_maintenance_date = payload.service_date
    item.next_maintenance_date = payload.next_due_date
    item.maintenance_notes = clean_text(payload.notes) or item.maintenance_notes
    if item.status in {"Under maintenance", "Repair needed"}:
        item.status = "Active"
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.maintenance_recorded", target_type="it_asset", target_id=str(item.id), metadata_json={"service_date": payload.service_date.isoformat(), "next_due_date": payload.next_due_date.isoformat() if payload.next_due_date else None}))
    await session.commit()
    await session.refresh(item)
    return asset_read(item)


@router.get("/{asset_id}/maintenance", response_model=list[MaintenanceRead])
async def maintenance_history(
    asset_id: uuid.UUID,
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> list[MaintenanceRead]:
    await ensure_asset_access(user, session)
    await get_asset(session, asset_id, user)
    events = list(await session.scalars(select(AssetMaintenanceEvent).where(AssetMaintenanceEvent.asset_id == asset_id, AssetMaintenanceEvent.organization_id == user.organization_id).order_by(AssetMaintenanceEvent.service_date.desc())))
    return [MaintenanceRead.model_validate(event) for event in events]


@router.post("/import")
async def import_assets(
    file: UploadFile = File(...),
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> dict:
    await ensure_asset_access(user, session)
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=422, detail="Upload the IT asset register as an Excel .xlsx file.")
    content = await file.read(get_settings().max_excel_upload_size_mb * 1024 * 1024 + 1)
    if len(content) > get_settings().max_excel_upload_size_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"The Excel file exceeds {get_settings().max_excel_upload_size_mb} MB.")
    try:
        rows = read_asset_rows(io.BytesIO(content))
    except Exception as error:
        raise HTTPException(status_code=422, detail="The IT asset workbook could not be read.") from error
    if not rows:
        raise HTTPException(status_code=422, detail="No asset rows were found in the workbook.")
    created, updated = await merge_asset_rows(session, user.organization_id, rows)
    session.add(AuditEvent(organization_id=user.organization_id, actor_user_id=user.id, action="asset.register_imported", target_type="it_asset", target_id="register", metadata_json={"filename": file.filename, "created": created, "updated": updated}))
    await session.commit()
    return {"created": created, "updated": updated, "total_rows": len(rows)}


@router.get("/export/register")
async def export_assets(
    user: User = Depends(require_permissions("users.manage")),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    await ensure_asset_access(user, session)
    if not ASSET_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="The IT asset Excel template is unavailable.")
    items = sorted(list(await session.scalars(select(ITAsset).where(ITAsset.organization_id == user.organization_id))), key=source_order)
    workbook = load_workbook(ASSET_TEMPLATE_PATH)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    template_styles = [copy(cell._style) for cell in sheet[2]]
    template_formats = [cell.number_format for cell in sheet[2]]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    red_fill = PatternFill("solid", fgColor="FF0000")
    green_fill = PatternFill("solid", fgColor="92D050")
    for row_index, item in enumerate(items, start=2):
        source_sn: object = int(item.source_sn) if item.source_sn and item.source_sn.isdigit() else item.source_sn
        invoice_value: object = item.invoice_date or (item.invoice_date_raw if item.invoice_date_raw and not item.invoice_date_raw.startswith("#") else None)
        price_value: object = float(item.price) if item.price is not None else (item.price_raw if item.price_raw and not item.price_raw.startswith("#") else None)
        values = [source_sn, item.employee, item.physical_location, item.department_name, item.home_office, item.category, item.brand, item.model, item.serial_imei, item.sim_no, item.ups, item.label_no, invoice_value, item.invoice_no, item.supplier_name, price_value, item.warranty]
        sheet.append(values)
        for column, cell in enumerate(sheet[row_index], start=1):
            cell._style = copy(template_styles[column - 1])
            cell.number_format = template_formats[column - 1]
            if item.status in {"Recovery required", "Scrap proposed", "Approved for scrap", "Scrapped", "Disposed"}:
                cell.fill = copy(red_fill)
            elif item.status == "Spare":
                cell.fill = copy(green_fill)
        if isinstance(invoice_value, date):
            sheet.cell(row_index, 13).number_format = "dd-mm-yyyy"
        if item.price is not None:
            sheet.cell(row_index, 16).number_format = "#,##0.00"
    last_row = max(2, len(items) + 1)
    for table in sheet.tables.values():
        table.ref = f"A1:Q{last_row}"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="IT ASSETS ORG.xlsx"'})
