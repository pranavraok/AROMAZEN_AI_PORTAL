import io
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


AMOUNT_TOLERANCE = 1.0


def _text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _header(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", _text(value).upper()).strip()


def _key_part(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", _text(value).upper())


def _amount(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    try:
        return round(float(str(value).replace(",", "").replace("₹", "").strip()), 2)
    except ValueError:
        return None


def _date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = _text(value)
    for pattern in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _iso(value: date | None) -> str | None:
    return value.isoformat() if value else None


def _workbook(content: bytes, label: str):
    try:
        return load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError(f"{label} could not be read as an Excel workbook. Upload the original .xlsx file.") from exc


def _find_header(sheet, required_groups: list[tuple[str, ...]], label: str, max_rows: int = 40) -> tuple[int, dict[str, int]]:
    best: tuple[int, dict[str, int]] | None = None
    for row_number in range(1, min(sheet.max_row, max_rows) + 1):
        values = {_header(sheet.cell(row_number, column).value): column for column in range(1, min(sheet.max_column, 500) + 1) if _header(sheet.cell(row_number, column).value)}
        matched: dict[str, int] = {}
        for aliases in required_groups:
            found = [values[_header(alias)] for alias in aliases if _header(alias) in values]
            if found:
                matched[aliases[0]] = found[0]
        if best is None or len(matched) > len(best[1]):
            best = (row_number, matched)
        if len(matched) == len(required_groups):
            return row_number, matched
    missing = [aliases[0] for aliases in required_groups if not best or aliases[0] not in best[1]]
    raise ValueError(f"{label}: required columns were not found ({', '.join(missing)}). Keep these column headings or restore them in the uploaded sheet.")


def _find_sheet(workbook, required_groups: list[tuple[str, ...]], label: str) -> tuple[Any, int, dict[str, int]]:
    candidates = []
    for sheet in workbook.worksheets:
        try:
            row, columns = _find_header(sheet, required_groups, label)
            candidates.append((sheet, row, columns))
        except ValueError:
            continue
    if len(candidates) != 1:
        if not candidates:
            raise ValueError(f"{label}: no sheet contains all required columns. Do not rename or remove the invoice, GSTIN, date and amount headings.")
        names = ", ".join(item[0].title for item in candidates)
        raise ValueError(f"{label}: more than one possible data sheet was found ({names}). Keep only one register sheet with the required headings.")
    return candidates[0]


BOOK_COMMON = [
    ("Date",), ("Particulars", "Supplier"), ("Voucher No.", "Voucher No"),
    ("GSTIN/UIN", "GSTIN"), ("Gross Total",),
]
PURCHASE_EXTRA = [("Supplier Invoice No.", "Supplier Invoice No"), ("Supplier Invoice Date",)]
JOURNAL_EXTRA = [("Voucher Ref No", "Voucher Ref No."), ("Voucher Ref Date",)]


def _parse_books(content: bytes, kind: str, label: str) -> tuple[list[dict], list[dict], int]:
    workbook = _workbook(content, label)
    extra = PURCHASE_EXTRA if kind == "purchase" else JOURNAL_EXTRA
    sheet, header_row, columns = _find_sheet(workbook, BOOK_COMMON + extra, label)
    invoice_col = columns[extra[0][0]]
    invoice_date_col = columns[extra[1][0]]
    records: list[dict] = []
    incomplete: list[dict] = []
    ignored = 0
    for row in range(header_row + 1, sheet.max_row + 1):
        transaction_date = _date(sheet.cell(row, columns["Date"]).value)
        supplier = _text(sheet.cell(row, columns["Particulars"]).value)
        voucher = _text(sheet.cell(row, columns["Voucher No."]).value)
        gstin = _text(sheet.cell(row, columns["GSTIN/UIN"]).value).upper()
        invoice = _text(sheet.cell(row, invoice_col).value)
        invoice_date = _date(sheet.cell(row, invoice_date_col).value)
        gross = _amount(sheet.cell(row, columns["Gross Total"]).value)
        if supplier.upper() in {"GRAND TOTAL", "TOTAL"}:
            continue
        if not any((transaction_date, supplier, voucher, gstin, invoice, gross not in (None, 0))):
            ignored += 1
            continue
        # Pure accounting/adjustment rows without a GST identity are outside invoice reconciliation.
        if not gstin and not invoice and gross in (None, 0):
            ignored += 1
            continue
        missing = []
        if not gstin:
            missing.append("GSTIN")
        if not invoice:
            missing.append("invoice number")
        if not invoice_date:
            missing.append("invoice date")
        if gross is None:
            missing.append("gross total")
        base = {
            "source": "Purchase Register" if kind == "purchase" else "Journal Register",
            "source_row": row,
            "supplier": supplier,
            "voucher_number": voucher,
            "gstin": gstin,
            "invoice_number": invoice,
            "invoice_date": _iso(invoice_date),
            "transaction_date": _iso(transaction_date),
            "invoice_value": gross,
        }
        if missing:
            incomplete.append({**base, "status": "incomplete_books", "issues": [f"Missing {item}" for item in missing]})
            continue
        records.append({**base, "key": f"{_key_part(gstin)}|{_key_part(invoice)}"})
    return records, incomplete, ignored


PORTAL_REQUIRED = [
    ("GSTIN of supplier",), ("Trade/Legal name",), ("Invoice number",),
    ("Invoice Date",), ("Invoice Value(₹)", "Invoice Value"), ("Taxable Value (₹)", "Taxable Value"),
    ("Integrated Tax(₹)", "Integrated Tax"), ("Central Tax(₹)", "Central Tax"),
    ("State/UT Tax(₹)", "State/UT Tax"), ("Cess(₹)", "Cess"),
]


def _portal_sheet(workbook):
    candidates = [sheet for sheet in workbook.worksheets if _header(sheet.title) == "B2B"]
    if len(candidates) != 1:
        raise ValueError("GST Portal GSTR-2B: the required B2B sheet was not found. Upload the original GSTR-2B Excel downloaded from the GST portal.")
    sheet = candidates[0]
    row = 0
    values: dict[str, int] = {}
    for parent_row in range(1, min(sheet.max_row, 14) + 1):
        candidate: dict[str, int] = {}
        for col in range(1, min(sheet.max_column, 500) + 1):
            for value in (sheet.cell(parent_row, col).value, sheet.cell(parent_row + 1, col).value):
                if _header(value):
                    candidate[_header(value)] = col
        if _header("GSTIN of supplier") in candidate and _header("Invoice number") in candidate:
            row, values = parent_row + 1, candidate
            break
    if not values:
        raise ValueError("GST Portal GSTR-2B: the two-row B2B column headings were not found. Upload the original portal download without changing its heading rows.")
    columns: dict[str, int] = {}
    missing = []
    for aliases in PORTAL_REQUIRED:
        found = next((values[_header(alias)] for alias in aliases if _header(alias) in values), None)
        if found is None:
            missing.append(aliases[0])
        else:
            columns[aliases[0]] = found
    if missing:
        raise ValueError(f"GST Portal GSTR-2B: required columns were not found ({', '.join(missing)}). Upload a fresh GSTR-2B Excel from the GST portal.")
    # Optional portal fields are resolved without making the file unusable.
    for name, aliases in {
        "ITC Availability": ("ITC Availability",), "Reason": ("Reason",),
        "Reverse Charge": ("Supply Attract Reverse Charge", "Reverse Charge"),
    }.items():
        for alias in aliases:
            if _header(alias) in values:
                columns[name] = values[_header(alias)]
                break
    return sheet, row, columns


def _parse_portal(content: bytes) -> tuple[list[dict], dict, int, int]:
    workbook = _workbook(content, "GST Portal GSTR-2B")
    sheet, header_row, columns = _portal_sheet(workbook)
    records = []
    for row in range(header_row + 1, sheet.max_row + 1):
        gstin = _text(sheet.cell(row, columns["GSTIN of supplier"]).value).upper()
        invoice = _text(sheet.cell(row, columns["Invoice number"]).value)
        if not gstin and not invoice:
            continue
        missing = []
        if not gstin:
            missing.append("GSTIN")
        if not invoice:
            missing.append("invoice number")
        invoice_date = _date(sheet.cell(row, columns["Invoice Date"]).value)
        invoice_value = _amount(sheet.cell(row, columns["Invoice Value(₹)"]).value)
        if not invoice_date:
            missing.append("invoice date")
        if invoice_value is None:
            missing.append("invoice value")
        if missing:
            raise ValueError(f"GST Portal GSTR-2B row {row} is incomplete ({', '.join(missing)}). Download a fresh GSTR-2B Excel from the GST portal.")
        def optional(name: str) -> str:
            return _text(sheet.cell(row, columns[name]).value) if name in columns else ""
        records.append({
            "key": f"{_key_part(gstin)}|{_key_part(invoice)}", "source_row": row,
            "supplier": _text(sheet.cell(row, columns["Trade/Legal name"]).value),
            "gstin": gstin, "invoice_number": invoice, "invoice_date": _iso(invoice_date),
            "invoice_value": invoice_value,
            "taxable_value": _amount(sheet.cell(row, columns["Taxable Value (₹)"]).value) or 0,
            "igst": _amount(sheet.cell(row, columns["Integrated Tax(₹)"]).value) or 0,
            "cgst": _amount(sheet.cell(row, columns["Central Tax(₹)"]).value) or 0,
            "sgst": _amount(sheet.cell(row, columns["State/UT Tax(₹)"]).value) or 0,
            "cess": _amount(sheet.cell(row, columns["Cess(₹)"]).value) or 0,
            "itc_availability": optional("ITC Availability"), "portal_reason": optional("Reason"),
            "reverse_charge": optional("Reverse Charge"),
        })
    if not records:
        raise ValueError("GST Portal GSTR-2B: the B2B sheet contains no invoice rows.")
    metadata = _portal_metadata(workbook)
    credit_notes = _count_data_rows(workbook, "B2B CDNR", ("Note number", "GSTIN of supplier"))
    imports = _count_data_rows(workbook, "IMPG", ("Bill of entry number", "Port Code"))
    return records, metadata, credit_notes, imports


def _portal_metadata(workbook) -> dict:
    metadata = {"period": "", "company_gstin": "", "company_name": ""}
    readme = next((sheet for sheet in workbook.worksheets if _header(sheet.title) == "READ ME"), None)
    if not readme:
        return metadata
    for row_number in range(1, min(readme.max_row, 60) + 1):
        values = [_text(readme.cell(row_number, column).value) for column in range(1, min(readme.max_column, 20) + 1)]
        for index, value in enumerate(values):
            key = _header(value)
            following = next((item for item in values[index + 1:] if item), "")
            if key == "TAX PERIOD": metadata["period"] = following
            elif key == "GSTIN": metadata["company_gstin"] = following
            elif key in {"LEGAL NAME", "TRADE NAME"} and not metadata["company_name"]: metadata["company_name"] = following
    return metadata


def _count_data_rows(workbook, sheet_name: str, identifiers: tuple[str, ...]) -> int:
    sheet = next((item for item in workbook.worksheets if _header(item.title) == sheet_name), None)
    if not sheet:
        return 0
    for row in range(1, min(sheet.max_row, 15) + 1):
        headers = {_header(sheet.cell(row, col).value): col for col in range(1, min(sheet.max_column, 100) + 1)}
        columns = [headers[_header(name)] for name in identifiers if _header(name) in headers]
        if columns:
            return sum(1 for current in range(row + 1, sheet.max_row + 1) if any(_text(sheet.cell(current, col).value) for col in columns))
    return 0


def _duplicate_keys(records: list[dict]) -> set[str]:
    counts = Counter(item["key"] for item in records)
    return {key for key, count in counts.items() if count > 1}


def reconcile(purchase_content: bytes, journal_content: bytes, portal_content: bytes) -> dict:
    purchases, incomplete_purchases, ignored_purchases = _parse_books(purchase_content, "purchase", "Tally Purchase Register")
    journals, incomplete_journals, ignored_journals = _parse_books(journal_content, "journal", "Tally Journal Register")
    books = purchases + journals
    incomplete = incomplete_purchases + incomplete_journals
    portal, metadata, credit_notes, imports = _parse_portal(portal_content)
    if metadata.get("period") and not re.search(r"\d{4}", metadata["period"]):
        years = Counter(item["invoice_date"][:4] for item in portal if item.get("invoice_date"))
        if years:
            metadata["period"] = f"{metadata['period']} {years.most_common(1)[0][0]}"
    duplicate_book_keys = _duplicate_keys(books)
    duplicate_portal_keys = _duplicate_keys(portal)
    duplicates = duplicate_book_keys | duplicate_portal_keys
    books_by_key: dict[str, list[dict]] = defaultdict(list)
    portal_by_key: dict[str, list[dict]] = defaultdict(list)
    for item in books: books_by_key[item["key"]].append(item)
    for item in portal: portal_by_key[item["key"]].append(item)
    rows: list[dict] = []
    all_keys = set(books_by_key) | set(portal_by_key)
    for key in all_keys:
        book_items, portal_items = books_by_key.get(key, []), portal_by_key.get(key, [])
        if key in duplicates:
            exemplar = (book_items or portal_items)[0]
            rows.append({
                "id": key, "status": "duplicate", "issues": ["Duplicate GSTIN and invoice number; review before matching"],
                "supplier": exemplar["supplier"], "gstin": exemplar["gstin"], "invoice_number": exemplar["invoice_number"],
                "books_source": ", ".join(sorted({item["source"] for item in book_items})),
                "books_row": ", ".join(str(item["source_row"]) for item in book_items),
                "books_date": book_items[0]["invoice_date"] if book_items else None,
                "portal_date": portal_items[0]["invoice_date"] if portal_items else None,
                "books_invoice_value": sum(item["invoice_value"] for item in book_items),
                "portal_invoice_value": sum(item["invoice_value"] for item in portal_items),
                "difference": round(sum(item["invoice_value"] for item in book_items) - sum(item["invoice_value"] for item in portal_items), 2),
                "portal_taxable_value": sum(item.get("taxable_value", 0) for item in portal_items),
                "igst": sum(item.get("igst", 0) for item in portal_items), "cgst": sum(item.get("cgst", 0) for item in portal_items),
                "sgst": sum(item.get("sgst", 0) for item in portal_items), "cess": sum(item.get("cess", 0) for item in portal_items),
                "itc_availability": portal_items[0].get("itc_availability", "") if portal_items else "", "portal_reason": portal_items[0].get("portal_reason", "") if portal_items else "",
            })
            continue
        if not book_items:
            item = portal_items[0]
            rows.append(_result_row("portal_only", ["Present in GSTR-2B but not found in Tally registers"], None, item, key))
            continue
        if not portal_items:
            item = book_items[0]
            rows.append(_result_row("books_only", ["Present in Tally but not found in GSTR-2B"], item, None, key))
            continue
        book, portal_item = book_items[0], portal_items[0]
        difference = round(book["invoice_value"] - portal_item["invoice_value"], 2)
        issues = []
        if abs(difference) > AMOUNT_TOLERANCE:
            issues.append("Invoice value mismatch")
        if book["invoice_date"] != portal_item["invoice_date"]:
            issues.append("Invoice date mismatch")
        if portal_item.get("itc_availability") and portal_item["itc_availability"].strip().lower() not in {"yes", "y", "available"}:
            issues.append("ITC not available in GSTR-2B")
        rows.append(_result_row("mismatch" if issues else "matched", issues, book, portal_item, key))
    rows.extend({
        "id": f"incomplete-{item['source']}-{item['source_row']}", **item,
        "books_source": item["source"], "books_row": str(item["source_row"]), "books_date": item["invoice_date"],
        "portal_date": None, "books_invoice_value": item["invoice_value"], "portal_invoice_value": None,
        "difference": None, "portal_taxable_value": None, "igst": None, "cgst": None, "sgst": None, "cess": None,
        "itc_availability": "", "portal_reason": "",
    } for item in incomplete)
    priority = {"duplicate": 0, "mismatch": 1, "incomplete_books": 2, "books_only": 3, "portal_only": 4, "matched": 5}
    rows.sort(key=lambda item: (priority[item["status"]], item.get("supplier", ""), item.get("invoice_number", "")))
    counts = Counter(item["status"] for item in rows)
    summary = {
        "book_invoices": len(books), "portal_invoices": len(portal), "matched": counts["matched"],
        "mismatched": counts["mismatch"], "books_only": counts["books_only"], "portal_only": counts["portal_only"],
        "incomplete_books": counts["incomplete_books"], "duplicates": counts["duplicate"],
        "portal_credit_notes": credit_notes, "portal_imports": imports,
        "book_invoice_value": round(sum(item["invoice_value"] for item in books), 2),
        "portal_invoice_value": round(sum(item["invoice_value"] for item in portal), 2),
    }
    warnings = []
    if incomplete: warnings.append(f"{len(incomplete)} Tally rows need missing GST details to be filled before they can be matched.")
    if duplicates: warnings.append(f"{len(duplicates)} duplicate invoice keys were not matched automatically.")
    if credit_notes: warnings.append(f"{credit_notes} portal credit/debit note records are shown in the summary and are not mixed with B2B invoice matching.")
    if imports: warnings.append(f"{imports} import records are shown in the summary and are not mixed with B2B invoice matching.")
    return {
        **metadata, "summary": summary, "rows": rows, "warnings": warnings,
        "ignored_non_invoice_rows": ignored_purchases + ignored_journals,
        "amount_tolerance": AMOUNT_TOLERANCE,
    }


def _result_row(status: str, issues: list[str], book: dict | None, portal: dict | None, key: str) -> dict:
    exemplar = book or portal or {}
    book_value = book["invoice_value"] if book else None
    portal_value = portal["invoice_value"] if portal else None
    return {
        "id": key, "status": status, "issues": issues, "supplier": exemplar.get("supplier", ""),
        "gstin": exemplar.get("gstin", ""), "invoice_number": exemplar.get("invoice_number", ""),
        "books_source": book["source"] if book else "", "books_row": str(book["source_row"]) if book else "",
        "books_date": book["invoice_date"] if book else None, "portal_date": portal["invoice_date"] if portal else None,
        "books_invoice_value": book_value, "portal_invoice_value": portal_value,
        "difference": round(book_value - portal_value, 2) if book_value is not None and portal_value is not None else None,
        "portal_taxable_value": portal.get("taxable_value") if portal else None,
        "igst": portal.get("igst") if portal else None, "cgst": portal.get("cgst") if portal else None,
        "sgst": portal.get("sgst") if portal else None, "cess": portal.get("cess") if portal else None,
        "itc_availability": portal.get("itc_availability", "") if portal else "",
        "portal_reason": portal.get("portal_reason", "") if portal else "",
    }
