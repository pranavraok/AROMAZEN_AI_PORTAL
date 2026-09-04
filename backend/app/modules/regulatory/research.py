from __future__ import annotations

import asyncio
from collections import Counter
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from io import BytesIO
import json
from pathlib import Path
import re
from urllib.parse import quote, urlparse

import httpx
from docx import Document
from openpyxl import load_workbook


PUBCHEM_API = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
PUBCHEM_VIEW_API = "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound"
PUBCHEM_PAGE = "https://pubchem.ncbi.nlm.nih.gov/compound"
ECHA_CANDIDATE_API = "https://chem.echa.europa.eu/api-obligation-list/v1/candidateList"
ECHA_CANDIDATE_PAGE = "https://chem.echa.europa.eu/obligation-lists/candidateList"
IFRA_LIBRARY_PAGE = "https://ifrafragrance.org/initiatives-positions/safe-use-fragrance-science/ifra-standards/ifra-standards-documentation"
IFRA_OVERVIEW_URL = "https://d3t14p1xronwr0.cloudfront.net/docs/Standards-Documentation/ifra-51st-amendment-ifra-standards-overview.xlsx"
NITE_GHS_PAGE = "https://www.chem-info.nite.go.jp/chem/english/ghs/ghs_nite_download_e.html"
NITE_GHS_URL = "https://www.chem-info.nite.go.jp/chem/english/ghs/files/list_nite_all_e.xlsx"
EPA_COMPTOX_API = "https://comptox.epa.gov/ctx-api"
EPA_COMPTOX_PAGE = "https://comptox.epa.gov/dashboard/chemical/details"
EU_COSMETICS_REGULATION = "https://eur-lex.europa.eu/eli/reg/2009/1223/oj"
CAS_PATTERN = re.compile(r"^\d{2,7}-\d{2}-\d$")
EC_PATTERN = re.compile(r"^\d{3}-\d{3}-\d$")
CAS_IN_TEXT = re.compile(r"\b\d{2,7}-\d{2}-\d\b")
EC_IN_TEXT = re.compile(r"\b\d{3}-\d{3}-\d\b")
TRUSTED_GHS_HOSTS = (
    "echa.europa.eu",
    "eur-lex.europa.eu",
    "pubchem.ncbi.nlm.nih.gov",
    "chem-info.nite.go.jp",
    "safeworkaustralia.gov.au",
    "industrialchemicals.gov.au",
    "epa.gov",
    "fda.gov",
)
REFERENCE_CACHE_DAYS = 7


class AsyncRequestThrottle:
    """Evenly space requests so concurrent PubChem tasks never burst above the limit."""

    def __init__(self, requests_per_second: int = 4) -> None:
        self._spacing = 1.0 / max(1, requests_per_second)
        self._next_slot = 0.0
        self._lock = asyncio.Lock()

    async def wait(self) -> None:
        loop = asyncio.get_running_loop()
        async with self._lock:
            now = loop.time()
            slot = max(now, self._next_slot)
            self._next_slot = slot + self._spacing
        if slot > now:
            await asyncio.sleep(slot - now)


# The deployment runs one API worker, so this gate is shared by every
# Regulatory workflow in that worker rather than resetting per ingredient.
PUBCHEM_THROTTLE = AsyncRequestThrottle(4)


def valid_cas(value: str) -> bool:
    """Return True only when a CAS-looking identifier has a valid checksum."""
    if not CAS_PATTERN.fullmatch(value):
        return False
    first, second, check = value.split("-")
    body = first + second
    checksum = sum(int(digit) * weight for weight, digit in enumerate(reversed(body), start=1)) % 10
    return checksum == int(check)


def _lookup_names(name: str) -> list[str]:
    cleaned = re.sub(r"\s+", " ", name).strip()
    candidates = [cleaned]
    normalized_trade_name = re.sub(r"\s+SS$", "", cleaned, flags=re.IGNORECASE)
    normalized_trade_name = re.sub(r"CARRYOPHELLENE", "CARYOPHYLLENE", normalized_trade_name, flags=re.IGNORECASE)
    if normalized_trade_name.casefold() != cleaned.casefold():
        candidates.append(normalized_trade_name)
    collapsed_iso = re.sub(r"^ISO\s+(?=[A-Z]{4,}\b)", "ISO", cleaned, count=1, flags=re.IGNORECASE)
    if collapsed_iso.casefold() != cleaned.casefold():
        candidates.append(collapsed_iso)
    if re.fullmatch(r"TERPINYL\s+ACETATE", cleaned, flags=re.IGNORECASE):
        candidates.append("alpha-terpinyl acetate")
    if cleaned.casefold() == "dhm":
        candidates.append("dihydromyrcenol")
    return candidates


def _unique(values: list[str], *, limit: int) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = re.sub(r"\s+", " ", str(value)).strip(" ;,|")
        key = cleaned.casefold()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
        if len(result) >= limit:
            break
    return result


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _host(url: object) -> str:
    return (urlparse(str(url or "")).hostname or "").lower().rstrip(".")


def _trusted_host(url: object) -> bool:
    host = _host(url)
    return any(host == allowed or host.endswith(f".{allowed}") for allowed in TRUSTED_GHS_HOSTS)


async def _official_get(
    client: httpx.AsyncClient,
    url: str,
    *,
    params: dict | None = None,
    headers: dict[str, str] | None = None,
    throttle: AsyncRequestThrottle | None = None,
) -> httpx.Response:
    response: httpx.Response | None = None
    for attempt in range(4):
        if throttle is not None:
            await throttle.wait()
        response = await client.get(url, params=params, headers=headers)
        if response.status_code not in {429, 502, 503, 504}:
            return response
        retry_after = response.headers.get("retry-after")
        delay = float(retry_after) if retry_after and retry_after.replace(".", "", 1).isdigit() else 0.6 * (2 ** attempt)
        await asyncio.sleep(min(delay, 5.0))
    assert response is not None
    return response


async def _pubchem_get(
    client: httpx.AsyncClient,
    url: str,
    throttle: AsyncRequestThrottle | None = None,
) -> httpx.Response:
    return await _official_get(client, url, throttle=throttle)


def _section_information(record: dict, heading: str) -> list[dict]:
    found: list[dict] = []

    def visit(value: object) -> None:
        if not isinstance(value, dict):
            return
        if str(value.get("TOCHeading") or "").casefold() == heading.casefold():
            found.extend(item for item in value.get("Information") or [] if isinstance(item, dict))
        for child in value.get("Section") or []:
            visit(child)

    visit(record)
    return found


def _information_strings(information: dict) -> list[str]:
    value = information.get("Value") or {}
    strings: list[str] = []
    for entry in value.get("StringWithMarkup") or []:
        if isinstance(entry, dict) and str(entry.get("String") or "").strip():
            strings.append(str(entry["String"]).strip())
    if isinstance(value.get("String"), str) and value["String"].strip():
        strings.append(value["String"].strip())
    return strings


def _information_icons(information: dict) -> list[str]:
    icons: list[str] = []
    for entry in (information.get("Value") or {}).get("StringWithMarkup") or []:
        for markup in entry.get("Markup") or [] if isinstance(entry, dict) else []:
            if markup.get("Type") == "Icon" and str(markup.get("Extra") or "").strip():
                icons.append(str(markup["Extra"]).strip())
    return icons


def _ghs_reference_rank(reference: dict) -> int | None:
    source = str(reference.get("SourceName") or "").casefold()
    host = _host(reference.get("URL"))
    if "regulation (ec) no 1272/2008" in source or host.endswith("eur-lex.europa.eu"):
        return 0
    if "european chemicals agency" in source or host.endswith("echa.europa.eu"):
        return 1
    if _trusted_host(reference.get("URL")):
        return 2
    return None


def _classification_from_hazards(hazards: list[str]) -> str:
    classes: list[str] = []
    for hazard in hazards:
        code_match = re.search(r"\bH\d{3}(?:\+H\d{3})*\b", hazard)
        code = code_match.group(0) if code_match else ""
        for bracketed in re.findall(r"\[([^\]]+)\]", hazard):
            label = re.sub(r"^(?:Danger|Warning)\s+", "", bracketed, flags=re.IGNORECASE).strip()
            if label:
                classes.append(f"{label} ({code})" if code else label)
    return "; ".join(_unique(classes, limit=20))


def _clean_hazard_statement(value: str) -> str:
    # PubChem can display the share of ECHA C&L notifications beside an
    # H-code. That percentage is not the ingredient concentration and must
    # never be copied into an SDS.
    return re.sub(
        r"\b(H\d{3}(?:\+H\d{3})*)\s*\(\d+(?:\.\d+)?%\)\s*:",
        r"\1:",
        value,
    ).strip()


def _parse_ghs(record: dict) -> tuple[dict, list[str]]:
    references = {
        int(item.get("ReferenceNumber")): item
        for item in record.get("Reference") or []
        if isinstance(item, dict) and item.get("ReferenceNumber") is not None
    }
    ranked = []
    for item in _section_information(record, "GHS Classification"):
        reference = references.get(int(item.get("ReferenceNumber") or -1), {})
        rank = _ghs_reference_rank(reference)
        if rank is not None:
            ranked.append((rank, item, reference))
    if not ranked:
        return {}, []
    best_rank = min(rank for rank, _, _ in ranked)
    selected = [(item, reference) for rank, item, reference in ranked if rank == best_rank]
    hazards: list[str] = []
    precaution_codes: list[str] = []
    signals: list[str] = []
    pictograms: list[str] = []
    source_urls: list[str] = []
    for item, reference in selected:
        name = str(item.get("Name") or "").casefold()
        values = [_clean_hazard_statement(value) for value in _information_strings(item)]
        if name == "ghs hazard statements":
            hazards.extend(values)
        elif name == "precautionary statement codes":
            for value in values:
                precaution_codes.extend(re.findall(r"\bP\d{3}(?:\+P\d{3})*\b", value))
        elif name == "signal":
            signals.extend(values)
        elif name == "pictogram(s)":
            pictograms.extend(_information_icons(item))
        if _trusted_host(reference.get("URL")):
            source_urls.append(str(reference["URL"]))
    hazards = _unique(hazards, limit=30)
    signal_values = _unique(signals, limit=5)
    signal = "Danger" if any(value.casefold() == "danger" for value in signal_values) else "Warning" if signal_values else ""
    suggestion = {
        "classification": _classification_from_hazards(hazards),
        "hazard_statements": "; ".join(hazards),
        "precautionary_statements": ", ".join(_unique(precaution_codes, limit=40)),
        "signal_word": signal,
        "pictograms": ", ".join(_unique(pictograms, limit=10)),
    }
    return {key: value for key, value in suggestion.items() if value}, _unique(source_urls, limit=12)


async def _primary_cas(
    client: httpx.AsyncClient,
    cid: int,
    fallback: list[str],
    throttle: AsyncRequestThrottle | None = None,
) -> str:
    valid = [value for value in fallback if valid_cas(value)]
    if len(set(valid)) <= 1:
        return valid[0] if valid else ""
    response = await _pubchem_get(client, f"{PUBCHEM_VIEW_API}/{cid}/JSON?heading=CAS", throttle)
    if response.status_code >= 400:
        return Counter(valid).most_common(1)[0][0]
    values = [
        value
        for item in _section_information(response.json().get("Record") or {}, "CAS")
        for value in _information_strings(item)
        if valid_cas(value)
    ]
    return Counter(values or valid).most_common(1)[0][0]


async def pubchem_regulatory_record(
    client: httpx.AsyncClient,
    name: str,
    throttle: AsyncRequestThrottle | None = None,
) -> tuple[dict, list[str], dict, dict]:
    """Resolve identity and defensible GHS fields from PubChem without an AI key."""
    properties: dict | None = None
    for candidate in _lookup_names(name):
        response = await _pubchem_get(
            client,
            f"{PUBCHEM_API}/compound/name/{quote(candidate, safe='')}/property/Title,IUPACName/JSON",
            throttle,
        )
        if response.status_code == 404:
            continue
        response.raise_for_status()
        rows = ((response.json().get("PropertyTable") or {}).get("Properties") or [])
        if rows:
            properties = rows[0]
            break
    checked_at = _now()
    if not properties or not properties.get("CID"):
        return {}, [], {"identity": {"status": "not_found", "source": "PubChem", "checked_at": checked_at}}, {"pubchem": checked_at}

    cid = int(properties["CID"])
    synonyms_response = await _pubchem_get(client, f"{PUBCHEM_API}/compound/cid/{cid}/synonyms/JSON", throttle)
    synonyms_response.raise_for_status()
    information = ((synonyms_response.json().get("InformationList") or {}).get("Information") or [{}])[0]
    synonyms = [str(value).strip() for value in information.get("Synonym") or [] if str(value).strip()]
    cas_candidates = [value for value in synonyms if valid_cas(value)]
    cas = await _primary_cas(client, cid, cas_candidates, throttle)
    ec_candidates = [value for value in synonyms if EC_PATTERN.fullmatch(value) and not valid_cas(value)]
    ec = Counter(ec_candidates).most_common(1)[0][0] if ec_candidates else ""
    canonical = str(properties.get("Title") or properties.get("IUPACName") or name).strip()
    excluded = {name.casefold(), canonical.casefold(), cas.casefold(), ec.casefold(), ""}
    aliases = _unique(
        [value for value in synonyms if value.casefold() not in excluded and not CAS_PATTERN.fullmatch(value) and not EC_PATTERN.fullmatch(value)],
        limit=12,
    )
    suggestion: dict = {"canonical_name": canonical, "aliases": aliases, "cas": cas, "ec": ec}
    urls = [f"{PUBCHEM_PAGE}/{cid}"]

    ghs_response = await _pubchem_get(client, f"{PUBCHEM_VIEW_API}/{cid}/JSON?heading=GHS%20Classification", throttle)
    if ghs_response.status_code < 400:
        ghs, ghs_urls = _parse_ghs(ghs_response.json().get("Record") or {})
        suggestion.update(ghs)
        urls.extend(ghs_urls)
    else:
        ghs = {}
    checks = {
        "identity": {"status": "matched", "source": "PubChem", "checked_at": checked_at},
        "ghs": {"status": "matched" if ghs else "not_found", "source": "Official references indexed by PubChem", "checked_at": checked_at},
    }
    return {key: value for key, value in suggestion.items() if value}, _unique(urls, limit=20), checks, {"pubchem": checked_at}


async def pubchem_identity(client: httpx.AsyncClient, name: str) -> tuple[dict, list[str]]:
    """Backward-compatible identity-only wrapper used by existing callers and tests."""
    suggestion, urls, _, _ = await pubchem_regulatory_record(client, name)
    identity_fields = {key: suggestion[key] for key in ("canonical_name", "aliases", "cas", "ec") if suggestion.get(key)}
    return identity_fields, urls


def _snapshot_is_fresh(snapshot: dict, days: int = REFERENCE_CACHE_DAYS) -> bool:
    try:
        fetched_at = datetime.fromisoformat(str(snapshot["fetched_at"]).replace("Z", "+00:00"))
    except (KeyError, TypeError, ValueError):
        return False
    return (datetime.now(timezone.utc) - fetched_at).total_seconds() < days * 86_400


def _read_snapshot(path: Path) -> dict | None:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
        return value if isinstance(value, dict) else None
    except (OSError, json.JSONDecodeError):
        return None


def _write_snapshot(path: Path, snapshot: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    temporary.write_text(json.dumps(snapshot, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    temporary.replace(path)


def _source_version(response: httpx.Response, label: str) -> str:
    last_modified = response.headers.get("last-modified")
    if last_modified:
        try:
            source_date = parsedate_to_datetime(last_modified).date().isoformat()
            return f"{label} {source_date}"
        except (TypeError, ValueError):
            pass
    return f"{label} snapshot {datetime.now(timezone.utc).date().isoformat()}"


async def _versioned_snapshot(
    client: httpx.AsyncClient,
    *,
    cache_dir: Path,
    key: str,
    source_url: str,
    label: str,
    parser,
    params: dict | None = None,
    cache_days: int = REFERENCE_CACHE_DAYS,
) -> dict:
    path = cache_dir / f"{key}.json"
    cached = _read_snapshot(path)
    if cached and _snapshot_is_fresh(cached, cache_days):
        return cached
    try:
        response = await _official_get(client, source_url, params=params)
        response.raise_for_status()
        snapshot = {
            "source_url": source_url,
            "version": _source_version(response, label),
            "fetched_at": _now(),
            "records": parser(response),
        }
        _write_snapshot(path, snapshot)
        return snapshot
    except (httpx.HTTPError, OSError, ValueError, TypeError):
        if cached:
            cached["stale"] = True
            return cached
        raise


def _parse_echa_snapshot(response: httpx.Response) -> list[dict]:
    records: list[dict] = []
    for item in response.json().get("items") or []:
        if not isinstance(item, dict):
            continue
        records.append({
            "substanceName": [str(value).strip() for value in item.get("substanceName") or [] if str(value).strip()],
            "casNumber": [str(value).strip() for value in item.get("casNumber") or [] if str(value).strip()],
            "ecNumber": [str(value).strip() for value in item.get("ecNumber") or [] if str(value).strip()],
            "dateOfInclusion": str(item.get("dateOfInclusion") or "").strip(),
            "reasonForInclusion": [str(value).strip() for value in item.get("reasonForInclusion") or [] if str(value).strip()],
            "decision": [
                {"decisionPath": str(value.get("decisionPath") or "").strip()}
                for value in item.get("decision") or []
                if isinstance(value, dict) and str(value.get("decisionPath") or "").strip()
            ],
        })
    return records


async def load_echa_candidate_snapshot(client: httpx.AsyncClient, cache_dir: Path) -> dict:
    path = cache_dir / "echa-candidate-list.json"
    cached = _read_snapshot(path)
    if cached and _snapshot_is_fresh(cached):
        return cached
    try:
        records: list[dict] = []
        first_response: httpx.Response | None = None
        page = 1
        total_pages = 1
        while page <= total_pages:
            response = await _official_get(
                client,
                ECHA_CANDIDATE_API,
                params={"pageIndex": page, "pageSize": 100},
            )
            response.raise_for_status()
            if first_response is None:
                first_response = response
            records.extend(_parse_echa_snapshot(response))
            state = response.json().get("state") or {}
            total_pages = min(10, max(1, int(state.get("totalPages") or 1)))
            page += 1
        assert first_response is not None
        snapshot = {
            "source_url": ECHA_CANDIDATE_API,
            "version": _source_version(first_response, "ECHA Candidate List"),
            "fetched_at": _now(),
            "records": records,
        }
        _write_snapshot(path, snapshot)
        return snapshot
    except (httpx.HTTPError, OSError, ValueError, TypeError):
        if cached:
            cached["stale"] = True
            return cached
        raise


def _parse_nite_snapshot(response: httpx.Response) -> list[dict]:
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    ignored = {
        "", "-", "not classified", "not classified (not applicable)",
        "classification not possible", "classification not possible (insufficient data)",
    }
    records: list[dict] = []
    for row in rows:
        cas = str(row[0] or "").strip()
        if not valid_cas(cas):
            continue
        hazards = []
        for index in range(4, min(len(headers) - 1, len(row))):
            value = re.sub(r"\s+", " ", str(row[index] or "")).strip()
            if value.casefold() not in ignored:
                hazards.append(f"{headers[index]}: {value}")
        records.append({
            "cas": cas,
            "name": re.sub(r"\s+", " ", str(row[2] or "")).strip(),
            "classification": "; ".join(hazards),
            "detail_url": str(row[len(headers) - 1] or "").strip(),
        })
    workbook.close()
    return records


async def load_nite_ghs_snapshot(client: httpx.AsyncClient, cache_dir: Path) -> dict:
    return await _versioned_snapshot(
        client,
        cache_dir=cache_dir,
        key="nite-japan-ghs-en",
        source_url=NITE_GHS_URL,
        label="NITE Japan-GHS",
        parser=_parse_nite_snapshot,
    )


def nite_ghs_match(snapshot: dict | None, item: dict) -> tuple[dict, list[str], dict, dict]:
    checked_at = _now()
    if snapshot is None:
        check = {"status": "unavailable", "source": "NITE Japan-GHS", "checked_at": checked_at}
        return {}, [], {"nite_ghs": check}, {}
    cas_values = {value for value in CAS_IN_TEXT.findall(str(item.get("cas") or "")) if valid_cas(value)}
    version = str(snapshot.get("version") or "NITE Japan-GHS")
    record = next((value for value in snapshot.get("records") or [] if value.get("cas") in cas_values), None)
    check = {
        "status": "matched" if record else "not_found",
        "source": "NITE Japan-GHS",
        "checked_at": checked_at,
        "details": version,
    }
    suggestion = {"classification": str(record.get("classification") or "")} if record else {}
    urls = [NITE_GHS_PAGE]
    if record and str(record.get("detail_url") or "").startswith("https://"):
        urls.append(str(record["detail_url"]))
    return suggestion, urls, {"nite_ghs": check}, {"nite_ghs": version}


def _parse_ifra_snapshot(response: httpx.Response) -> list[dict]:
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [str(worksheet.cell(3, index).value or "").strip() for index in range(1, worksheet.max_column + 1)]
    positions = {header: index for index, header in enumerate(headers)}
    category_positions = [(index, header) for index, header in enumerate(headers) if header.startswith("Category ")]
    records: list[dict] = []
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        name = re.sub(r"\s+", " ", str(row[positions["Name of the IFRA Standard"]] or "")).strip()
        if not name:
            continue
        cas_values = [value for value in CAS_IN_TEXT.findall(str(row[positions["CAS numbers"]] or "")) if valid_cas(value)]
        synonyms = [re.sub(r"\s+", " ", value).strip() for value in str(row[positions["Synonyms"]] or "").splitlines() if value.strip()]
        limits = []
        for index, header in category_positions:
            value = row[index]
            if value is None or str(value).strip() == "":
                continue
            rendered = f"{value:g}" if isinstance(value, (int, float)) else re.sub(r"\s+", " ", str(value)).strip()
            limits.append(f"{header.removesuffix(' (%)')}: {rendered}{'%' if isinstance(value, (int, float)) else ''}")
        records.append({
            "name": name,
            "normalized_name": _normalise(name),
            "cas": cas_values,
            "synonyms": synonyms,
            "normalized_synonyms": [_normalise(value) for value in synonyms],
            "standard_type": str(row[positions["IFRA Standard type"]] or "").strip(),
            "intrinsic_property": str(row[positions["Intrinsic property driving the risk management measure"]] or "").strip(),
            "amendment": str(row[positions["Amendment number"]] or "").strip(),
            "publication_year": str(row[positions["Year of last publication"]] or "").strip(),
            "limits": "; ".join(limits),
        })
    workbook.close()
    return records


async def load_ifra_snapshot(client: httpx.AsyncClient, cache_dir: Path) -> dict:
    snapshot = await _versioned_snapshot(
        client,
        cache_dir=cache_dir,
        key="ifra-51st-amendment-overview",
        source_url=IFRA_OVERVIEW_URL,
        label="IFRA 51st Amendment overview",
        parser=_parse_ifra_snapshot,
        cache_days=30,
    )
    snapshot["version"] = "IFRA 51st Amendment official overview"
    return snapshot


def ifra_snapshot_match(snapshot: dict | None, item: dict) -> tuple[dict, list[str], dict, dict]:
    checked_at = _now()
    if snapshot is None:
        check = {"status": "unavailable", "source": "IFRA Standards official overview", "checked_at": checked_at}
        return {}, [], {"ifra": check}, {}
    cas_values = {value for value in CAS_IN_TEXT.findall(str(item.get("cas") or "")) if valid_cas(value)}
    names = {
        _normalise(value)
        for value in [item.get("name"), item.get("canonical_name"), *(item.get("aliases") or [])]
        if _normalise(value)
    }
    record = next((
        value for value in snapshot.get("records") or []
        if cas_values.intersection(value.get("cas") or [])
        or value.get("normalized_name") in names
        or names.intersection(value.get("normalized_synonyms") or [])
    ), None)
    version = str(snapshot.get("version") or "IFRA official overview")
    details = ""
    suggestion: dict = {}
    if record:
        details = "; ".join(filter(None, [record.get("standard_type"), record.get("intrinsic_property"), f"Amendment {record.get('amendment')}" if record.get("amendment") else ""]))
        if record.get("limits"):
            suggestion["ifra_limits"] = record["limits"]
    check = {
        "status": "listed" if record else "not_listed",
        "source": "IFRA Standards official overview",
        "checked_at": checked_at,
        "details": "; ".join(filter(None, [version, details])),
    }
    return suggestion, [IFRA_LIBRARY_PAGE], {"ifra": check}, {"ifra": version}


async def epa_comptox_identity(
    client: httpx.AsyncClient,
    *,
    name: str,
    cas: str = "",
    api_key: str,
) -> tuple[dict, list[str], dict, dict]:
    """Supplement identity only with exact EPA DSSTox matches."""
    checked_at = _now()
    headers = {"x-api-key": api_key, "accept": "application/json"}
    terms = [*CAS_IN_TEXT.findall(cas), *_lookup_names(name)]
    match: dict | None = None
    for term in _unique(terms, limit=4):
        response = await _official_get(
            client,
            f"{EPA_COMPTOX_API}/chemical/search/equal/{quote(term, safe='')}",
            headers=headers,
        )
        if response.status_code in {400, 404}:
            continue
        response.raise_for_status()
        payload = response.json()
        candidates = payload if isinstance(payload, list) else [payload]
        match = next((candidate for candidate in candidates if isinstance(candidate, dict) and candidate.get("dtxsid")), None)
        if match:
            break
    if not match:
        return {}, [], {"epa_comptox": {"status": "not_found", "source": "EPA CompTox", "checked_at": checked_at}}, {"epa_comptox": checked_at}

    dtxsid = str(match.get("dtxsid") or "").strip()
    aliases: list[str] = []
    synonym_response = await _official_get(
        client,
        f"{EPA_COMPTOX_API}/chemical/synonym/search/by-dtxsid/{quote(dtxsid, safe='')}",
        headers=headers,
    )
    if synonym_response.status_code < 400:
        synonyms = synonym_response.json()
        if isinstance(synonyms, dict):
            for key in ("valid", "good", "alternate", "other"):
                aliases.extend(str(value) for value in synonyms.get(key) or [])
        elif isinstance(synonyms, list):
            aliases.extend(str(value.get("synonym") or "") for value in synonyms if isinstance(value, dict))
    preferred_name = str(match.get("preferredName") or match.get("searchName") or name).strip()
    casrn = str(match.get("casrn") or "").strip()
    excluded = {name.casefold(), preferred_name.casefold(), casrn.casefold(), ""}
    suggestion = {
        "canonical_name": preferred_name,
        "cas": casrn if valid_cas(casrn) else "",
        "aliases": _unique([value for value in aliases if value.casefold() not in excluded], limit=12),
    }
    source_url = f"{EPA_COMPTOX_PAGE}/{dtxsid}"
    check = {"status": "matched", "source": "EPA CompTox / DSSTox", "checked_at": checked_at, "details": dtxsid}
    return {key: value for key, value in suggestion.items() if value}, [source_url], {"epa_comptox": check}, {"epa_comptox": dtxsid}


def _exact_echa_match(item: dict, cas_values: set[str], ec_values: set[str], names: set[str]) -> bool:
    item_cas = {value for value in item.get("casNumber") or [] if isinstance(value, str)}
    item_ec = {value for value in item.get("ecNumber") or [] if isinstance(value, str)}
    item_names = {re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip() for value in item.get("substanceName") or [] if isinstance(value, str)}
    return bool(cas_values.intersection(item_cas) or ec_values.intersection(item_ec) or names.intersection(item_names))


async def echa_svhc_status(
    client: httpx.AsyncClient,
    *,
    name: str,
    cas: str = "",
    ec: str = "",
    snapshot: dict | None = None,
) -> tuple[dict, list[str], dict, dict]:
    """Check a versioned ECHA snapshot, or the live list for legacy callers."""
    cas_values = {value for value in CAS_IN_TEXT.findall(cas) if valid_cas(value)}
    ec_values = {value for value in EC_IN_TEXT.findall(ec) if not valid_cas(value)}
    normalized_names = {
        re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip()
        for value in _lookup_names(name)
        if value.strip()
    }
    checked_at = _now()
    version = str((snapshot or {}).get("version") or checked_at)
    if snapshot is not None:
        batches = [(snapshot.get("records") or [])]
    else:
        # Kept for direct callers and tests. Production enrichment supplies a
        # single cached snapshot instead of repeating a request per ingredient.
        searches = list(cas_values) or list(ec_values) or list(normalized_names)[:1]
        batches = []
        for search in searches:
            response = await _official_get(client, ECHA_CANDIDATE_API, params={
                "pageIndex": 1,
                "pageSize": 100,
                "searchText": search,
            })
            response.raise_for_status()
            batches.append(response.json().get("items") or [])
    for batch in batches:
        for item in batch:
            if not isinstance(item, dict) or not _exact_echa_match(item, cas_values, ec_values, normalized_names):
                continue
            substance_name = next((str(value).strip() for value in item.get("substanceName") or [] if str(value).strip()), name)
            reasons = [str(value).strip() for value in item.get("reasonForInclusion") or [] if str(value).strip()]
            decision_urls = [
                str(decision.get("decisionPath"))
                for decision in item.get("decision") or []
                if isinstance(decision, dict) and str(decision.get("decisionPath") or "").startswith("https://chem.echa.europa.eu/")
            ]
            check = {
                "status": "listed",
                "source": "ECHA Candidate List",
                "checked_at": checked_at,
                "details": "; ".join(filter(None, [str(item.get("dateOfInclusion") or "").strip(), "; ".join(reasons)])),
            }
            return {"svhc_identity": substance_name}, _unique([ECHA_CANDIDATE_PAGE, *decision_urls], limit=10), {"svhc": check}, {"echa_candidate_list": version}
    check = {"status": "not_listed", "source": "ECHA Candidate List", "checked_at": checked_at}
    return {}, [ECHA_CANDIDATE_PAGE], {"svhc": check}, {"echa_candidate_list": version}


def _normalise(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def _template_records(path: Path, kind: str) -> list[dict]:
    records: list[dict] = []
    if not path.exists():
        return records
    document = Document(path)
    for table in document.tables:
        for row in table.rows:
            texts = [re.sub(r"\s+", " ", cell.text).strip(" |") for cell in row.cells]
            joined = " | ".join(text for text in texts if text)
            if not joined or "name of ingredient" in joined.casefold() or joined.casefold().startswith("ingredient"):
                continue
            first = next((text for text in texts if text), "")
            name = re.split(r"\bCAS\s+(?:Number|No\.)\s*:", first, maxsplit=1, flags=re.IGNORECASE)[0].strip(" |")
            cases = {value for value in CAS_IN_TEXT.findall(joined) if valid_cas(value)}
            ecs = {value for value in EC_IN_TEXT.findall(joined) if not valid_cas(value)}
            record = {"name": name, "normalized_name": _normalise(name), "cas": cases, "ec": ecs}
            if kind == "ifra":
                standards = _unique(re.findall(r"\b(?:Restriction|Prohibition|Specification)\b", joined, flags=re.IGNORECASE), limit=3)
                years = re.findall(r"\b(?:19|20)\d{2}\b", joined)
                record.update({"standards": standards, "publication_year": years[-1] if years else ""})
            records.append(record)
    return records


def template_reference_match(path: Path | None, kind: str, item: dict) -> dict | None:
    if path is None:
        return None
    cases = {value for value in CAS_IN_TEXT.findall(str(item.get("cas") or "")) if valid_cas(value)}
    ecs = {value for value in EC_IN_TEXT.findall(str(item.get("ec") or "")) if not valid_cas(value)}
    names = {
        _normalise(value)
        for value in [item.get("name"), item.get("canonical_name"), *(item.get("aliases") or [])]
        if _normalise(value)
    }
    for record in _template_records(path, kind):
        if cases.intersection(record["cas"]) or ecs.intersection(record["ec"]) or record["normalized_name"] in names:
            return record
    return None
