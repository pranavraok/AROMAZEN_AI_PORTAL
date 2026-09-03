import io
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen.canvas import Canvas


TEMPLATE_DIR = Path(__file__).with_name("templates")
LOGO_PATH = Path(__file__).resolve().parents[4] / "aromazen-ai" / "public" / "AROMAZEN_AI_LOGO.png"
CASH_FLOW_TEMPLATE = TEMPLATE_DIR / "AROMAZEN_Monthly_Cash_Flow_Report_Same_Format_Template.xlsx"
ASSET_TEMPLATE = TEMPLATE_DIR / "AROMAZEN_Optional_Fixed_Assets_Same_Format_Template.xlsx"

NAVY = colors.HexColor("#102535")
TEAL = colors.HexColor("#17A6A1")
ORANGE = colors.HexColor("#F59E52")
GREEN = colors.HexColor("#2EAE78")
MUTED = colors.HexColor("#667785")
SOFT = colors.HexColor("#EDF3F5")
WHITE = colors.white


@dataclass
class BankSummary:
    name: str
    opening: float | None
    closing: float | None
    balance_type: str | None


def number(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            pass
    return 0.0


def indian(value: float) -> str:
    rounded = int(round(abs(value)))
    digits = str(rounded)
    if len(digits) > 3:
        digits = digits[:-3][::-1]
        digits = ",".join(digits[i:i + 2] for i in range(0, len(digits), 2))[::-1] + "," + str(rounded)[-3:]
    return ("-" if value < 0 else "") + "₹" + digits


def _sheet(workbook, name: str):
    for current in workbook.sheetnames:
        if current.strip().upper() == name:
            return workbook[current]
    raise ValueError(f"The required {name} sheet is missing.")


def _cell(sheet, reference: str) -> float:
    return number(sheet[reference].value)


def _normalized(value: object) -> str:
    text = re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()
    corrections = {"SUPPILER": "SUPPLIER", "TRASNPORT": "TRANSPORT", "MAINTAINANCE": "MAINTENANCE", "SECUIRTY": "SECURITY", "INSENTIVE": "INCENTIVE"}
    for wrong, right in corrections.items():
        text = text.replace(wrong, right)
    return re.sub(r"\s+", " ", text)


def _matches(value: object, aliases: tuple[str, ...]) -> bool:
    text = _normalized(value)
    return any(text == _normalized(alias) for alias in aliases)


SHEET_ANCHORS = {
    "MAIN": (("TELEPHONE",), ("PROVISIONS",), ("BANK CHARGES BOB", "BANK CHARGE BOB")),
    "TRANSPORT": (("TRANSPORT",), ("COURIER AGENT",), ("FREIGHT PAID",)),
    "SUPPLIER": (("RAW MATERIAL SUPPLIER",), ("R D RAW MATERIAL",), ("R D LAB",)),
    "PACKING": (("PACKING MATERIAL",), ("PACKAING", "PACKING")),
    "REPAIR": (("R D REPAIRS FITTINGS",), ("REPAIR MAINTENANCE",)),
    "CUSTOMER": (("CUSTOMER RECEIPT",), ("SCRAP DELAR", "SCRAP DEALER"), ("CUSTOMER",)),
}

SECTION_ALIASES = (
    ("CUSTOMER RECEIPT",), ("CUSTOMER",), ("SCRAP DELAR", "SCRAP DEALER"),
    ("RAW MATERIAL SUPPLIER",), ("R D RAW MATERIAL",), ("R D LAB",),
    ("TELEPHONE",), ("REED PRODUCTION",), ("COMMISSION",), ("LOAN PAID",),
    ("REFRESHMENT",), ("MARKETING EXPENSES", "MARKETING EXP"),
    ("SAFETY GLOVES SHOES",), ("VEHICLE MAINTENANCE",), ("MADHAVAKRUPA", "MADHAVAKRIPA"),
    ("LICENSE AND RENEWAL",), ("INSURANCE RENEWALS",), ("TRAVELLING EXP", "TRAVEL EXP"),
    ("SHIVAMOGGA",), ("SHIVAMOGGA FUEL",), ("ELECTRICAL FITTING SHIVAMOGGA",),
    ("SECURITY",), ("DESIGN LAB",), ("FUELS EXP", "FUEL EXP"),
    ("OFFICE EXPENSES", "OFFICE EXP"), ("PRINTING",), ("LOAN STAMP DUTY",),
    ("STAFF WELFARE",), ("SALARY DELAYED",), ("PROVISIONS",),
    ("BANK CHARGES INDUSIND",), ("BANK CHARGES BOB",), ("BANK CHARGE AXIS",),
)


def _all_text_cells(sheet):
    return [cell for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.strip()]


def _locate_sheets(workbook) -> dict[str, object]:
    located = {}
    used = set()
    for kind, anchors in SHEET_ANCHORS.items():
        scores = []
        for sheet in workbook.worksheets:
            cells = _all_text_cells(sheet)
            score = sum(any(_matches(cell.value, aliases) for cell in cells) for aliases in anchors)
            scores.append((score, sheet))
        best = max(score for score, _ in scores)
        choices = [sheet for score, sheet in scores if score == best and score > 0]
        minimum = 2 if kind not in {"PACKING", "REPAIR"} else 1
        if best < minimum or len(choices) != 1 or choices[0].title in used:
            found = ", ".join(f"{sheet.title} ({score}/{len(anchors)} headings)" for score, sheet in sorted(scores, key=lambda item: item[0], reverse=True)[:3])
            raise ValueError(f"Could not safely identify the {kind.title()} sheet. Keep at least {minimum} recognizable headings. Best matches: {found}.")
        located[kind] = choices[0]; used.add(choices[0].title)
    return located


def _heading_cells(sheet, aliases: tuple[str, ...]):
    return [cell for cell in _all_text_cells(sheet) if _matches(cell.value, aliases)]


def _next_heading_row(sheet, start_cell) -> int:
    rows = []
    for aliases in SECTION_ALIASES:
        for cell in _heading_cells(sheet, aliases):
            if cell.column == start_cell.column and cell.row > start_cell.row:
                rows.append(cell.row)
    return min(rows, default=sheet.max_row + 1)


def _section_from_cell(sheet, heading, category: str) -> float:
    value_column = heading.column + 1
    end_row = _next_heading_row(sheet, heading)
    details = []
    totals = []
    start_amount = number(sheet.cell(heading.row, value_column).value)
    if start_amount:
        details.append(start_amount)
    for row in range(heading.row + 1, end_row):
        label = sheet.cell(row, heading.column).value
        amount = number(sheet.cell(row, value_column).value)
        if not amount:
            continue
        if isinstance(label, str) and label.strip():
            details.append(amount)
        elif details:
            totals.append((row, amount, sum(details)))
    if totals:
        row, stated, calculated = totals[-1]
        tolerance = max(1.0, abs(calculated) * 0.0001)
        if abs(stated - calculated) > tolerance:
            raise ValueError(f"{category} does not reconcile on sheet '{sheet.title}': visible entries total {indian(calculated)}, but row {row} shows {indian(stated)}. Correct or remove that total.")
        return stated
    if details:
        return sum(details)
    return 0.0


def _section_total(sheet, aliases: tuple[str, ...], category: str, required: bool = True) -> float:
    headings = _heading_cells(sheet, aliases)
    if not headings:
        if required:
            raise ValueError(f"'{category}' heading was not found on sheet '{sheet.title}'. Restore that heading or use the downloadable template.")
        return 0.0
    return sum(_section_from_cell(sheet, cell, category) for cell in headings)


def _labeled_total(sheet, aliases: tuple[str, ...], category: str, required: bool = False) -> float:
    matches = _heading_cells(sheet, aliases)
    if required and not matches:
        raise ValueError(f"'{category}' heading was not found on sheet '{sheet.title}'. Restore that heading or use the downloadable template.")
    values = []
    for cell in matches:
        for column in range(cell.column + 1, min(sheet.max_column, cell.column + 3) + 1):
            amount = number(sheet.cell(cell.row, column).value)
            if amount:
                values.append(amount); break
    return sum(values)


def _sum_labeled_details(sheet, header_aliases: tuple[tuple[str, ...], ...], category: str) -> float:
    header_cells = [cell for aliases in header_aliases for cell in _heading_cells(sheet, aliases)]
    if not header_cells:
        raise ValueError(f"No recognizable {category} heading was found on sheet '{sheet.title}'.")
    label_column = min(cell.column for cell in header_cells)
    value_column = label_column + 1
    total = 0.0
    for row in range(1, sheet.max_row + 1):
        label = sheet.cell(row, label_column).value
        amount = number(sheet.cell(row, value_column).value)
        if not isinstance(label, str) or not label.strip() or not amount:
            continue
        if any(_matches(label, aliases) for aliases in header_aliases):
            continue
        total += amount
    if not total:
        raise ValueError(f"No numeric {category} entries were found on sheet '{sheet.title}'.")
    return total


def read_cash_flow(content: bytes) -> tuple[list[tuple[str, float]], list[tuple[str, float]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError("The monthly cash-flow file could not be read as an Excel workbook.") from exc
    if len(workbook.worksheets) < 6:
        raise ValueError(f"The workbook contains {len(workbook.worksheets)} sheets; six recognizable cash-flow sheets are required.")
    sheets = _locate_sheets(workbook)
    main = sheets["MAIN"]; customer = sheets["CUSTOMER"]; supplier = sheets["SUPPLIER"]
    transport = sheets["TRANSPORT"]; packing = sheets["PACKING"]; repair = sheets["REPAIR"]
    receipts = [
        ("Export collections", _section_total(customer, ("CUSTOMER",), "Export collections")),
        ("Domestic collections", _section_total(customer, ("CUSTOMER RECEIPT",), "Domestic collections")),
        ("GST refund", _labeled_total(main, ("GST REFUND",), "GST refund", required=True)),
        ("Customs duty drawback", _labeled_total(main, ("CUSTOMS DUTY DRAWBACK",), "Customs duty drawback", required=True)),
        ("Scrap receipts", _section_total(customer, ("SCRAP DELAR", "SCRAP DEALER"), "Scrap receipts")),
    ]
    payments = [
        ("Raw materials", _section_total(supplier, ("RAW MATERIAL SUPPLIER",), "Raw materials")),
        ("R&D raw materials", _section_total(supplier, ("R D RAW MATERIAL",), "R&D raw materials")),
        ("R&D laboratory", _section_total(supplier, ("R D LAB",), "R&D laboratory")),
        ("Transport", _sum_labeled_details(transport, (("TRANSPORT",),("TRANSPORT RAW MATERIAL IN",),("TRANSPORT FOR CUSTOMER",),("TEMPO CHARGES",),("COURIER AGENT",),("FREIGHT PAID",)), "transport")),
        ("Packing", _sum_labeled_details(packing, (("PACKING MATERIAL",),("PACKAING", "PACKING")), "packing")),
        ("Repairs & maintenance", _sum_labeled_details(repair, (("R D REPAIRS FITTINGS",),("REPAIR MAINTENANCE",)), "repair")),
        ("Telephone", _section_total(main, ("TELEPHONE",), "Telephone")),
        ("Reed production", _section_total(main, ("REED PRODUCTION",), "Reed production")),
        ("Commission", _section_total(main, ("COMMISSION",), "Commission")),
        ("Loans paid", _section_total(main, ("LOAN PAID",), "Loans paid") + _labeled_total(main, ("LOAN RECOVERY EMI",), "Loan recovery")),
        ("Refreshment", _section_total(main, ("REFRESHMENT",), "Refreshment")),
        ("Marketing", _section_total(main, ("MARKETING EXPENSES", "MARKETING EXP"), "Marketing")),
        ("Safety", _section_total(main, ("SAFETY GLOVES SHOES",), "Safety")),
        ("Vehicle maintenance", _section_total(main, ("VEHICLE MAINTENANCE",), "Vehicle maintenance")),
        ("Madhavakrupa", _section_total(main, ("MADHAVAKRUPA", "MADHAVAKRIPA"), "Madhavakrupa")),
        ("Licences & renewals", _section_total(main, ("LICENSE AND RENEWAL",), "Licences and renewals")),
        ("Insurance & renewals", _section_total(main, ("INSURANCE RENEWALS",), "Insurance renewals")),
        ("Travel", _section_total(main, ("TRAVELLING EXP", "TRAVEL EXP"), "Travel")),
        ("Shivamogga operations", _section_total(main, ("SHIVAMOGGA",), "Shivamogga operations") + _section_total(main, ("SHIVAMOGGA FUEL",), "Shivamogga fuel") + _section_total(main, ("ELECTRICAL FITTING SHIVAMOGGA",), "Shivamogga electrical")),
        ("Security", _section_total(main, ("SECURITY",), "Security")),
        ("Design lab", _section_total(main, ("DESIGN LAB",), "Design lab")),
        ("Fuel", _section_total(main, ("FUELS EXP", "FUEL EXP"), "Fuel")),
        ("Office & printing", _section_total(main, ("OFFICE EXPENSES", "OFFICE EXP"), "Office expenses") + _section_total(main, ("PRINTING",), "Printing")),
        ("Loan stamp duty", _section_total(main, ("LOAN STAMP DUTY",), "Loan stamp duty")),
        ("Staff welfare", _section_total(main, ("STAFF WELFARE",), "Staff welfare")),
        ("Delayed salary", _section_total(main, ("SALARY DELAYED",), "Delayed salary")),
        ("Salary", _labeled_total(main, ("SALARY",), "Salary")),
        ("Incentive", _labeled_total(main, ("INCENTIVE",), "Incentive")),
        ("GST tax", _labeled_total(main, ("GST TAX",), "GST tax")),
        ("ESI", _labeled_total(main, ("ESI",), "ESI")),
        ("EPFO", _labeled_total(main, ("EPFO",), "EPFO")),
        ("TDS", _labeled_total(main, ("TDS",), "TDS")),
        ("Professional tax", _labeled_total(main, ("PT",), "Professional tax")),
        ("Electricity", _labeled_total(main, ("MESCOM", "MESCOM BILL SHIVAMOGGA"), "Electricity")),
        ("Bank charges", _section_total(main, ("BANK CHARGES INDUSIND",), "IndusInd bank charges") + _section_total(main, ("BANK CHARGES BOB",), "BOB bank charges") + _section_total(main, ("BANK CHARGE AXIS",), "Axis bank charges")),
        ("Bank interest", _labeled_total(main, ("BANK INTEREST SIDBI",), "Bank interest")),
    ]
    receipts = [(label, value) for label, value in receipts if value]
    payments = [(label, value) for label, value in payments if value]
    if not receipts or not payments or sum(value for _, value in receipts) <= 0 or sum(value for _, value in payments) <= 0:
        raise ValueError("The cash-flow workbook does not contain usable monthly receipt and payment values.")
    return receipts, sorted(payments, key=lambda item: item[1], reverse=True)


def read_assets(content: bytes | None) -> list[tuple[str, float]]:
    if not content:
        return []
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError("The optional fixed-assets file could not be read as Excel.") from exc
    sheet = workbook.active
    rows = []
    in_fixed_assets = False
    for row in range(1, sheet.max_row + 1):
        name = sheet.cell(row, 4).value
        normalized = str(name or "").strip().lower()
        if normalized == "fixed assets":
            in_fixed_assets = True
            continue
        if in_fixed_assets and normalized == "current assets":
            break
        if not in_fixed_assets:
            continue
        value = number(sheet.cell(row, 5).value)
        if name and value and normalized not in {"particulars", "total"}:
            rows.append((str(name).strip(), value))
    return rows


def read_bank(name: str, content: bytes) -> BankSummary:
    try:
        reader = PdfReader(io.BytesIO(content))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as exc:
        raise ValueError(f"The {name} statement is unreadable. Upload an unlocked, text-readable bank PDF.") from exc
    if not text.strip():
        raise ValueError(f"The {name} statement has no readable text.")
    opening = re.search(r"opening\s+balance\D{0,30}([\d,]+\.\d{2})\s*(Cr|Dr)?", text, re.I)
    explicit = re.search(r"closing\s+balance\D{0,30}([\d,]+\.\d{2})\s*(Cr|Dr)?", text, re.I)
    matches = re.findall(r"([\d,]+\.\d{2})\s*(Cr|Dr)", text, re.I)
    match = (explicit.group(1), explicit.group(2) or "Cr") if explicit else (matches[0] if matches else None)
    opening_value = number(opening.group(1)) if opening else None
    if opening and (opening.group(2) or "Cr").lower() == "dr": opening_value = -opening_value
    return BankSummary(name, opening_value, number(match[0]) if match else None, match[1].title() if match else None)


def _text(canvas, x, y, value, size=9, color=NAVY, font="Helvetica", align="left"):
    canvas.setFillColor(color); canvas.setFont(font, size)
    if align == "right": canvas.drawRightString(x, y, str(value))
    elif align == "center": canvas.drawCentredString(x, y, str(value))
    else: canvas.drawString(x, y, str(value))


def _header(canvas, title: str, subtitle: str, page: int):
    width, height = landscape(A4)
    canvas.setFillColor(NAVY); canvas.rect(0, height - 76, width, 76, fill=1, stroke=0)
    if LOGO_PATH.exists():
        canvas.drawImage(ImageReader(LOGO_PATH), 26, height - 61, width=38, height=38, preserveAspectRatio=True, mask="auto")
    _text(canvas, width / 2, height - 34, title, 17, WHITE, "Helvetica-Bold", "center")
    _text(canvas, width / 2, height - 53, subtitle, 8, colors.HexColor("#BFD5DE"), align="center")
    _text(canvas, width - 30, 22, f"AROMAZEN  •  {page}", 7, MUTED, align="right")


def _card(canvas, x, y, w, label, value, accent):
    canvas.setFillColor(WHITE); canvas.roundRect(x, y, w, 78, 10, fill=1, stroke=0)
    canvas.setStrokeColor(colors.HexColor("#DCE6EA")); canvas.roundRect(x, y, w, 78, 10, fill=0, stroke=1)
    canvas.setFillColor(accent); canvas.roundRect(x, y, 5, 78, 3, fill=1, stroke=0)
    _text(canvas, x + 17, y + 51, label.upper(), 7, MUTED, "Helvetica-Bold")
    _text(canvas, x + 17, y + 22, value, 15, NAVY, "Helvetica-Bold")


def _bars(canvas, rows, x, y, w, row_height=28):
    maximum = max((value for _, value in rows), default=1)
    for index, (label, value) in enumerate(rows):
        yy = y - index * row_height
        _text(canvas, x, yy + 7, label[:38], 8)
        canvas.setFillColor(SOFT); canvas.roundRect(x + 205, yy, w - 330, 12, 6, fill=1, stroke=0)
        canvas.setFillColor(TEAL if index % 2 == 0 else ORANGE); canvas.roundRect(x + 205, yy, max(3, (w - 330) * value / maximum), 12, 6, fill=1, stroke=0)
        _text(canvas, x + w, yy + 3, indian(value), 8, NAVY, "Helvetica-Bold", "right")


def _percentage_change(current: float, previous: float) -> str:
    if previous == 0:
        return "New" if current else "No change"
    change = (current - previous) / abs(previous) * 100
    return f"{change:+.1f}%"


def _comparison_page(canvas, month_label: str, previous: dict, receipts, payments, page: int):
    width, height = landscape(A4)
    previous_label = datetime.strptime(previous["report_month"], "%Y-%m").strftime("%B %Y")
    current_receipts = sum(value for _, value in receipts)
    current_payments = sum(value for _, value in payments)
    current_net = current_receipts - current_payments
    previous_receipts = float(previous["total_receipts"])
    previous_payments = float(previous["total_payments"])
    previous_net = float(previous["net_movement"])

    _header(canvas, "MONTH-ON-MONTH CASH FLOW COMPARISON", f"{previous_label} vs {month_label}", page)
    metrics = (
        ("Cash receipts", previous_receipts, current_receipts, TEAL),
        ("Cash payments", previous_payments, current_payments, ORANGE),
        ("Net movement", previous_net, current_net, GREEN if current_net >= 0 else ORANGE),
    )
    for index, (label, old, current, accent) in enumerate(metrics):
        x = 30 + index * 273
        canvas.setFillColor(WHITE); canvas.roundRect(x, height - 205, 245, 105, 10, fill=1, stroke=0)
        canvas.setStrokeColor(colors.HexColor("#DCE6EA")); canvas.roundRect(x, height - 205, 245, 105, 10, fill=0, stroke=1)
        _text(canvas, x + 16, height - 124, label.upper(), 8, accent, "Helvetica-Bold")
        _text(canvas, x + 16, height - 151, previous_label[:3].upper(), 7, MUTED, "Helvetica-Bold")
        _text(canvas, x + 64, height - 151, indian(old), 10, NAVY, "Helvetica-Bold")
        _text(canvas, x + 16, height - 181, month_label[:3].upper(), 7, MUTED, "Helvetica-Bold")
        _text(canvas, x + 64, height - 181, indian(current), 10, NAVY, "Helvetica-Bold")
        _text(canvas, x + 229, height - 181, _percentage_change(current, old), 9, accent, "Helvetica-Bold", "right")

    _text(canvas, 38, height - 243, "PAYMENT CATEGORY MOVEMENT", 9, NAVY, "Helvetica-Bold")
    old_map = {str(label): float(value) for label, value in previous.get("payments", [])}
    current_map = {str(label): float(value) for label, value in payments}
    changes = sorted(
        ((label, old_map.get(label, 0.0), current_map.get(label, 0.0)) for label in set(old_map) | set(current_map)),
        key=lambda row: abs(row[2] - row[1]), reverse=True,
    )[:7]
    maximum = max((max(old, current) for _, old, current in changes), default=1) or 1
    for index, (label, old, current) in enumerate(changes):
        y = height - 278 - index * 38
        _text(canvas, 42, y + 7, label[:30], 8)
        canvas.setFillColor(colors.HexColor("#D8E1E5")); canvas.roundRect(235, y + 12, 310 * old / maximum, 7, 3, fill=1, stroke=0)
        canvas.setFillColor(TEAL); canvas.roundRect(235, y, 310 * current / maximum, 7, 3, fill=1, stroke=0)
        _text(canvas, 565, y + 9, indian(old), 7, MUTED, align="right")
        _text(canvas, 680, y + 9, indian(current), 7, NAVY, "Helvetica-Bold", "right")
        _text(canvas, 795, y + 9, indian(current - old), 7, GREEN if current - old <= 0 else ORANGE, "Helvetica-Bold", "right")
    _text(canvas, 235, 42, previous_label, 7, MUTED)
    _text(canvas, 480, 42, month_label, 7, TEAL, "Helvetica-Bold")
    _text(canvas, 745, 42, "CHANGE", 7, MUTED, "Helvetica-Bold")


def _insights_page(canvas, month_label: str, receipts, payments, banks, previous: dict | None, page: int):
    width, height = landscape(A4)
    inflow = sum(value for _, value in receipts)
    outflow = sum(value for _, value in payments)
    net = inflow - outflow
    top_receipt = max(receipts, key=lambda item: item[1])
    top_payment = max(payments, key=lambda item: item[1])
    _header(canvas, "AI CASH-FLOW INSIGHTS", "Factual observations from the report data", page)

    observations = []
    if previous:
        previous_label = datetime.strptime(previous["report_month"], "%Y-%m").strftime("%B %Y")
        previous_receipts = float(previous["total_receipts"])
        previous_payments = float(previous["total_payments"])
        previous_net = float(previous["net_movement"])
        observations.extend((
            ("RECEIPTS", indian(inflow), f"{_percentage_change(inflow, previous_receipts)} vs {previous_label}"),
            ("PAYMENTS", indian(outflow), f"{_percentage_change(outflow, previous_payments)} vs {previous_label}"),
            ("NET MOVEMENT", indian(net), f"{indian(net - previous_net)} change vs {previous_label}"),
        ))
        old_payments = {str(label): float(value) for label, value in previous.get("payments", [])}
        observations.append(("LARGEST PAYMENT", top_payment[0], f"{indian(top_payment[1])} | {indian(top_payment[1] - old_payments.get(top_payment[0], 0))} change"))
    else:
        observations.extend((
            ("CASH COVERAGE", f"{(inflow / outflow * 100):.1f}%" if outflow else "N/A", "Receipts as a share of payments"),
            ("NET MOVEMENT", indian(net), "Receipts less payments"),
            ("LARGEST RECEIPT", top_receipt[0], f"{indian(top_receipt[1])} | {top_receipt[1] / inflow * 100:.1f}% of receipts"),
            ("LARGEST PAYMENT", top_payment[0], f"{indian(top_payment[1])} | {top_payment[1] / outflow * 100:.1f}% of payments"),
        ))

    for index, (label, value, detail) in enumerate(observations):
        y = height - 176 - index * 92
        canvas.setFillColor(WHITE); canvas.roundRect(55, y, width - 110, 68, 10, fill=1, stroke=0)
        canvas.setStrokeColor(colors.HexColor("#DCE6EA")); canvas.roundRect(55, y, width - 110, 68, 10, fill=0, stroke=1)
        canvas.setFillColor((TEAL, ORANGE, GREEN, NAVY)[index]); canvas.roundRect(55, y, 6, 68, 3, fill=1, stroke=0)
        _text(canvas, 78, y + 43, label, 7, MUTED, "Helvetica-Bold")
        _text(canvas, 78, y + 17, str(value)[:42], 14, NAVY, "Helvetica-Bold")
        _text(canvas, width - 78, y + 25, detail, 9, MUTED, align="right")


def build_report(month: str, receipts, payments, banks, assets, password: str, previous: dict | None = None) -> bytes:
    try:
        month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    except ValueError as exc:
        raise ValueError("Choose a valid report month.") from exc
    raw = io.BytesIO(); canvas = Canvas(raw, pagesize=landscape(A4)); width, height = landscape(A4)
    inflow = sum(v for _, v in receipts); outflow = sum(v for _, v in payments); net = inflow - outflow
    known_openings = [b.opening for b in banks if b.opening is not None]
    known_closings = [(-b.closing if b.balance_type == "Dr" else b.closing) for b in banks if b.closing is not None]
    _header(canvas, "MONTHLY CASH FLOW OVERVIEW", month_label, 1)
    _text(canvas, width / 2, height - 109, "OWNER DASHBOARD", 8, TEAL, "Helvetica-Bold", "center")
    gap = 14; card_w = (width - 60 - gap * 3) / 4
    cards = [("Opening bank position", indian(sum(known_openings)) if len(known_openings) == 3 else "See bank page", NAVY), ("Cash receipts", indian(inflow), TEAL), ("Cash payments", indian(outflow), ORANGE), ("Closing bank position", indian(sum(known_closings)) if len(known_closings) == 3 else "See bank page", GREEN)]
    for i, item in enumerate(cards): _card(canvas, 30 + i * (card_w + gap), height - 215, card_w, *item)
    _text(canvas, 40, height - 258, "MONTH AT A GLANCE", 9, NAVY, "Helvetica-Bold")
    _bars(canvas, (receipts + payments)[:9], 40, height - 292, width - 80)
    canvas.showPage()

    _header(canvas, "ACTUAL CASH FLOW STATEMENT", "Direct method • monthly receipts and payments", 2)
    _card(canvas, 30, height - 180, 235, "Total receipts", indian(inflow), TEAL); _card(canvas, 303, height - 180, 235, "Total payments", indian(outflow), ORANGE); _card(canvas, 576, height - 180, 235, "Net movement", indian(net), GREEN if net >= 0 else ORANGE)
    _text(canvas, 45, height - 230, "RECEIPTS", 9, TEAL, "Helvetica-Bold"); _text(canvas, 430, height - 230, "PAYMENTS", 9, ORANGE, "Helvetica-Bold")
    for i,(label,value) in enumerate(receipts): _text(canvas,45,height-260-i*27,label,8); _text(canvas,360,height-260-i*27,indian(value),8,NAVY,"Helvetica-Bold","right")
    for i,(label,value) in enumerate(payments[:11]): _text(canvas,430,height-260-i*27,label,8); _text(canvas,795,height-260-i*27,indian(value),8,NAVY,"Helvetica-Bold","right")
    canvas.showPage()

    page = 3
    if previous:
        _comparison_page(canvas, month_label, previous, receipts, payments, page); canvas.showPage(); page += 1

    _header(canvas, "CASH RECEIPTS", "Complete visual breakdown", page); _bars(canvas, receipts, 45, height - 125, width - 90, 48); canvas.showPage(); page += 1
    for start in range(0, len(payments), 14):
        _header(canvas, "CASH PAYMENTS", f"Complete detail • items {start + 1}–{min(start + 14, len(payments))}", page)
        _bars(canvas, payments[start:start + 14], 45, height - 115, width - 90, 31); canvas.showPage(); page += 1
    _header(canvas, "BANK-WISE CLOSING POSITION", "Balances read from the three uploaded official statements", page)
    for i, bank in enumerate(banks):
        value = f"{indian(bank.closing)} {bank.balance_type or ''}" if bank.closing is not None else "Not stated in PDF"
        _card(canvas, 45 + i * 260, height - 205, 235, bank.name, value, (TEAL, ORANGE, GREEN)[i])
    canvas.showPage(); page += 1
    if assets:
        for start in range(0, len(assets), 40):
            _header(canvas, "FIXED-ASSET REGISTER", f"Optional register • items {start + 1}–{min(start + 40, len(assets))}", page)
            rows = assets[start:start + 40]; columns = [rows[:20], rows[20:]]
            for col, values in enumerate(columns):
                x = 42 + col * 395
                for i,(label,value) in enumerate(values):
                    y = height - 115 - i * 22; _text(canvas,x,y,label[:45],7.5); _text(canvas,x+355,y,indian(value),7.5,NAVY,"Helvetica-Bold","right")
            canvas.showPage(); page += 1
    _insights_page(canvas, month_label, receipts, payments, banks, previous, page); canvas.showPage()
    canvas.save(); raw.seek(0)
    reader = PdfReader(raw); writer = PdfWriter(); [writer.add_page(p) for p in reader.pages]
    writer.encrypt(password, algorithm="AES-256"); protected = io.BytesIO(); writer.write(protected)
    return protected.getvalue()
